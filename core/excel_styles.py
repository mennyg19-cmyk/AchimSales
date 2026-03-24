"""
Shared Excel styling constants.

Single source of truth for all fills, fonts, borders, and number formats
used across report writers. Prevents drift between reports.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- Borders ---
THIN_SIDE = Side(style="thin", color="000000")

BORDER_THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# --- Fonts ---
FONT_HEADER = Font(bold=True)
FONT_TITLE = Font(bold=True, size=16)
FONT_LINK = Font(color="0563C1", underline="single")

# --- Fills ---
FILL_HEADER = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
FILL_HEADER_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_SUMMARY_HEADER = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
FILL_LIGHT_GREY = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
FILL_TOTAL_ROW = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
FILL_TOTALS = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# --- Number formats ---
FMT_CURRENCY = '"$"#,##0.00'
FMT_DATE = "M/D/YYYY"

# --- Alignment ---
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def score_to_fill(score: float) -> PatternFill:
    """Fulfillment score gradient: red (0) -> yellow (0.5) -> green (1)."""
    try:
        s = float(score)
        if s < 0 or s != s:
            return FILL_LIGHT_GREY
    except (TypeError, ValueError):
        return FILL_LIGHT_GREY

    s = max(0.0, min(1.0, s))
    red = (255, 199, 206)
    yellow = (255, 235, 156)
    green = (198, 239, 206)

    if s <= 0:
        r, g, b = red
    elif s >= 1:
        r, g, b = green
    elif s < 0.5:
        t = s * 2
        r = int(red[0] + (yellow[0] - red[0]) * t)
        g = int(red[1] + (yellow[1] - red[1]) * t)
        b = int(red[2] + (yellow[2] - red[2]) * t)
    else:
        t = (s - 0.5) * 2
        r = int(yellow[0] + (green[0] - yellow[0]) * t)
        g = int(yellow[1] + (green[1] - yellow[1]) * t)
        b = int(yellow[2] + (green[2] - yellow[2]) * t)

    hex_color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
