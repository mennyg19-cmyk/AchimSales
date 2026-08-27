"""
Shared Excel writing utilities.

Provides autosize, currency formatting, streaming cell helpers,
and timezone stripping for Excel-safe DataFrames.
"""

import logging
import math

import pandas as pd
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from core.excel_styles import FMT_CURRENCY

log = logging.getLogger(__name__)

STREAMING_ROW_THRESHOLD = 2000
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n")


def neutralize_excel_value(value):
    """Prefix formula-leading strings so Excel treats them as text."""
    if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def autosize_columns(ws, max_width: int = 60, min_width: int = 10, max_rows: int | None = None) -> None:
    """Auto-size columns based on content width.

    When ``max_rows`` is set, only samples that many rows to determine widths
    (much faster and lighter on memory for large sheets).
    """
    widths: dict[int, int] = {}
    row_limit = max_rows or ws.max_row
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= row_limit:
            break
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            widths[i] = max(widths.get(i, 0), min(len(str(v)), max_width))
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_width, w + 2)


def format_money_columns(ws, money_keywords: set[str] | None = None, header_row: int = 1) -> None:
    """Apply currency format to columns whose headers match money keywords."""
    if money_keywords is None:
        money_keywords = {
            "subtotal invoices", "tariff charges", "freight charges", "cc charges",
            "total invoice", "sales", "freight", "cc", "tariff", "commissions",
            "total invoices", "totnofrt",
        }

    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            headers[str(v).strip().lower()] = c

    for key, col in headers.items():
        if "count" in key:
            continue
        if key in money_keywords or any(kw in key for kw in ("charge", "invoice", "sales", "tot")):
            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)) and not math.isnan(cell.value):
                    cell.number_format = FMT_CURRENCY


def make_streaming_cell(ws, value, fill=None, font=None, border=None, fmt=None) -> WriteOnlyCell:
    """Create a WriteOnlyCell with optional styling."""
    cell = WriteOnlyCell(ws, value=neutralize_excel_value(value))
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell


def strip_datetime_tz(df: pd.DataFrame) -> pd.DataFrame:
    """Convert tz-aware datetime columns to naive for Excel compatibility."""
    for col in df.columns:
        try:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                if hasattr(df[col].dtype, "tz") and df[col].dtype.tz is not None:
                    df[col] = df[col].dt.tz_localize(None)
        except Exception:
            log.debug("Could not strip timezone from column %s", col, exc_info=True)
    for col in df.columns:
        if df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            df[col] = df[col].map(neutralize_excel_value)
    return df
