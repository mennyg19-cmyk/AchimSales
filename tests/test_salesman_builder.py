"""Unit tests for the Salesman Report builder and writer logic."""

import tempfile
from pathlib import Path

import pandas as pd
import pytest

from reports.salesman.builder import build_salesman_full_year_data
from reports.salesman.writer import write_monthly_salesmen_workbook


class TestBuildSalesmanFullYearData:
    """Tests for ``build_salesman_full_year_data()``."""

    def test_returns_12_months(self, sample_invoice_detail):
        result = build_salesman_full_year_data(sample_invoice_detail, year=2026)
        assert len(result) == 12
        for m in range(1, 13):
            assert m in result

    def test_month_has_expected_columns(self, sample_invoice_detail):
        result = build_salesman_full_year_data(sample_invoice_detail, year=2026)
        feb = result[2]
        if not feb.empty:
            for col in ["Sales_Current", "Sales_Prior", "Sales_YTD_Current", "Sales_YTD_Prior", "Salesman", "CustomerAccount"]:
                assert col in feb.columns, f"Expected column '{col}' missing"

    def test_empty_input(self):
        result = build_salesman_full_year_data(pd.DataFrame(), year=2026)
        assert result == {}


class TestWriteMonthlySalesmenWorkbook:
    """Tests for ``write_monthly_salesmen_workbook()``."""

    def test_writes_12_sheets(self, sample_invoice_detail):
        month_data = build_salesman_full_year_data(sample_invoice_detail, year=2026)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name
        try:
            write_monthly_salesmen_workbook(month_data, 2026, out_path)
            from openpyxl import load_workbook
            wb = load_workbook(out_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            assert len(sheet_names) == 12
            expected = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            assert sheet_names == expected
        finally:
            Path(out_path).unlink(missing_ok=True)
