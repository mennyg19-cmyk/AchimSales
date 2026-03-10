"""
Cell-by-cell Excel comparison tool.

Compares two Excel workbooks sheet-by-sheet, cell-by-cell.
Reports differences in values, formatting, and structure.

Usage:
  python -m tests.compare_reports old_report.xlsx new_report.xlsx
  python -m tests.compare_reports old_report.xlsx new_report.xlsx --tolerance 0.01
"""

import argparse
import math
import sys
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class ComparisonResult:
    """Result of comparing two workbooks."""
    old_path: str
    new_path: str
    sheet_diffs: dict[str, list[str]] = field(default_factory=dict)
    missing_sheets_in_new: list[str] = field(default_factory=list)
    extra_sheets_in_new: list[str] = field(default_factory=list)
    total_cells_compared: int = 0
    total_diffs: int = 0

    @property
    def is_match(self) -> bool:
        return (
            self.total_diffs == 0
            and not self.missing_sheets_in_new
            and not self.extra_sheets_in_new
        )

    def summary(self) -> str:
        lines = [
            f"Comparison: {self.old_path} vs {self.new_path}",
            f"Cells compared: {self.total_cells_compared}",
            f"Differences: {self.total_diffs}",
        ]
        if self.missing_sheets_in_new:
            lines.append(f"Missing sheets in new: {', '.join(self.missing_sheets_in_new)}")
        if self.extra_sheets_in_new:
            lines.append(f"Extra sheets in new: {', '.join(self.extra_sheets_in_new)}")
        if self.is_match:
            lines.append("RESULT: MATCH")
        else:
            lines.append("RESULT: DIFFERENCES FOUND")
            for sheet, diffs in self.sheet_diffs.items():
                lines.append(f"\n  Sheet '{sheet}': {len(diffs)} difference(s)")
                for d in diffs[:20]:
                    lines.append(f"    {d}")
                if len(diffs) > 20:
                    lines.append(f"    ... and {len(diffs) - 20} more")
        return "\n".join(lines)


def compare_workbooks(
    old_path: str,
    new_path: str,
    tolerance: float = 0.001,
    ignore_formatting: bool = False,
    max_diffs_per_sheet: int = 100,
) -> ComparisonResult:
    """Compare two Excel workbooks cell-by-cell.

    Args:
        tolerance: numeric tolerance for float comparisons
        ignore_formatting: if True, only compare values (not fills/fonts/formats)
        max_diffs_per_sheet: stop collecting diffs after this many per sheet
    """
    result = ComparisonResult(old_path=old_path, new_path=new_path)

    wb_old = load_workbook(old_path, data_only=True)
    wb_new = load_workbook(new_path, data_only=True)

    old_sheets = set(wb_old.sheetnames)
    new_sheets = set(wb_new.sheetnames)

    result.missing_sheets_in_new = sorted(old_sheets - new_sheets)
    result.extra_sheets_in_new = sorted(new_sheets - old_sheets)

    common_sheets = sorted(old_sheets & new_sheets, key=lambda s: wb_old.sheetnames.index(s))

    for sheet_name in common_sheets:
        ws_old = wb_old[sheet_name]
        ws_new = wb_new[sheet_name]
        diffs = []

        max_row = max(ws_old.max_row or 1, ws_new.max_row or 1)
        max_col = max(ws_old.max_column or 1, ws_new.max_column or 1)

        if ws_old.max_row != ws_new.max_row:
            diffs.append(f"Row count: old={ws_old.max_row} new={ws_new.max_row}")
        if ws_old.max_column != ws_new.max_column:
            diffs.append(f"Column count: old={ws_old.max_column} new={ws_new.max_column}")

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if len(diffs) >= max_diffs_per_sheet:
                    break

                cell_old = ws_old.cell(row=r, column=c)
                cell_new = ws_new.cell(row=r, column=c)
                cell_ref = f"{get_column_letter(c)}{r}"

                result.total_cells_compared += 1

                v_old = cell_old.value
                v_new = cell_new.value

                if not _values_equal(v_old, v_new, tolerance):
                    diffs.append(f"{cell_ref}: value old={_repr(v_old)} new={_repr(v_new)}")
                    result.total_diffs += 1

                if not ignore_formatting:
                    fmt_diffs = _compare_formatting(cell_old, cell_new, cell_ref)
                    for fd in fmt_diffs:
                        if len(diffs) < max_diffs_per_sheet:
                            diffs.append(fd)
                            result.total_diffs += 1

        if diffs:
            result.sheet_diffs[sheet_name] = diffs

    wb_old.close()
    wb_new.close()
    return result


def _values_equal(v1, v2, tolerance: float) -> bool:
    """Compare two cell values with numeric tolerance."""
    if v1 is None and v2 is None:
        return True
    if v1 is None or v2 is None:
        v1_empty = v1 is None or (isinstance(v1, str) and v1.strip() == "")
        v2_empty = v2 is None or (isinstance(v2, str) and v2.strip() == "")
        return v1_empty and v2_empty

    if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
        if math.isnan(v1) and math.isnan(v2):
            return True
        return abs(v1 - v2) <= tolerance

    if isinstance(v1, str) and isinstance(v2, str):
        return v1.strip() == v2.strip()

    return str(v1).strip() == str(v2).strip()


def _compare_formatting(cell_old, cell_new, cell_ref: str) -> list[str]:
    """Compare cell formatting. Returns list of diff descriptions."""
    diffs = []

    if cell_old.number_format != cell_new.number_format:
        diffs.append(f"{cell_ref}: number_format old='{cell_old.number_format}' new='{cell_new.number_format}'")

    if cell_old.font and cell_new.font:
        if cell_old.font.bold != cell_new.font.bold:
            diffs.append(f"{cell_ref}: bold old={cell_old.font.bold} new={cell_new.font.bold}")

    if cell_old.fill and cell_new.fill:
        old_color = getattr(cell_old.fill, "start_color", None)
        new_color = getattr(cell_new.fill, "start_color", None)
        if old_color and new_color:
            old_rgb = getattr(old_color, "rgb", None)
            new_rgb = getattr(new_color, "rgb", None)
            if old_rgb != new_rgb and old_rgb and new_rgb:
                if old_rgb != "00000000" or new_rgb != "00000000":
                    diffs.append(f"{cell_ref}: fill old={old_rgb} new={new_rgb}")

    return diffs


def _repr(v) -> str:
    """Short repr for display."""
    if v is None:
        return "None"
    if isinstance(v, float):
        return f"{v:.4f}"
    s = str(v)
    return s[:50] + "..." if len(s) > 50 else s


def main():
    parser = argparse.ArgumentParser(description="Compare two Excel workbooks cell-by-cell")
    parser.add_argument("old", help="Path to old (baseline) workbook")
    parser.add_argument("new", help="Path to new (refactored) workbook")
    parser.add_argument("--tolerance", type=float, default=0.001, help="Numeric tolerance (default: 0.001)")
    parser.add_argument("--ignore-formatting", action="store_true", help="Only compare values, not formatting")
    parser.add_argument("--max-diffs", type=int, default=100, help="Max diffs per sheet (default: 100)")
    args = parser.parse_args()

    result = compare_workbooks(
        args.old, args.new,
        tolerance=args.tolerance,
        ignore_formatting=args.ignore_formatting,
        max_diffs_per_sheet=args.max_diffs,
    )
    print(result.summary())
    sys.exit(0 if result.is_match else 1)


if __name__ == "__main__":
    main()
