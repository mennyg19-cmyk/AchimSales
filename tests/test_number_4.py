"""
Tests for the Number 4 Report aggregator.

The Number 4 report pivots invoice lines into monthly qty/dollar columns.
"""

from datetime import datetime

import numpy as np
import pandas as pd
import pytest

from reports.number_4.aggregator import (
    aggregate_by_item_customer,
    build_month_labels,
    rolling_12_months,
    ytd_months,
)


def _make_invoice_lines(rows):
    """Build a minimal Number 4 invoice-lines DataFrame."""
    df = pd.DataFrame(rows)
    df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"])
    for col in ["Qty", "Total_$"]:
        if col not in df.columns:
            df[col] = 0.0
    for col in ["Item_#", "Item_Name", "CustomerAccount", "CustomerName", "Salesman", "InvoiceNumber"]:
        if col not in df.columns:
            df[col] = ""
    return df


class TestRolling12Months:

    def test_normal(self):
        rd = datetime(2026, 6, 15)
        months = rolling_12_months(rd)
        assert len(months) == 12
        assert months[0] == (2025, 7)
        assert months[-1] == (2026, 6)

    def test_january(self):
        """January wraps around to prior year."""
        rd = datetime(2026, 1, 10)
        months = rolling_12_months(rd)
        assert months[0] == (2025, 2)
        assert months[-1] == (2026, 1)

    def test_december(self):
        rd = datetime(2026, 12, 1)
        months = rolling_12_months(rd)
        assert months[0] == (2026, 1)
        assert months[-1] == (2026, 12)


class TestYtdMonths:

    def test_mid_year(self):
        rd = datetime(2026, 6, 15)
        months = ytd_months(rd)
        assert months == [(2026, m) for m in range(1, 7)]

    def test_january(self):
        rd = datetime(2026, 1, 1)
        months = ytd_months(rd)
        assert months == [(2026, 1)]


class TestAggregateByItemCustomer:

    def test_rolling_12_months_pivot(self):
        """18 months of data → only last 12 appear as columns."""
        rows = []
        for m in range(1, 19):
            y = 2025 if m <= 12 else 2026
            mm = m if m <= 12 else m - 12
            rows.append({
                "InvoiceDate": f"{y}-{mm:02d}-15",
                "Item_#": "ITEM-A", "Item_Name": "Widget A",
                "CustomerAccount": "C100", "CustomerName": "Customer C100",
                "Salesman": "SM01", "InvoiceNumber": f"INV-{m}",
                "Qty": 10.0, "Total_$": 100.0,
            })
        lines = _make_invoice_lines(rows)
        rd = datetime(2026, 6, 15)
        months = rolling_12_months(rd)
        agg, qty_keys, dol_keys = aggregate_by_item_customer(lines, months)

        assert len(qty_keys) == 12
        # All 12 month columns should have data
        for k in qty_keys:
            assert agg[k].sum() > 0

    def test_ytd_slice(self):
        """YTD aggregation only includes current year months."""
        rows = [
            {"InvoiceDate": "2025-11-15", "Item_#": "A", "Item_Name": "A",
             "CustomerAccount": "C1", "CustomerName": "C1", "Salesman": "S",
             "Qty": 5, "Total_$": 50},
            {"InvoiceDate": "2026-02-15", "Item_#": "A", "Item_Name": "A",
             "CustomerAccount": "C1", "CustomerName": "C1", "Salesman": "S",
             "Qty": 10, "Total_$": 100},
        ]
        lines = _make_invoice_lines(rows)
        # Only keep 2026 lines for YTD
        lines_ytd = lines[lines["InvoiceDate"].dt.year == 2026].copy()

        rd = datetime(2026, 6, 15)
        months = ytd_months(rd)
        agg, qty_keys, dol_keys = aggregate_by_item_customer(lines_ytd, months)

        assert agg["Total_Qty"].iloc[0] == pytest.approx(10.0)

    def test_empty_lines(self):
        """Empty lines → empty aggregation, no crash."""
        lines = pd.DataFrame(columns=[
            "InvoiceDate", "Item_#", "Item_Name", "CustomerAccount",
            "CustomerName", "Salesman", "Qty", "Total_$",
        ])
        lines["InvoiceDate"] = pd.to_datetime(lines["InvoiceDate"])
        months = [(2026, m) for m in range(1, 7)]
        agg, qty_keys, dol_keys = aggregate_by_item_customer(lines, months)
        assert agg.empty

    def test_avg_price(self):
        """Average price = Total_$ / Total_Qty."""
        rows = [
            {"InvoiceDate": "2026-03-15", "Item_#": "A", "Item_Name": "A",
             "CustomerAccount": "C1", "CustomerName": "C1", "Salesman": "S",
             "Qty": 10.0, "Total_$": 200.0},
            {"InvoiceDate": "2026-04-15", "Item_#": "A", "Item_Name": "A",
             "CustomerAccount": "C1", "CustomerName": "C1", "Salesman": "S",
             "Qty": 10.0, "Total_$": 300.0},
        ]
        lines = _make_invoice_lines(rows)
        months = [(2026, 3), (2026, 4)]
        agg, _, _ = aggregate_by_item_customer(lines, months)
        assert agg["Avg_Price"].iloc[0] == pytest.approx(25.0)

    def test_multiple_items_customers(self):
        """Different items/customers are grouped separately."""
        rows = [
            {"InvoiceDate": "2026-03-15", "Item_#": "A", "Item_Name": "A",
             "CustomerAccount": "C1", "CustomerName": "C1", "Salesman": "S",
             "Qty": 5, "Total_$": 50},
            {"InvoiceDate": "2026-03-15", "Item_#": "B", "Item_Name": "B",
             "CustomerAccount": "C2", "CustomerName": "C2", "Salesman": "S",
             "Qty": 10, "Total_$": 200},
        ]
        lines = _make_invoice_lines(rows)
        months = [(2026, 3)]
        agg, _, _ = aggregate_by_item_customer(lines, months)
        assert len(agg) == 2


class TestBuildMonthLabels:

    def test_labels_format(self):
        months = [(2026, 1), (2026, 2)]
        labels = build_month_labels(months)
        assert labels[0] == ("Jan-26 Qty", "Jan-26 $")
        assert labels[1] == ("Feb-26 Qty", "Feb-26 $")


class TestWriteByItemTrailingColumns:

    def test_headers_match_by_customer_trailing(self, tmp_path):
        from openpyxl import load_workbook

        from reports.number_4.writer_item import write_by_item

        agg = pd.DataFrame([{
            "Item_#": "A", "Item_Name": "Widget",
            "CustomerAccount": "100", "CustomerName": "Acme",
            "Salesman": "S", "2026-03": 5.0, "2026-03_$": 50.0,
            "Total_Qty": 5.0, "Total_$": 50.0, "Avg_Price": 10.0, "BookPrice": 12.0,
        }])
        labels = [("Mar-26 Qty", "Mar-26 $")]
        out = tmp_path / "n4_item.xlsx"
        write_by_item(agg, labels, ["2026-03"], ["2026-03_$"],
                      agg, labels, ["2026-03"], ["2026-03_$"], str(out))
        wb = load_workbook(out, read_only=True)
        try:
            headers = [c.value for c in next(wb["12 Months"].iter_rows(max_row=1))]
        finally:
            wb.close()
        assert headers == [
            "Item #", "Item Name", "Customer #", "Customer Name",
            "Mar-26 Qty", "Mar-26 $",
            "Total Qty", "Total $", "Avg Price", "Book Price", "Salesman",
        ]


class TestWriteByCustomerPricesBeforeSalesman:

    def test_headers_put_avg_and_book_price_before_salesman(self, tmp_path):
        from openpyxl import load_workbook

        from reports.number_4.writer_customer import write_by_customer

        agg = pd.DataFrame([{
            "Item_#": "A", "Item_Name": "Widget",
            "CustomerAccount": "100", "CustomerName": "Acme",
            "Salesman": "S", "2026-03": 5.0, "2026-03_$": 50.0,
            "Total_Qty": 5.0, "Total_$": 50.0, "Avg_Price": 10.0, "BookPrice": 12.0,
        }])
        labels = [("Mar-26 Qty", "Mar-26 $")]
        out = tmp_path / "n4_cust.xlsx"
        write_by_customer(agg, labels, ["2026-03"], ["2026-03_$"],
                          agg, labels, ["2026-03"], ["2026-03_$"], str(out))
        wb = load_workbook(out, read_only=True)
        try:
            headers = [c.value for c in next(wb["12 Months"].iter_rows(max_row=1))]
        finally:
            wb.close()
        assert headers[-5:] == ["Total Qty", "Total $", "Avg Price", "Book Price", "Salesman"]
