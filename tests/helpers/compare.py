"""
Excel workbook comparison and structural assertions.

Provides lightweight checks for sheet names, row counts, and cell values
so tests can verify generated reports without pixel-level screenshot comparison.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

log = logging.getLogger(__name__)


@dataclass
class CellMismatch:
    sheet: str
    row: int
    col: int
    expected: object
    actual: object

    def __str__(self):
        return f"[{self.sheet}] R{self.row}C{self.col}: expected={self.expected!r}  actual={self.actual!r}"


@dataclass
class WorkbookDiff:
    missing_sheets: list[str] = field(default_factory=list)
    extra_sheets: list[str] = field(default_factory=list)
    row_count_diffs: dict[str, tuple[int, int]] = field(default_factory=dict)
    cell_mismatches: list[CellMismatch] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return (not self.missing_sheets
                and not self.extra_sheets
                and not self.row_count_diffs
                and not self.cell_mismatches)

    def summary(self) -> str:
        parts = []
        if self.missing_sheets:
            parts.append(f"Missing sheets: {self.missing_sheets}")
        if self.extra_sheets:
            parts.append(f"Extra sheets: {self.extra_sheets}")
        for sheet, (exp, act) in self.row_count_diffs.items():
            parts.append(f"Row count [{sheet}]: expected {exp}, got {act}")
        for m in self.cell_mismatches[:20]:
            parts.append(str(m))
        if len(self.cell_mismatches) > 20:
            parts.append(f"... and {len(self.cell_mismatches) - 20} more cell mismatches")
        return "\n".join(parts) if parts else "Workbooks match"


def compare_workbooks(
    actual_path: str | Path,
    golden_path: str | Path,
    numeric_tolerance: float = 0.01,
    max_cell_mismatches: int = 100,
) -> WorkbookDiff:
    """Compare two Excel workbooks and return a diff report."""
    wb_actual = load_workbook(str(actual_path), data_only=True)
    wb_golden = load_workbook(str(golden_path), data_only=True)
    diff = WorkbookDiff()

    actual_sheets = set(wb_actual.sheetnames)
    golden_sheets = set(wb_golden.sheetnames)
    diff.missing_sheets = sorted(golden_sheets - actual_sheets)
    diff.extra_sheets = sorted(actual_sheets - golden_sheets)

    for sheet_name in sorted(actual_sheets & golden_sheets):
        ws_a = wb_actual[sheet_name]
        ws_g = wb_golden[sheet_name]

        if ws_a.max_row != ws_g.max_row:
            diff.row_count_diffs[sheet_name] = (ws_g.max_row, ws_a.max_row)

        max_row = min(ws_a.max_row or 0, ws_g.max_row or 0, 500)
        max_col = min(ws_a.max_column or 0, ws_g.max_column or 0, 50)

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                va = ws_a.cell(r, c).value
                vg = ws_g.cell(r, c).value
                if _values_equal(va, vg, numeric_tolerance):
                    continue
                diff.cell_mismatches.append(CellMismatch(sheet_name, r, c, vg, va))
                if len(diff.cell_mismatches) >= max_cell_mismatches:
                    break
            if len(diff.cell_mismatches) >= max_cell_mismatches:
                break

    wb_actual.close()
    wb_golden.close()
    return diff


def _values_equal(a, b, tol: float) -> bool:
    if a == b:
        return True
    if a is None or b is None:
        return str(a or "").strip() == str(b or "").strip()
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip() == str(b).strip()


def assert_workbook_structure(
    path: str | Path,
    expected_sheets: list[str],
    min_rows: dict[str, int] | None = None,
) -> None:
    """Assert that the workbook has the expected sheets and minimum row counts.

    Raises ``AssertionError`` with a descriptive message on failure.
    """
    wb = load_workbook(str(path), read_only=True)
    actual = wb.sheetnames

    for sheet in expected_sheets:
        assert sheet in actual, (
            f"Expected sheet '{sheet}' not found. Actual sheets: {actual}"
        )

    if min_rows:
        for sheet, min_r in min_rows.items():
            if sheet not in actual:
                continue
            ws = wb[sheet]
            row_count = ws.max_row or 0
            assert row_count >= min_r, (
                f"Sheet '{sheet}': expected >= {min_r} rows, got {row_count}"
            )

    wb.close()


def dump_workbook_summary(path: str | Path) -> dict[str, int]:
    """Return ``{sheet_name: row_count}`` for debugging."""
    wb = load_workbook(str(path), read_only=True)
    result = {name: (wb[name].max_row or 0) for name in wb.sheetnames}
    wb.close()
    return result
