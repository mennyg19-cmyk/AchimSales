from core.excel_writer import neutralize_excel_value


def test_formula_leaders_get_apostrophe():
    assert neutralize_excel_value("=cmd") == "'=cmd"
    assert neutralize_excel_value("+1+1") == "'+1+1"
    assert neutralize_excel_value("-2") == "'-2"
    assert neutralize_excel_value("@SUM(1)") == "'@SUM(1)"
    assert neutralize_excel_value("safe") == "safe"
    assert neutralize_excel_value(12) == 12
    assert neutralize_excel_value(None) is None


def test_number_4_cells_neutralize_formula_leaders():
    from openpyxl import Workbook

    from reports.number_4._styles import make_cell

    ws = Workbook(write_only=True).create_sheet("t")
    cell = make_cell(ws, "=1+1")
    assert cell.value == "'=1+1"


def test_salesman_excel_val_neutralizes_formula_leaders():
    from reports.salesman.writer import _excel_val

    assert _excel_val("=1+1") == "'=1+1"
    assert _excel_val(12) == 12


def test_ordered_summary_neutralizes_formula_leaders():
    import pandas as pd
    from openpyxl import Workbook

    from reports.ordered.writer import _write_summary_sheet

    df = pd.DataFrame([{"Customer Name": "=cmd", "Salesman": "safe"}])
    ws = Workbook().active
    _write_summary_sheet(ws, df)
    assert ws.cell(row=2, column=1).value == "'=cmd"


def test_invoiced_sheet_neutralizes_formula_leaders():
    import pandas as pd
    from openpyxl import Workbook

    from reports.invoiced.writer import _write_data_sheet

    wb = Workbook()
    df = pd.DataFrame([{"Customer": "=cmd", "Amount": 1}])
    _write_data_sheet(wb, "t", df, is_first=True)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "'=cmd"
