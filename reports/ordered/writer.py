"""
Ordered Report Excel writer.

Produces a multi-sheet workbook: Summary, By Customer, By Item, By Order,
By Salesman, Full Data. Supports both normal and streaming (write_only) modes.
"""

import logging
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from core.columns import pick_col, to_number
from core.excel_styles import (
    ALIGN_CENTER,
    BORDER_ORDER_END,
    BORDER_THIN,
    FILL_HEADER,
    FILL_LIGHT_GREY,
    FILL_SUMMARY_HEADER,
    FILL_TOTAL_ROW,
    FMT_CURRENCY,
    FMT_DATE,
    FONT_HEADER,
    FONT_LINK,
    score_to_fill,
)
from core.excel_writer import (
    STREAMING_ROW_THRESHOLD,
    make_streaming_cell,
    strip_datetime_tz,
)
from reports.ordered.builder import AGG_COLS, FULL_DATA_ORDER, SUMMARY_COLS

log = logging.getLogger(__name__)


def write_report(df: pd.DataFrame, out_path: str, report_variant: str | None = None) -> None:
    """Write the ordered report to an Excel file.

    report_variant: If "amazon_weekly", produces Amazon-specific layout: no legacy Summary
    sheet, Summary = By Customer (with CustomerRequisition; no SalesOrderName/ItemName),
    By Item, By Order only; no By Salesman, no Full Data; no sheet-to-sheet links.
    """
    if "_StatusCategory" not in df.columns:
        df["_StatusCategory"] = "Other"
    strip_datetime_tz(df)

    order_col = pick_col(df, ["SalesOrderNumber"])
    if order_col:
        df = df.sort_values(by=[order_col], na_position="last").reset_index(drop=True)
        next_order = df[order_col].astype(str).shift(-1)
        df["_LastLineOfOrder"] = (df[order_col].astype(str) != next_order) | next_order.isna()
    else:
        df["_LastLineOfOrder"] = False

    out_cols = [c for c in df.columns if not c.startswith("_")]
    currency_cols = [c for c in out_cols if c.endswith(" $")]
    ordered_out_cols = [c for c in FULL_DATA_ORDER if c in df.columns] or list(out_cols)

    large = len(df) > STREAMING_ROW_THRESHOLD
    log.info("Writing Excel: %d rows, mode=%s, variant=%s",
             len(df), "streaming" if large else "normal", report_variant or "default")
    if large:
        _write_streaming(df, out_path, ordered_out_cols, AGG_COLS, currency_cols, report_variant=report_variant)
    else:
        _write_normal(df, out_path, ordered_out_cols, AGG_COLS, currency_cols, report_variant=report_variant)
    try:
        size_kb = os.path.getsize(out_path) / 1024
        log.info("Excel saved: %.0f KB", size_kb)
    except OSError:
        pass


def _fulfillment_score(grp: pd.DataFrame) -> pd.Series:
    qo = grp["QtyOrdered"].fillna(0) if "QtyOrdered" in grp.columns else pd.Series(0.0, index=grp.index)
    qc = grp["QtyCancelled"].fillna(0) if "QtyCancelled" in grp.columns else pd.Series(0.0, index=grp.index)
    score = pd.Series(float("nan"), index=grp.index)
    mask = qo > 1e-6
    score.loc[mask] = ((qo - qc) / qo).loc[mask]
    return score.clip(0, 1)


def _build_summary_data(df: pd.DataFrame) -> pd.DataFrame:
    """Build Summary tab: aggregate by Customer + Item Number, total per customer."""
    df = df.copy()
    df["Customer Name"] = df.get("CustomerName", df.get("CustomerAccount", "")).fillna("").astype(str).str.strip()
    df["Salesman"] = df.get("Salesman", "").fillna("").astype(str).str.strip()
    df["Item Number"] = df.get("Item#", "").fillna("").astype(str).str.strip()
    df["Line Description"] = df.get("ItemName", df.get("LineDescription", "")).fillna("").astype(str).str.strip()

    net_price = df.get("SalesPrice", df.get("UnitPrice", 0))
    df["Net Price"] = to_number(net_price) if hasattr(net_price, "fillna") else net_price
    df["Extended Price - Ordered"] = df.get("Extended Price Ordered", df["QtyOrdered"].fillna(0) * df["Net Price"])
    df["Extended Price Remainder"] = df.get("Extended Price Remainder", df["QtyRemainder"].fillna(0) * df["Net Price"])

    sum_keys = ["QtyOrdered", "QtyCancelled", "QtyRemainder", "Extended Price - Ordered", "Extended Price Remainder"]

    agg_dict = {k: "sum" for k in sum_keys if k in df.columns}
    agg_dict["Line Description"] = "first"
    agg_dict["Salesman"] = "first"

    grouped = (
        df.groupby(["Customer Name", "Item Number"], dropna=False, sort=True)
        .agg(agg_dict)
        .reset_index()
    )
    for k in sum_keys:
        if k in grouped.columns:
            grouped[k] = to_number(grouped[k])
    qty_ord = grouped.get("QtyOrdered", 0)
    ext_ord = grouped.get("Extended Price - Ordered", 0)
    grouped["Net Price"] = ext_ord / qty_ord.replace(0, float("nan"))
    grouped["Net Price"] = grouped["Net Price"].fillna(0)

    grouped = grouped.sort_values(["Customer Name", "Item Number"], na_position="last").reset_index(drop=True)

    rows = []
    for cust, grp in grouped.groupby("Customer Name", dropna=False, sort=True):
        for _, r in grp.iterrows():
            rows.append({
                "Customer Name": cust,
                "Salesman": r.get("Salesman", ""),
                "Item Number": r.get("Item Number", ""),
                "Line Description": r.get("Line Description", ""),
                "QtyOrdered": r.get("QtyOrdered", 0),
                "QtyCancelled": r.get("QtyCancelled", 0),
                "QtyRemainder": r.get("QtyRemainder", 0),
                "Net Price": r.get("Net Price", 0),
                "Extended Price - Ordered": r.get("Extended Price - Ordered", 0),
                "Extended Price Remainder": r.get("Extended Price Remainder", 0),
                "_is_total": False,
                "_is_spacer": False,
            })
        total = {c: to_number(grp[c]).sum() if c in grp.columns else 0 for c in sum_keys}
        total["Customer Name"] = str(cust)
        total["Salesman"] = grp["Salesman"].iloc[0] if "Salesman" in grp.columns else ""
        total["Item Number"] = "TOTALS"
        total["Line Description"] = ""
        total["Net Price"] = ""
        total["_is_total"] = True
        total["_is_spacer"] = False
        rows.append(total)
        rows.append({k: "" for k in SUMMARY_COLS} | {"_is_total": False, "_is_spacer": True})

    grand = {k: 0 for k in sum_keys}
    for r in rows:
        if r.get("_is_total") or r.get("_is_spacer"):
            continue
        for k in sum_keys:
            v = r.get(k, 0)
            try:
                grand[k] += float(v) if v not in ("", None) else 0
            except (TypeError, ValueError):
                pass
    grand_row = {
        "Customer Name": "GRAND TOTAL",
        "Salesman": "",
        "Item Number": "",
        "Line Description": "",
        "Net Price": "",
        "_is_total": True,
        "_is_spacer": False,
    }
    grand_row.update(grand)
    if rows and not rows[-1].get("_is_spacer"):
        rows.append({k: "" for k in SUMMARY_COLS} | {"_is_total": False, "_is_spacer": True})
    rows.append(grand_row)

    return pd.DataFrame(rows)


def _write_summary_sheet(ws, summary_df: pd.DataFrame, is_write_only: bool = False) -> None:
    """Write Summary tab."""
    write_cols = [c for c in SUMMARY_COLS if c in summary_df.columns]
    if not write_cols:
        return
    currency_set = {"Net Price", "Extended Price - Ordered", "Extended Price Remainder"}

    if is_write_only:
        header_cells = [WriteOnlyCell(ws, c) for c in write_cols]
        for c in header_cells:
            c.fill = FILL_SUMMARY_HEADER
            c.font = FONT_HEADER
        ws.append(header_cells)
        for _, row in summary_df.iterrows():
            cells = []
            for col in write_cols:
                v = row.get(col, "")
                cell = WriteOnlyCell(ws, value=v)
                if row.get("_is_total"):
                    cell.fill = FILL_TOTAL_ROW
                    cell.font = FONT_HEADER
                if col in currency_set and v not in ("", None) and not (isinstance(v, str) and v.strip() == ""):
                    try:
                        cell.value = float(v) if not isinstance(v, (int, float)) else v
                        cell.number_format = FMT_CURRENCY
                    except (TypeError, ValueError):
                        pass
                cells.append(cell)
            ws.append(cells)
    else:
        for c_idx, col in enumerate(write_cols, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.fill = FILL_SUMMARY_HEADER
            cell.font = FONT_HEADER
        for r_idx, (_, row) in enumerate(summary_df.iterrows(), 2):
            for c_idx, col in enumerate(write_cols, 1):
                v = row.get(col, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                if row.get("_is_total"):
                    cell.fill = FILL_TOTAL_ROW
                    cell.font = FONT_HEADER
                if col in currency_set and v not in ("", None) and not (isinstance(v, str) and v.strip() == ""):
                    try:
                        cell.value = float(v) if not isinstance(v, (int, float)) else v
                        cell.number_format = FMT_CURRENCY
                    except (TypeError, ValueError):
                        pass


def _build_agg_sheet(df: pd.DataFrame, group_cols: list[str], agg_cols: list[str]) -> pd.DataFrame:
    """Build an aggregated sheet (By Customer, By Item, By Order, By Salesman)."""
    sum_cols = [c for c in agg_cols if c in df.columns]
    df_agg = df.copy()
    for c in sum_cols:
        df_agg[c] = to_number(df_agg[c])
    agg_dict = {c: "sum" for c in sum_cols}
    for extra in ["SalesOrderName", "ItemName"]:
        if extra in df_agg.columns and extra not in group_cols:
            agg_dict[extra] = "first"
    grp = df_agg.groupby(group_cols, dropna=False).agg(agg_dict).reset_index()
    grp = grp.sort_values(by=group_cols[0], na_position="last")
    grp["_FulfillmentScore"] = _fulfillment_score(grp)
    grp["Fulfillment %"] = grp["_FulfillmentScore"].apply(
        lambda x: f"{x * 100:.0f}%" if pd.notna(x) and x == x else ""
    )
    return grp


def _write_streaming(df, out_path, out_cols, agg_cols, currency_cols, report_variant=None):
    """Write using streaming (write_only) mode for large reports."""
    wb = Workbook(write_only=True)
    amazon = report_variant == "amazon_weekly"

    sheet_specs = _build_sheet_specs(df, out_cols, agg_cols, report_variant=report_variant)

    if not amazon and report_variant != "filtered":
        summary_df = _build_summary_data(df)
        ws_summary = wb.create_sheet(title="Summary", index=0)
        _write_summary_sheet(ws_summary, summary_df, is_write_only=True)
        if not summary_df.empty:
            try:
                ws_summary.freeze_panes = "A2"
                ws_summary.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLS))}{len(summary_df) + 1}"
            except Exception:
                log.debug("Could not set freeze_panes/auto_filter on Summary sheet", exc_info=True)

    for sheet_idx, (sheet_name, subset, write_cols, use_score, skip_totals_row) in enumerate(sheet_specs):
        log.info("  Writing sheet: %s (%d rows)", sheet_name, len(subset))
        if (amazon or report_variant == "filtered") and sheet_idx == 0:
            ws = wb.create_sheet(title=sheet_name, index=0)
        else:
            ws = wb.create_sheet(title=sheet_name)
        if subset.empty:
            ws.append([make_streaming_cell(ws, "No data")])
            continue

        header_cells = [
            make_streaming_cell(ws, col, fill=FILL_HEADER, font=FONT_HEADER, border=BORDER_THIN)
            for col in write_cols
        ]
        ws.append(header_cells)

        ff_idx = write_cols.index("Fulfillment %") if "Fulfillment %" in write_cols else None
        use_hyperlink = use_score and not amazon

        for row_idx, (_, row) in enumerate(subset.iterrows()):
            cells = []
            for j, col_name in enumerate(write_cols):
                v = row.get(col_name)
                if col_name in currency_cols and v not in (None, "", "nan"):
                    try:
                        v = float(str(v).replace(",", "").replace("$", "").strip())
                    except (TypeError, ValueError):
                        pass
                fmt = FMT_CURRENCY if col_name in currency_cols else None
                if col_name == "OrderDate" and hasattr(v, "strftime"):
                    fmt = FMT_DATE

                cell_fill = None
                if ff_idx is not None and j == ff_idx:
                    cell_fill = score_to_fill(row.get("_FulfillmentScore"))
                elif row_idx % 2 == 1:
                    cell_fill = FILL_LIGHT_GREY

                font = None
                if use_hyperlink and j == 0 and v not in (None, "", "Total"):
                    font = FONT_LINK

                cell = make_streaming_cell(ws, v, fill=cell_fill, font=font, border=BORDER_THIN, fmt=fmt)
                if use_hyperlink and j == 0 and v not in (None, "", "Total"):
                    cell.hyperlink = "#'Full Data'!A1"
                cells.append(cell)
            ws.append(cells)

        if not skip_totals_row:
            totals_cells = []
            for col_name in write_cols:
                if col_name == write_cols[0]:
                    val = "Total"
                elif col_name in agg_cols and col_name in subset.columns:
                    val = to_number(subset[col_name]).sum()
                else:
                    val = ""
                fmt = FMT_CURRENCY if col_name in currency_cols else None
                totals_cells.append(make_streaming_cell(ws, val, fill=FILL_HEADER, font=FONT_HEADER, border=BORDER_THIN, fmt=fmt))
            ws.append(totals_cells)

        try:
            ws.freeze_panes = "A2"
            n_rows = len(subset) + (2 if not skip_totals_row else 1)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(write_cols))}{n_rows}"
        except Exception:
            log.debug("Could not set freeze_panes/auto_filter on %s sheet", sheet_name, exc_info=True)

    wb.save(out_path)


def _write_normal(df, out_path, out_cols, agg_cols, currency_cols, report_variant=None):
    """Write using normal (non-streaming) mode."""
    wb = Workbook()
    amazon = report_variant == "amazon_weekly"

    sheet_specs = _build_sheet_specs(df, out_cols, agg_cols, report_variant=report_variant)

    if not amazon and report_variant != "filtered":
        summary_df = _build_summary_data(df)
        ws_summary = wb.active
        ws_summary.title = "Summary"
        _write_summary_sheet(ws_summary, summary_df, is_write_only=False)
        if not summary_df.empty:
            ws_summary.freeze_panes = "A2"
            ws_summary.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLS))}{len(summary_df) + 1}"

    for sheet_idx, (sheet_name, subset, write_cols, use_score, skip_totals_row) in enumerate(sheet_specs):
        log.info("  Writing sheet: %s (%d rows)", sheet_name, len(subset))
        if (amazon or report_variant == "filtered") and sheet_idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        if subset.empty:
            ws.cell(row=1, column=1, value="No data")
            continue

        is_full_data = sheet_name == "Full Data"

        for c_idx, col_name in enumerate(write_cols, 1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.border = BORDER_THIN

        ff_idx = write_cols.index("Fulfillment %") + 1 if "Fulfillment %" in write_cols else None

        if is_full_data:
            order_col = pick_col(subset, ["SalesOrderNumber"])
            if order_col:
                subset["_OrderIndex"] = subset.groupby(order_col).ngroup()
            else:
                subset["_OrderIndex"] = 0

        for r_idx, row in enumerate(dataframe_to_rows(subset[write_cols], index=False, header=False), 2):
            data_idx = r_idx - 2
            for c_idx, value in enumerate(row, 1):
                col_name = write_cols[c_idx - 1] if c_idx <= len(write_cols) else ""
                if col_name in currency_cols and value not in (None, "", "nan"):
                    try:
                        value = float(str(value).replace(",", "").replace("$", "").strip())
                    except (TypeError, ValueError):
                        pass

                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = BORDER_THIN

                if ff_idx and c_idx == ff_idx:
                    score = None
                    if "_FulfillmentScore" in subset.columns and 0 <= data_idx < len(subset):
                        score = subset["_FulfillmentScore"].iloc[data_idx]
                    cell.fill = score_to_fill(score)
                elif is_full_data and "_OrderIndex" in subset.columns:
                    idx = subset["_OrderIndex"].iloc[data_idx] if 0 <= data_idx < len(subset) else 0
                    if idx % 2 == 1:
                        cell.fill = FILL_LIGHT_GREY
                elif not is_full_data and data_idx % 2 == 1:
                    cell.fill = FILL_LIGHT_GREY

                use_hyperlink = use_score and report_variant != "amazon_weekly"
                if use_hyperlink and c_idx == 1 and value not in (None, "", "Total"):
                    cell.hyperlink = "#'Full Data'!A1"
                    cell.font = FONT_LINK

                if col_name in currency_cols:
                    cell.number_format = FMT_CURRENCY
                if col_name == "OrderDate" and pd.notna(value) and hasattr(value, "strftime"):
                    cell.number_format = FMT_DATE

        if not skip_totals_row:
            totals_row = len(subset) + 2
            for c_idx, col_name in enumerate(write_cols, 1):
                if col_name == write_cols[0]:
                    ws.cell(row=totals_row, column=c_idx, value="Total")
                elif col_name in agg_cols and col_name in subset.columns:
                    val = to_number(subset[col_name]).sum()
                    cell = ws.cell(row=totals_row, column=c_idx, value=val)
                    if col_name in currency_cols:
                        cell.number_format = FMT_CURRENCY
                else:
                    ws.cell(row=totals_row, column=c_idx, value="")
                ws.cell(row=totals_row, column=c_idx).border = BORDER_THIN
                ws.cell(row=totals_row, column=c_idx).font = FONT_HEADER

        for c_idx, col_name in enumerate(write_cols, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = max(10, min(35, len(str(col_name)) + 2))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(write_cols))}{ws.max_row}"

    wb.save(out_path)


def _build_summary_one_row(df: pd.DataFrame, agg_cols: list[str]) -> pd.DataFrame:
    """Build a single-row Summary tab: one line of totals (for amazon_weekly)."""
    sum_cols = [c for c in agg_cols if c in df.columns]
    if not sum_cols:
        return pd.DataFrame()
    row = {c: to_number(df[c]).sum() for c in sum_cols}
    row["Total"] = "Total"
    # Fulfillment % for the whole report
    qo = row.get("QtyOrdered", 0) or 0
    qc = row.get("QtyCancelled", 0) or 0
    score = ((qo - qc) / qo * 100) if qo and qo > 1e-6 else None
    if score is not None:
        row["Fulfillment %"] = f"{score:.0f}%"
    else:
        row["Fulfillment %"] = ""
    # Column order: Total, Fulfillment %, then agg cols
    cols = ["Total", "Fulfillment %"] + [c for c in agg_cols if c in row]
    return pd.DataFrame([{c: row.get(c, "") for c in cols}])


def _build_sheet_specs(df, out_cols, agg_cols, report_variant=None):
    """Build list of (sheet_name, subset_df, write_cols, use_score_fill, skip_totals_row) for all tabs.

    Variants:
      None (default)     -- all tabs: By Customer, By Item, By Order, By Salesman, Full Data
      "amazon_weekly"    -- Summary (one-row totals), By Item, By Order (with PO #)
      "filtered"         -- Summary (aggregated by customer), By Item, By Order
    """
    amazon = report_variant == "amazon_weekly"
    filtered = report_variant == "filtered"
    specs = []

    if amazon:
        one_row = _build_summary_one_row(df, agg_cols)
        write_cols = list(one_row.columns) if not one_row.empty else ["Total"]
        specs.append(("Summary", one_row, write_cols, False, True))
    elif filtered:
        cust_cols = [c for c in ["CustomerAccount", "CustomerName"] if c in df.columns]
        if cust_cols:
            grp = _build_agg_sheet(df, cust_cols, agg_cols)
            write_cols = [c for c in grp.columns if c not in ("_StatusCategory", "_FulfillmentScore", "ItemName", "SalesOrderName")]
            specs.append(("Summary", grp, write_cols, False, False))
        else:
            specs.append(("Summary", pd.DataFrame(), out_cols, False, False))
    else:
        cust_cols = [c for c in ["CustomerAccount", "CustomerName"] if c in df.columns]
        if "Salesman" in df.columns:
            cust_cols = cust_cols + ["Salesman"]
        if cust_cols:
            grp = _build_agg_sheet(df, cust_cols, agg_cols)
            write_cols = [c for c in grp.columns if c not in ("_StatusCategory", "_FulfillmentScore", "ItemName", "SalesOrderName")]
            specs.append(("By Customer", grp, write_cols, True, False))
        else:
            specs.append(("By Customer", pd.DataFrame(), out_cols, False, False))

    if not amazon:
        item_cols = ["Item#"] if "Item#" in df.columns else []
        if item_cols:
            grp = _build_agg_sheet(df, item_cols, agg_cols)
            lead = [c for c in ["Item#", "ItemName"] if c in grp.columns]
            exclude = {"_StatusCategory", "_FulfillmentScore", "CustomerName", "SalesOrderName"}
            rest = [c for c in grp.columns if c not in exclude and c not in lead]
            specs.append(("By Item", grp, lead + rest, not amazon, False))
        else:
            specs.append(("By Item", pd.DataFrame(), out_cols, False, False))

    order_col = pick_col(df, ["SalesOrderNumber"])
    if order_col:
        ctx = [c for c in ["OrderDate", "CustomerAccount", "Salesman", "CustomerRequisition", "SalesOrderName"] if c in df.columns]
        grp = _build_agg_sheet(df, [order_col] + ctx, agg_cols)
        write_cols = [c for c in grp.columns if c not in ("_StatusCategory", "_FulfillmentScore", "ItemName")]
        if "CustomerRequisition" in grp.columns:
            grp = grp.rename(columns={"CustomerRequisition": "PO #"})
            write_cols = ["PO #" if c == "CustomerRequisition" else c for c in write_cols]
        specs.append(("By Order", grp, write_cols, not amazon and not filtered, False))
    else:
        specs.append(("By Order", pd.DataFrame(), out_cols, False, False))

    if not amazon and not filtered:
        if "Salesman" in df.columns:
            grp = _build_agg_sheet(df, ["Salesman"], agg_cols)
            write_cols = [c for c in grp.columns if c not in ("_StatusCategory", "_FulfillmentScore", "SalesOrderName", "ItemName")]
            specs.append(("By Salesman", grp, write_cols, True, False))
        else:
            specs.append(("By Salesman", pd.DataFrame(), out_cols, False, False))

        full_data = df[out_cols + ["_StatusCategory", "_LastLineOfOrder", "_FulfillmentScore"]].copy()
        write_cols = [c for c in FULL_DATA_ORDER if c in full_data.columns]
        if not write_cols:
            write_cols = [c for c in full_data.columns if not c.startswith("_")]
        specs.append(("Full Data", full_data, write_cols, False, False))

    return specs
