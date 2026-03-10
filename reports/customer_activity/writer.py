"""
Customer Activity Report -- Excel writer.

Produces two types of output:
  1. Individual salesman file: single sheet with that salesman's customers.
  2. Management workbook: All tab, one sheet per salesman, and Unassigned tab.
"""

import logging
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.excel_styles import (
    ALIGN_CENTER,
    BORDER_THIN,
    FILL_HEADER,
    FILL_LIGHT_GREY,
    FILL_TOTAL_ROW,
    FMT_DATE,
    FONT_HEADER,
    FONT_TITLE,
)
from core.excel_writer import autosize_columns, strip_datetime_tz
from reports.customer_activity.builder import OUTPUT_COLUMNS

log = logging.getLogger(__name__)


def _prep_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Replace NaT Last Order Date with 'N/A' string for display."""
    df = df.copy()
    if "Last Order Date" in df.columns:
        mask = df["Last Order Date"].isna()
        df["Last Order Date"] = df["Last Order Date"].astype(object)
        df.loc[mask, "Last Order Date"] = "N/A"
    return df


def _write_sheet(ws, df: pd.DataFrame, title: str | None = None, columns: list[str] | None = None) -> None:
    """Write a single customer activity sheet."""
    df = _prep_for_display(strip_datetime_tz(df))
    if columns is None:
        columns = [c for c in OUTPUT_COLUMNS if c in df.columns]
    else:
        columns = [c for c in columns if c in df.columns]
    if not columns:
        ws.cell(row=1, column=1, value="No data")
        return

    start_row = 1
    if title:
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = FONT_TITLE
        start_row = 3

    for c_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER

    for r_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for c_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER_THIN

            if col_name == "Last Order Date" and value != "N/A" and pd.notna(value) and hasattr(value, "strftime"):
                cell.number_format = FMT_DATE

            if (r_idx - start_row) % 2 == 0:
                cell.fill = FILL_LIGHT_GREY

    totals_row = start_row + 1 + len(df)
    for c_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=totals_row, column=c_idx)
        cell.border = BORDER_THIN
        cell.font = FONT_HEADER
        cell.fill = FILL_TOTAL_ROW

        if col_name == columns[0]:
            cell.value = f"Total ({len(df)} customers)"
        else:
            cell.value = ""

    autosize_columns(ws)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{totals_row}"


def write_individual_report(df: pd.DataFrame, salesman_name: str, out_path: str) -> None:
    """Write a single-salesman Excel file."""
    log.info("Writing Customer Activity for %s (%d customers)", salesman_name, len(df))
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Activity"
    _write_sheet(ws, df, title=f"{salesman_name} - Customer Activity")
    wb.save(out_path)
    try:
        size_kb = os.path.getsize(out_path) / 1024
        log.info("Excel saved: %.0f KB -- %s", size_kb, out_path)
    except OSError:
        log.info("Wrote individual report: %s", out_path)


def write_management_report(
    all_data: dict[str, pd.DataFrame],
    out_path: str,
    unassigned_df: pd.DataFrame | None = None,
) -> None:
    """Write the combined management workbook.

    Tabs (in order):
      1. "All" -- every customer from every salesman + unassigned, with Salesman column
      2. One tab per salesman (sorted alphabetically)
      3. "Unassigned" -- customers with no salesman
    """
    total_sheets = len(all_data) + (1 if unassigned_df is not None and not unassigned_df.empty else 0) + 1
    log.info("Writing Customer Activity management report (%d salesmen, %d total sheets)", len(all_data), total_sheets)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb = Workbook()

    all_tab_cols = ["Salesman"] + list(OUTPUT_COLUMNS)

    combined_frames = []
    for salesman_name in sorted(all_data.keys()):
        df = all_data[salesman_name].copy()
        df["Salesman"] = salesman_name
        combined_frames.append(df)

    if unassigned_df is not None and not unassigned_df.empty:
        udf = unassigned_df.copy()
        udf["Salesman"] = "Unassigned"
        combined_frames.append(udf)

    ws_all = wb.active
    ws_all.title = "All"
    if combined_frames:
        combined_df = pd.concat(combined_frames, ignore_index=True)
        log.info("  Writing sheet: All (%d customers)", len(combined_df))
        _write_sheet(ws_all, combined_df, title="All Salesmen - Customer Activity", columns=all_tab_cols)
    else:
        ws_all.cell(row=1, column=1, value="No customer data available")

    for salesman_name in sorted(all_data.keys()):
        df = all_data[salesman_name]
        sheet_title = salesman_name[:31]
        log.info("  Writing sheet: %s (%d customers)", sheet_title, len(df))
        ws = wb.create_sheet(title=sheet_title)
        _write_sheet(ws, df, title=f"{salesman_name} - Customer Activity")

    if unassigned_df is not None and not unassigned_df.empty:
        log.info("  Writing sheet: Unassigned (%d customers)", len(unassigned_df))
        ws_un = wb.create_sheet(title="Unassigned")
        _write_sheet(ws_un, unassigned_df, title="Unassigned Customers - Customer Activity")

    wb.save(out_path)
    try:
        size_kb = os.path.getsize(out_path) / 1024
        log.info("Excel saved: %.0f KB -- %s (%d sheets)", size_kb, out_path, total_sheets)
    except OSError:
        log.info("Wrote management report: %s (%d sheets)", out_path, total_sheets)
