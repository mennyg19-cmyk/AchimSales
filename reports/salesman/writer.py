"""
Salesman Report Excel writer.

Produces the Monthly Salesmen Report workbook with per-salesman sections,
customer rows, totals, and color-banded columns.
"""

import calendar
import logging
import math

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.excel_styles import BORDER_THIN, FILL_HEADER_BLUE, FILL_TOTALS, FMT_CURRENCY, FONT_HEADER
from core.excel_writer import neutralize_excel_value

log = logging.getLogger(__name__)

FONT_BLUE = Font(color="0000CC")
FONT_GREEN = Font(color="008000")
FONT_PURPLE = Font(color="800080")
FONT_RED = Font(color="FF0000")

COL_WIDTHS = [13, 12, 48, 21, 13, 26, 13, 30, 13, 32, 13, 25, 13, 42, 13]


def write_monthly_salesmen_workbook(
    month_data_by_month: dict[int, pd.DataFrame],
    year: int,
    out_path: str,
) -> None:
    """Write one workbook with 12 month tabs (Jan-Dec), matching legacy structure."""
    wb = Workbook()
    wb.remove(wb.active)

    last_year = year - 1

    for m in range(1, 13):
        mon3 = calendar.month_abbr[m]
        mon_full = calendar.month_name[m]
        ws = wb.create_sheet(mon3)

        headers = _build_monthly_headers(year, last_year, mon_full)
        _write_header_row_monthly(ws, headers)

        month_data = month_data_by_month.get(m, pd.DataFrame())
        if month_data.empty:
            continue

        sort_num = 1
        salesmen_iter = month_data.groupby("Salesman", dropna=False, sort=False)
        # Sort by Sort Number (padded salesman #) to match legacy order
        salesmen = sorted(
            salesmen_iter,
            key=lambda x: (str(x[1]["Sort Number"].iloc[0]) if "Sort Number" in x[1].columns and len(x[1]) else "", str(x[0])),
        )

        for sm_name, grp in salesmen:
            sm_num = grp["SalesmanNumber"].iloc[0] if "SalesmanNumber" in grp.columns else str(sm_name)

            # Salesman header row: Sort Number, Salesman #, Salesman Name, blanks
            _write_salesman_header_row(ws, sort_num, sm_num, str(sm_name))
            sort_num += 1

            for _, r in grp.iterrows():
                _write_data_row_monthly(ws, sort_num, r)
                sort_num += 1

            _write_totals_row_monthly(ws, sort_num, grp, str(sm_name))
            sort_num += 1

            # Blank row
            _write_blank_row(ws, sort_num)
            sort_num += 1

        _write_grand_totals_monthly(ws, sort_num, month_data)
        _apply_color_bands(ws, 2, ws.max_row)
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(15)}{ws.max_row}"

    wb.save(out_path)


def write_individual_salesman_workbook(
    month_data_by_month: dict[int, pd.DataFrame],
    year: int,
    salesman_name: str,
    out_path: str,
) -> None:
    """Write a per-salesman workbook with 12 month tabs, filtered to one salesman."""
    import os
    filtered = {}
    for m, df in month_data_by_month.items():
        if df.empty or "Salesman" not in df.columns:
            filtered[m] = pd.DataFrame()
            continue
        subset = df[df["Salesman"].astype(str).str.strip() == salesman_name].copy()
        filtered[m] = subset

    if all(df.empty for df in filtered.values()):
        log.info("No data for salesman %s -- skipping individual file", salesman_name)
        return

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    write_monthly_salesmen_workbook(filtered, year, out_path)
    log.info("Wrote individual salesman report: %s", out_path)


def get_all_salesmen(month_data_by_month: dict[int, pd.DataFrame]) -> list[str]:
    """Return a sorted list of all unique salesman display names across all months."""
    names = set()
    for df in month_data_by_month.values():
        if not df.empty and "Salesman" in df.columns:
            names.update(df["Salesman"].dropna().astype(str).str.strip().unique())
    names.discard("")
    return sorted(names)


def _build_monthly_headers(year: int, last_year: int, mon_full: str) -> list[str]:
    return [
        "Sort Number",
        "Cust. #",
        "Customer Name",
        f"Sales {mon_full} {year}",
        f"Sales {mon_full} {last_year}",
        "$ This Year to Last Year",
        "% This Year to Last Year",
        f"Sales {year} Jan Thru {mon_full}",
        f"Sales {last_year} Jan Thru {mon_full}",
        "$ This Year to Last Year (YTD)",
        "% This Year to Last Year (YTD)",
        f"Sales Year to Date {year}",
        f"Sales Year to Date {last_year}",
        "$ This Year to Last Year (YTD Full Year)",
        "% This Year to Last Year (YTD Full Year)",
    ]


def _write_header_row_monthly(ws, headers: list[str]) -> None:
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_BLUE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    ws.freeze_panes = "A2"
    for c_idx, w in enumerate(COL_WIDTHS, 1):
        if c_idx <= len(headers):
            ws.column_dimensions[get_column_letter(c_idx)].width = w


def _write_salesman_header_row(ws, sort_num: int, sm_num: str, sm_name: str) -> None:
    row = ws.max_row + 1
    vals = [sort_num, sm_num, sm_name] + [None] * 12
    for c in range(1, 16):
        cell = ws.cell(row=row, column=c, value=_excel_val(vals[c - 1]) if c <= 3 else None)
        cell.border = BORDER_THIN
        if c in (4, 5, 6, 8, 9, 10, 12, 13, 14):
            cell.number_format = FMT_CURRENCY
        elif c in (7, 11, 15):
            cell.number_format = "0.00%"


def _write_data_row_monthly(ws, sort_num: int, r: dict) -> None:
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=sort_num)
    ws.cell(row=row, column=2, value=_excel_val(r.get("Cust. #", "")))
    ws.cell(row=row, column=3, value=_excel_val(r.get("Customer Name", "")))

    money_vals = [
        r.get("Sales_Current", 0), r.get("Sales_Prior", 0),
        r.get("$ Month Diff", 0), r.get("% Month Diff", 0),
        r.get("Sales_YTD_Current", 0), r.get("Sales_YTD_Prior", 0),
        r.get("$ YTD Diff", 0), r.get("% YTD Diff", 0),
        r.get("Sales_FullYear_Current", 0), r.get("Sales_FullYear_Prior", 0),
        r.get("$ FullYear Diff", 0), r.get("% FullYear Diff", 0),
    ]

    band_fonts = [FONT_BLUE] * 4 + [FONT_GREEN] * 4 + [FONT_PURPLE] * 4

    for i, val in enumerate(money_vals):
        c_idx = 4 + i
        cell = ws.cell(row=row, column=c_idx, value=_excel_val(val))
        cell.border = BORDER_THIN
        is_pct = i in (3, 7, 11)
        cell.number_format = "0.00%" if is_pct else FMT_CURRENCY
        num_val = _as_num(val)
        if num_val is not None and num_val < 0:
            cell.font = FONT_RED
        else:
            cell.font = band_fonts[i]


def _write_totals_row_monthly(ws, sort_num: int, grp: pd.DataFrame, sm_name: str) -> None:
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=sort_num)
    ws.cell(row=row, column=2, value="Total for:")
    ws.cell(row=row, column=3, value=_excel_val(sm_name))
    for c in range(1, 16):
        ws.cell(row=row, column=c).fill = FILL_TOTALS
        ws.cell(row=row, column=c).border = BORDER_THIN
        ws.cell(row=row, column=c).font = FONT_HEADER

    sum_cols = [
        "Sales_Current", "Sales_Prior", "$ Month Diff", None,
        "Sales_YTD_Current", "Sales_YTD_Prior", "$ YTD Diff", None,
        "Sales_FullYear_Current", "Sales_FullYear_Prior", "$ FullYear Diff", None,
    ]
    for i, col in enumerate(sum_cols):
        c_idx = 4 + i
        cell = ws.cell(row=row, column=c_idx)
        is_pct = i in (3, 7, 11)
        if is_pct:
            base_col = sum_cols[i - 1]
            prior_col = sum_cols[i - 2]
            if base_col and prior_col and base_col in grp.columns and prior_col in grp.columns:
                diff = grp[base_col].sum()
                prior = grp[prior_col].sum()
                cell.value = diff / prior if prior else 0
            cell.number_format = "0.00%"
        elif col and col in grp.columns:
            cell.value = grp[col].sum()
            cell.number_format = FMT_CURRENCY


def _write_blank_row(ws, sort_num: int) -> None:
    row = ws.max_row + 1
    for c in range(1, 16):
        cell = ws.cell(row=row, column=c, value=sort_num if c == 1 else None)
        cell.border = BORDER_THIN
        if c in (4, 5, 6, 8, 9, 10, 12, 13, 14):
            cell.number_format = FMT_CURRENCY
        elif c in (7, 11, 15):
            cell.number_format = "0.00%"


def _write_grand_totals_monthly(ws, sort_num: int, data: pd.DataFrame) -> None:
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=sort_num)
    ws.cell(row=row, column=2, value="Grand total:")
    ws.cell(row=row, column=3, value="")
    for c in range(1, 16):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_TOTALS
        cell.border = BORDER_THIN
        cell.font = Font(bold=True)

    sum_cols = [
        "Sales_Current", "Sales_Prior", "$ Month Diff", None,
        "Sales_YTD_Current", "Sales_YTD_Prior", "$ YTD Diff", None,
        "Sales_FullYear_Current", "Sales_FullYear_Prior", "$ FullYear Diff", None,
    ]
    for i, col in enumerate(sum_cols):
        c_idx = 4 + i
        cell = ws.cell(row=row, column=c_idx)
        is_pct = i in (3, 7, 11)
        if is_pct:
            base_col = sum_cols[i - 1]
            prior_col = sum_cols[i - 2]
            if base_col and prior_col and base_col in data.columns and prior_col in data.columns:
                diff = data[base_col].sum()
                prior = data[prior_col].sum()
                cell.value = diff / prior if prior else 0
            cell.number_format = "0.00%"
        elif col and col in data.columns:
            cell.value = data[col].sum()
            cell.number_format = FMT_CURRENCY


def _excel_val(v) -> None | int | float | str:
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, str) and v.strip().lower() == "nan":
        return None
    return neutralize_excel_value(v)


def _as_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return float(v)
    try:
        s = str(v).replace(",", "").replace("$", "").strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


def _apply_color_bands(ws, start_row: int, end_row: int) -> None:
    """Apply blue/green/purple bands to cols D-O; red for negatives."""
    for r in range(start_row, end_row + 1):
        for c in range(4, 16):
            cell = ws.cell(row=r, column=c)
            base_font = cell.font or Font()
            is_bold = bool(base_font.bold)
            num = _as_num(cell.value)
            if c in (7, 11, 15):
                cell.number_format = "0.00%"
            elif num is not None or (isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and math.isnan(cell.value))):
                cell.number_format = FMT_CURRENCY

            if c <= 7:
                band = "0000CC"
            elif c <= 11:
                band = "008000"
            else:
                band = "800080"

            if num is not None and num < 0:
                cell.font = Font(name=base_font.name, size=base_font.sz, bold=is_bold, color="FF0000")
            else:
                cell.font = Font(name=base_font.name, size=base_font.sz, bold=is_bold, color=band)
