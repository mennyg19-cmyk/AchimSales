"""
Invoiced Report Excel writer.

Produces multi-sheet workbooks for MTD and Daily reports:
  Summary by Customer, Commissions, Full Details, Credits, Invoices, Audit

For large workbooks (>10K total data rows), uses openpyxl write_only mode
to keep memory bounded -- rows are flushed to a temp file as they are written
instead of being held in memory.
"""

import logging
import math
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.excel_styles import (
    BORDER_THIN,
    FILL_HEADER,
    FILL_HEADER_BLUE,
    FMT_CURRENCY,
    FONT_HEADER,
)
from core.excel_writer import (
    autosize_columns,
    format_money_columns,
    neutralize_excel_value,
    strip_datetime_tz,
)

log = logging.getLogger(__name__)

STREAMING_THRESHOLD = 10_000


def _is_money_column(col_name: str) -> bool:
    """Same matching logic as format_money_columns in core/excel_writer.py."""
    key = str(col_name).strip().lower()
    if "count" in key:
        return False
    money_kw = {
        "subtotal invoices", "tariff charges", "freight charges", "cc charges",
        "total invoice", "total invoices", "total tariff charges",
        "total freight charges", "total cc charges",
        "sales", "freight", "cc", "tariff", "commissions",
        "commission base", "totnofrt",
    }
    if key in money_kw:
        return True
    return any(kw in key for kw in ("charge", "invoice", "sales", "tot"))


def write_invoiced_report(
    summary: pd.DataFrame,
    commissions: pd.DataFrame,
    details: pd.DataFrame,
    credits: pd.DataFrame,
    invoices: pd.DataFrame,
    audit: pd.DataFrame | None,
    out_path: str,
    *,
    year: int | None = None,
    full_detail: pd.DataFrame | None = None,
    ytd_credits: pd.DataFrame | None = None,
    ytd_invoices: pd.DataFrame | None = None,
    pct_map: dict[str, float] | None = None,
    current_month: int | None = None,
    skip_commissions: bool = False,
) -> None:
    """Write the Invoiced Report workbook.

    Automatically switches to openpyxl ``write_only`` mode when the total
    data rows exceed STREAMING_THRESHOLD to avoid OOM in memory-constrained
    environments (e.g. Azure Automation sandboxes).

    Set ``skip_commissions=True`` for per-salesman Shipped Reports where the
    commissions tab is not meaningful.
    """
    from core.logging import log_memory

    total_data_rows = len(details) + len(invoices)
    streaming = total_data_rows > STREAMING_THRESHOLD

    log.info("Writing Invoiced Excel: %d summary, %d detail, %d invoice rows%s",
             len(summary), len(details), len(invoices),
             " (streaming)" if streaming else "")
    log_memory("invoiced:writer:start (%d detail, %d invoice rows)" % (len(details), len(invoices)))

    if streaming:
        wb = Workbook(write_only=True)
        _wds = _write_data_sheet_streaming
    else:
        wb = Workbook()
        _wds = _write_data_sheet

    log.info("  Writing sheet: Summary by Customer (%d rows)", len(summary))
    _wds(wb, "Summary by Customer", summary, is_first=not streaming)

    if not skip_commissions:
        if year is not None and full_detail is not None and pct_map is not None:
            log.info("  Writing sheet: Commissions (legacy monthly format)")
            comm_credits = ytd_credits if ytd_credits is not None else credits
            comm_invoices = ytd_invoices if ytd_invoices is not None else invoices
            if streaming:
                _write_commissions_sheet_streaming(
                    wb, year, full_detail, comm_credits, comm_invoices, pct_map,
                    current_month=current_month)
            else:
                _write_commissions_sheet(
                    wb, year, full_detail, comm_credits, comm_invoices, pct_map,
                    current_month=current_month)
        else:
            log.info("  Writing sheet: Commissions (fallback -- no monthly data)")
            if streaming:
                _write_commissions_sheet_simple_streaming(wb, commissions)
            else:
                _write_commissions_sheet_simple(wb, commissions)

    log.info("  Writing sheet: Full Details (%d rows)", len(details))
    _wds(wb, "Full Details", details)
    log.info("  Writing sheet: Credits (%d rows)", len(credits))
    _wds(wb, "Credits", credits)
    log.info("  Writing sheet: Invoices (%d rows)", len(invoices))
    _wds(wb, "Invoices", invoices)

    if audit is not None and not audit.empty:
        log.info("  Writing sheet: Audit - Reversals (%d rows)", len(audit))
        _wds(wb, "Audit - Reversals", audit)

    _maybe_write_totals_by_salesman(wb, invoices, streaming=streaming)

    log_memory("invoiced:writer:before_save")
    wb.save(out_path)
    try:
        size_kb = os.path.getsize(out_path) / 1024
        log.info("Excel saved: %.0f KB -- %s", size_kb, out_path)
    except OSError:
        log.info("Wrote: %s", out_path)


# ---------------------------------------------------------------------------
# Normal-mode sheet writers (used for small reports, < STREAMING_THRESHOLD)
# ---------------------------------------------------------------------------

def _write_commissions_sheet_simple(wb: Workbook, comm_df: pd.DataFrame) -> None:
    """Fallback: simple commissions sheet when monthly data is not available."""
    ws = wb.create_sheet(title="Commissions")
    if comm_df.empty or "SalesmanNumber" not in comm_df.columns:
        ws.cell(row=1, column=1, value="No commission data")
        return
    commissioned = comm_df[comm_df["Percent"] > 0].copy()
    if commissioned.empty:
        ws.cell(row=1, column=1, value="No commissioned salesmen")
        return
    ws.cell(row=1, column=1, value="Commissions Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    row = 3
    for sm_num, grp in commissioned.groupby("SalesmanNumber", sort=True):
        sm_name = grp["SalesmanName"].iloc[0] if "SalesmanName" in grp.columns else ""
        pct = grp["Percent"].iloc[0]
        ws.cell(row=row, column=1, value=neutralize_excel_value(f"Salesman #{sm_num} - {sm_name}"))
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=row, column=3, value=f"{pct:.0%}")
        row += 1
        sub_total = grp["SubTotal Invoices"].sum() if "SubTotal Invoices" in grp.columns else 0
        tariff_total = grp["Total Tariff Charges"].sum() if "Total Tariff Charges" in grp.columns else 0
        freight_total = grp["Total Freight Charges"].sum() if "Total Freight Charges" in grp.columns else 0
        cc_total = grp["Total CC Charges"].sum() if "Total CC Charges" in grp.columns else 0
        inv_total = grp["Total Invoices"].sum() if "Total Invoices" in grp.columns else 0
        net = sub_total + tariff_total
        commission = net * pct
        for label, val in [("SubTotal Invoices", sub_total), ("Total Tariff Charges", tariff_total),
                           ("Total Freight Charges", freight_total), ("Total CC Charges", cc_total),
                           ("Total Invoices", inv_total), ("Net Commission Amount", net), ("Commission", commission)]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val).number_format = FMT_CURRENCY
            row += 1
        row += 1
    autosize_columns(ws)


def _write_data_sheet(wb: Workbook, title: str, df: pd.DataFrame, is_first: bool = False) -> None:
    """Write a simple data sheet with header styling and money formatting.

    Uses numpy array access instead of iterrows() to minimize memory overhead
    when writing large DataFrames (50K+ rows).
    """
    if is_first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    if df.empty:
        ws.cell(row=1, column=1, value="No data")
        return

    strip_datetime_tz(df)
    cols = [c for c in df.columns if not c.startswith("_")]
    col_indices = [df.columns.get_loc(c) for c in cols]

    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_BLUE
        cell.border = BORDER_THIN

    data = df.values
    num_rows = len(data)
    for r_idx in range(num_rows):
        row_data = data[r_idx]
        for c_idx, ci in enumerate(col_indices, 1):
            v = row_data[ci]
            if v is None:
                v = ""
            elif isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                v = 0.0
            ws.cell(row=r_idx + 2, column=c_idx, value=neutralize_excel_value(v))

    totals_row = num_rows + 2
    money_cols = {"SubTotal Invoices", "Total Tariff Charges", "Total Freight Charges",
                  "Total CC Charges", "Total Invoices", "Total Invoice",
                  "Tariff Charges", "Freight Charges", "CC Charges", "Commissions", "Commission Base"}
    for c_idx, col in enumerate(cols, 1):
        if col in money_cols and col in df.columns:
            val = pd.to_numeric(df[col], errors="coerce").fillna(0).sum()
            cell = ws.cell(row=totals_row, column=c_idx, value=val)
            cell.number_format = FMT_CURRENCY
            cell.font = FONT_HEADER
        elif c_idx == 1:
            ws.cell(row=totals_row, column=c_idx, value="Total").font = FONT_HEADER

    format_money_columns(ws, money_cols)
    autosize_columns(ws, max_rows=200)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"


def _maybe_write_totals_by_salesman(
    wb: Workbook,
    invoices: pd.DataFrame,
    *,
    streaming: bool = False,
) -> None:
    """Add a 'Totals by Salesman' sheet if the invoices span multiple salesmen."""
    if invoices.empty or "Salesman" not in invoices.columns:
        return

    unique_salesmen = invoices["Salesman"].dropna().astype(str).str.strip()
    unique_salesmen = unique_salesmen[unique_salesmen != ""].unique()
    if len(unique_salesmen) < 2:
        return

    log.info("  Writing sheet: Totals by Salesman (%d salesmen)", len(unique_salesmen))

    group_cols = []
    for c in ["SalesmanNumber", "SalesmanName", "Salesman"]:
        if c in invoices.columns:
            group_cols.append(c)
    if not group_cols:
        return

    money_cols = ["SubTotal Invoices", "Tariff Charges", "Freight Charges",
                  "CC Charges", "Total Invoice"]
    existing_money = [c for c in money_cols if c in invoices.columns]

    agg_dict: dict[str, tuple[str, str]] = {
        "InvoiceCount": ("InvoiceNumber", "nunique"),
    }
    for mc in existing_money:
        agg_dict[mc] = (mc, "sum")

    totals = invoices.groupby(group_cols, dropna=False).agg(**agg_dict).reset_index()
    totals = totals.sort_values(group_cols[0], na_position="last")

    if streaming:
        _write_data_sheet_streaming(wb, "Totals by Salesman", totals)
    else:
        _write_data_sheet(wb, "Totals by Salesman", totals)


def _write_commissions_sheet(
    wb: Workbook,
    year: int,
    detail_df: pd.DataFrame,
    credits_df: pd.DataFrame,
    invoices_df: pd.DataFrame,
    pct_map: dict[str, float],
    *,
    current_month: int | None = None,
) -> None:
    """Write the Commissions sheet with monthly columns up to current_month."""
    import re
    from datetime import datetime as dt

    from openpyxl.utils import get_column_letter

    from reports.invoiced._commissions_styles import (
        COMMISSIONS_BLOCK_STYLE_MAP,
        COMMISSIONS_COL_WIDTHS,
        COMMISSIONS_TOP_MERGES,
        COMMISSIONS_TOP_STYLE_MAP,
        apply_commissions_style,
        build_commissions_styles,
    )

    ws = wb.create_sheet(title="Commissions")

    num_months = current_month if current_month and 1 <= current_month <= 12 else 12
    ytd_col = 4 + num_months  # column index for the YTD Total column

    for col_letter, width in COMMISSIONS_COL_WIDTHS.items():
        if width is not None:
            ws.column_dimensions[col_letter].width = width

    for rng in COMMISSIONS_TOP_MERGES:
        try:
            ws.merge_cells(rng)
        except Exception:
            pass

    built_styles = build_commissions_styles()

    for (r, c, sid) in COMMISSIONS_TOP_STYLE_MAP:
        apply_commissions_style(ws, r, c, sid, built_styles)

    ws["B2"].value = f"Commissions Summary ({year})"
    ws.row_dimensions[2].height = 18.0

    def _norm_smn(v) -> str:
        s = str(v).strip()
        if re.fullmatch(r"\d+", s or ""):
            return str(int(s))
        return s

    commissioned = [(k, v) for k, v in pct_map.items() if float(v or 0) > 0]
    commissioned.sort(key=lambda x: int(x[0]) if str(x[0]).isdigit() else str(x[0]))

    if not commissioned:
        ws.cell(row=4, column=2, value="No commissioned salesmen")
        return

    smn_to_name: dict[str, str] = {}
    smn_col_name = "SalesmanNumber" if "SalesmanNumber" in detail_df.columns else None
    name_col_name = "SalesmanName" if "SalesmanName" in detail_df.columns else None
    if smn_col_name and name_col_name:
        smn_series = detail_df[smn_col_name].map(_norm_smn)
        for smn, idx_group in smn_series.groupby(smn_series):
            nm = detail_df.loc[idx_group.index, name_col_name].astype(str).str.strip()
            nm = nm[nm != ""]
            if len(nm):
                smn_to_name[str(smn)] = nm.iloc[0]

    det = detail_df
    crd = credits_df
    inv = invoices_df

    def _coerce_month(series):
        return pd.to_datetime(series, errors="coerce").dt.month

    date_col_map = {}
    for label, df in [("det", det), ("crd", crd), ("inv", inv)]:
        for c in ["InvoiceDate", "Invoice Date", "InvoiceDate1", "Date"]:
            if c in df.columns:
                date_col_map[label] = c
                break

    if "det" in date_col_map:
        det["_month"] = _coerce_month(det[date_col_map["det"]])
    else:
        det["_month"] = pd.NA
    if "crd" in date_col_map:
        crd["_month"] = _coerce_month(crd[date_col_map["crd"]])
    else:
        crd["_month"] = pd.NA
    if "inv" in date_col_map:
        inv["_month"] = _coerce_month(inv[date_col_map["inv"]])
    else:
        inv["_month"] = pd.NA

    def _sum_month(df_src, smn_key, col):
        if df_src is None or df_src.empty or col not in df_src.columns:
            return [0.0] * num_months
        smn_c = "SalesmanNumber" if "SalesmanNumber" in df_src.columns else None
        if not smn_c:
            return [0.0] * num_months
        mask = df_src[smn_c].map(_norm_smn) == _norm_smn(smn_key)
        filtered = df_src.loc[mask]
        if filtered.empty:
            return [0.0] * num_months
        g = filtered.groupby("_month")[col].sum()
        return [float(g.get(m, 0.0) or 0.0) for m in range(1, num_months + 1)]

    def _find_col(df_src, candidates):
        if df_src is None or df_src.empty:
            return None
        norm = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in df_src.columns}
        for cand in candidates:
            key = cand.strip().lower().replace(" ", "").replace("_", "")
            if key in norm:
                return norm[key]
        return None

    _row_style_map: dict[int, list[tuple[int, int]]] = {}
    for rr, c, sid in COMMISSIONS_BLOCK_STYLE_MAP:
        _row_style_map.setdefault(int(rr), []).append((int(c), int(sid)))

    def _apply_style_row(dst_abs_row, template_rel_row):
        for c, sid in _row_style_map.get(int(template_rel_row), []):
            apply_commissions_style(ws, dst_abs_row, c, sid, built_styles)

    start_row = 4
    block_height = 11

    for idx, (smn_key, pct) in enumerate(commissioned):
        r0 = start_row + idx * block_height
        smn = _norm_smn(smn_key)
        name = smn_to_name.get(smn, "")

        _apply_style_row(r0 + 0, 0)
        for k in range(1, 7):
            _apply_style_row(r0 + k, 1)
        _apply_style_row(r0 + 7, 6)
        _apply_style_row(r0 + 8, 7)
        _apply_style_row(r0 + 9, 8)
        _apply_style_row(r0 + 10, 9)

        ws.row_dimensions[r0 + 8].height = 15.0
        ws.row_dimensions[r0 + 9].height = 15.0

        ws.cell(r0, 2).value = neutralize_excel_value(f"{smn} - {name}".strip(" -"))
        ws.cell(r0, ytd_col).value = "YTD Total"

        for mi in range(1, num_months + 1):
            ws.cell(r0, 3 + mi).value = dt(year, mi, 1)

        labels = [
            "SubTotal Invoices:",
            "Total Tariff Charges:",
            "Total Freight Charges:",
            "Total CC Charges:",
            "Total Invoices: (SubTotal+Tariff+Freight+CC)",
            "Total Credits:",
            "Net Commission Amount (Less Freight and CC)",
            "Commission:",
        ]
        for i, lab in enumerate(labels, start=1):
            ws.cell(r0 + i, 2).value = lab

        sub_row = r0 + 1
        tar_row = r0 + 2
        fre_row = r0 + 3
        cc_row = r0 + 4
        totinv_row = r0 + 5
        cred_row = r0 + 6
        net_row = r0 + 7
        comm_row = r0 + 8
        pay_row = r0 + 9

        pct_cell = ws.cell(comm_row, 3)
        pct_cell.value = float(pct or 0)
        pct_cell.number_format = '0.00%'

        subtotal_col = _find_col(inv, ["SubTotal Invoices", "SubTotal", "Subtotal", "Sub Total", "SubTotalAmount"])
        tariff_col = _find_col(det, ["Tariff Charges", "Tariff", "TariffCharge"])
        freight_col = _find_col(det, ["Freight Charges", "Freight", "FreightCharge"])
        cc_col = _find_col(det, ["CC Charges", "CC", "Credit Card Charges", "CCCharge"])
        credits_col = _find_col(crd, ["Total Invoice", "Total Invoices", "Total", "Amount"])

        subtotal = _sum_month(inv, smn, subtotal_col) if subtotal_col else [0.0] * num_months
        tariff = _sum_month(det, smn, tariff_col) if tariff_col else [0.0] * num_months
        freight = _sum_month(det, smn, freight_col) if freight_col else [0.0] * num_months
        cc = _sum_month(det, smn, cc_col) if cc_col else [0.0] * num_months
        credits_vals = _sum_month(crd, smn, credits_col) if credits_col else [0.0] * num_months

        last_data_col = 3 + num_months  # last month data column
        first_data_cl = get_column_letter(4)
        last_data_cl = get_column_letter(last_data_col)
        ytd_cl = get_column_letter(ytd_col)

        base_rows = {
            sub_row: subtotal,
            tar_row: tariff,
            fre_row: freight,
            cc_row: cc,
            cred_row: credits_vals,
        }
        for rr, vals in base_rows.items():
            for mi in range(num_months):
                ws.cell(rr, 4 + mi).value = float(vals[mi] or 0.0)
            ws.cell(rr, ytd_col).value = sum(float(v or 0.0) for v in vals)

        totinv_vals = [
            float(subtotal[mi] or 0) + float(tariff[mi] or 0)
            + float(freight[mi] or 0) + float(cc[mi] or 0)
            for mi in range(num_months)
        ]
        for mi in range(num_months):
            ws.cell(totinv_row, 4 + mi).value = totinv_vals[mi]
        ws.cell(totinv_row, ytd_col).value = sum(totinv_vals)

        net_vals = [
            totinv_vals[mi] + float(credits_vals[mi] or 0)
            - float(freight[mi] or 0) - float(cc[mi] or 0)
            for mi in range(num_months)
        ]
        for mi in range(num_months):
            ws.cell(net_row, 4 + mi).value = net_vals[mi]
        ws.cell(net_row, ytd_col).value = sum(net_vals)

        pct_float = float(pct or 0)
        comm_vals = [net_vals[mi] * pct_float for mi in range(num_months)]
        for mi in range(num_months):
            ws.cell(comm_row, 4 + mi).value = comm_vals[mi]
        ws.cell(comm_row, ytd_col).value = sum(comm_vals)

        ws.cell(pay_row, 2).value = f"Total Payable: {smn} - {name}".strip(" -")
        ws.cell(pay_row, ytd_col).value = sum(comm_vals)


# ---------------------------------------------------------------------------
# Streaming-mode sheet writers (write_only=True, for large reports)
# ---------------------------------------------------------------------------

def _write_data_sheet_streaming(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    is_first: bool = False,
) -> None:
    """Write a data sheet using write_only append mode.

    Rows are flushed to a temp file immediately after append, so memory
    stays roughly proportional to one row at a time instead of the entire
    sheet.  Currency formatting is applied inline during cell creation rather
    than as a post-hoc pass over all cells.
    """
    ws = wb.create_sheet(title=title)

    if df.empty:
        ws.append(["No data"])
        return

    strip_datetime_tz(df)
    cols = [c for c in df.columns if not c.startswith("_")]
    col_indices = [df.columns.get_loc(c) for c in cols]
    money_flags = [_is_money_column(c) for c in cols]

    max_widths: dict[int, int] = {i: min(len(str(c)), 60) for i, c in enumerate(cols)}
    _WIDTH_SAMPLE = 200

    header_cells = []
    for col in cols:
        cell = WriteOnlyCell(ws, value=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_BLUE
        cell.border = BORDER_THIN
        header_cells.append(cell)
    ws.append(header_cells)

    data = df.values
    num_rows = len(data)
    for r_idx in range(num_rows):
        row_arr = data[r_idx]
        row_cells = []
        for c_idx, ci in enumerate(col_indices):
            v = row_arr[ci]
            if v is None:
                v = ""
            elif isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                v = 0.0
            cell = WriteOnlyCell(ws, value=neutralize_excel_value(v))
            if money_flags[c_idx] and isinstance(v, (int, float)):
                cell.number_format = FMT_CURRENCY
            row_cells.append(cell)
        ws.append(row_cells)

        if r_idx < _WIDTH_SAMPLE:
            for c_idx, ci in enumerate(col_indices):
                v = row_arr[ci]
                if v is not None:
                    max_widths[c_idx] = max(max_widths.get(c_idx, 0), min(len(str(v)), 60))

    totals_cells = []
    for c_idx, col in enumerate(cols):
        if money_flags[c_idx] and col in df.columns:
            val = pd.to_numeric(df[col], errors="coerce").fillna(0).sum()
            cell = WriteOnlyCell(ws, value=val)
            cell.number_format = FMT_CURRENCY
            cell.font = FONT_HEADER
            totals_cells.append(cell)
        elif c_idx == 0:
            cell = WriteOnlyCell(ws, value="Total")
            cell.font = FONT_HEADER
            totals_cells.append(cell)
        else:
            totals_cells.append(WriteOnlyCell(ws, value=None))
    ws.append(totals_cells)

    for c_idx, w in max_widths.items():
        ws.column_dimensions[get_column_letter(c_idx + 1)].width = max(10, w + 2)

    ws.freeze_panes = "A2"
    total_excel_rows = num_rows + 2
    if total_excel_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{total_excel_rows}"


def _write_commissions_sheet_simple_streaming(wb: Workbook, comm_df: pd.DataFrame) -> None:
    """Streaming fallback: simple commissions sheet (write_only mode)."""
    ws = wb.create_sheet(title="Commissions")
    if comm_df.empty or "SalesmanNumber" not in comm_df.columns:
        ws.append(["No commission data"])
        return
    commissioned = comm_df[comm_df["Percent"] > 0].copy()
    if commissioned.empty:
        ws.append(["No commissioned salesmen"])
        return

    title_cell = WriteOnlyCell(ws, value="Commissions Summary")
    title_cell.font = Font(bold=True, size=14)
    ws.append([title_cell])
    ws.append([])

    for sm_num, grp in commissioned.groupby("SalesmanNumber", sort=True):
        sm_name = grp["SalesmanName"].iloc[0] if "SalesmanName" in grp.columns else ""
        pct = grp["Percent"].iloc[0]
        hdr = WriteOnlyCell(ws, value=neutralize_excel_value(f"Salesman #{sm_num} - {sm_name}"))
        hdr.font = Font(bold=True, size=11)
        ws.append([hdr, None, f"{pct:.0%}"])

        sub_total = grp["SubTotal Invoices"].sum() if "SubTotal Invoices" in grp.columns else 0
        tariff_total = grp["Total Tariff Charges"].sum() if "Total Tariff Charges" in grp.columns else 0
        freight_total = grp["Total Freight Charges"].sum() if "Total Freight Charges" in grp.columns else 0
        cc_total = grp["Total CC Charges"].sum() if "Total CC Charges" in grp.columns else 0
        inv_total = grp["Total Invoices"].sum() if "Total Invoices" in grp.columns else 0
        net = sub_total + tariff_total
        commission = net * pct
        for label, val in [("SubTotal Invoices", sub_total), ("Total Tariff Charges", tariff_total),
                           ("Total Freight Charges", freight_total), ("Total CC Charges", cc_total),
                           ("Total Invoices", inv_total), ("Net Commission Amount", net), ("Commission", commission)]:
            val_cell = WriteOnlyCell(ws, value=val)
            val_cell.number_format = FMT_CURRENCY
            ws.append([label, val_cell])
        ws.append([])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def _write_commissions_sheet_streaming(
    wb: Workbook,
    year: int,
    detail_df: pd.DataFrame,
    credits_df: pd.DataFrame,
    invoices_df: pd.DataFrame,
    pct_map: dict[str, float],
    *,
    current_month: int | None = None,
) -> None:
    """Write the Commissions sheet in write_only mode.

    All data and styling are identical to the normal-mode version.
    The only difference is merge_cells are skipped (purely cosmetic)
    and rows are written via ws.append() with WriteOnlyCell objects.
    """
    import re
    from datetime import datetime as dt

    from reports.invoiced._commissions_styles import (
        COMMISSIONS_BLOCK_STYLE_MAP,
        COMMISSIONS_COL_WIDTHS,
        COMMISSIONS_TOP_STYLE_MAP,
        build_commissions_styles,
    )

    ws = wb.create_sheet(title="Commissions")

    num_months = current_month if current_month and 1 <= current_month <= 12 else 12
    ytd_col = 4 + num_months
    total_cols = max(16, ytd_col)

    for col_letter, width in COMMISSIONS_COL_WIDTHS.items():
        if width is not None:
            ws.column_dimensions[col_letter].width = width

    built_styles = build_commissions_styles()

    top_style_map: dict[tuple[int, int], int] = {}
    for r, c, sid in COMMISSIONS_TOP_STYLE_MAP:
        top_style_map[(r, c)] = sid

    _row_style_map: dict[int, dict[int, int]] = {}
    for rr, c, sid in COMMISSIONS_BLOCK_STYLE_MAP:
        _row_style_map.setdefault(int(rr), {})[int(c)] = int(sid)

    def _styled_cell(value, style_id):
        st = built_styles[style_id]
        cell = WriteOnlyCell(ws, value=neutralize_excel_value(value))
        cell.font = st["font"]
        cell.fill = st["fill"]
        cell.border = st["border"]
        cell.alignment = st["alignment"]
        cell.number_format = st["number_format"]
        cell.protection = st["protection"]
        return cell

    def _top_row(row_num, col_values=None):
        col_values = col_values or {}
        return [
            _styled_cell(col_values.get(c), top_style_map.get((row_num, c), 0))
            for c in range(1, total_cols + 1)
        ]

    def _block_row(template_rel, col_values=None):
        col_values = col_values or {}
        style_entries = _row_style_map.get(template_rel, {})
        return [
            _styled_cell(col_values.get(c), style_entries.get(c, 0))
            for c in range(1, total_cols + 1)
        ]

    # --- Top 3 rows ---
    ws.append(_top_row(1))
    ws.append(_top_row(2, {2: f"Commissions Summary ({year})"}))
    ws.append(_top_row(3))

    # --- Data preparation (same logic as normal-mode) ---
    def _norm_smn(v) -> str:
        s = str(v).strip()
        if re.fullmatch(r"\d+", s or ""):
            return str(int(s))
        return s

    commissioned = [(k, v) for k, v in pct_map.items() if float(v or 0) > 0]
    commissioned.sort(key=lambda x: int(x[0]) if str(x[0]).isdigit() else str(x[0]))

    if not commissioned:
        ws.append(_block_row(1, {2: "No commissioned salesmen"}))
        return

    smn_to_name: dict[str, str] = {}
    smn_col_name = "SalesmanNumber" if "SalesmanNumber" in detail_df.columns else None
    name_col_name = "SalesmanName" if "SalesmanName" in detail_df.columns else None
    if smn_col_name and name_col_name:
        smn_series = detail_df[smn_col_name].map(_norm_smn)
        for smn, idx_group in smn_series.groupby(smn_series):
            nm = detail_df.loc[idx_group.index, name_col_name].astype(str).str.strip()
            nm = nm[nm != ""]
            if len(nm):
                smn_to_name[str(smn)] = nm.iloc[0]

    det, crd, inv = detail_df, credits_df, invoices_df

    def _coerce_month(series):
        return pd.to_datetime(series, errors="coerce").dt.month

    date_col_map: dict[str, str] = {}
    for label, df in [("det", det), ("crd", crd), ("inv", inv)]:
        for c in ["InvoiceDate", "Invoice Date", "InvoiceDate1", "Date"]:
            if c in df.columns:
                date_col_map[label] = c
                break

    if "det" in date_col_map:
        det["_month"] = _coerce_month(det[date_col_map["det"]])
    else:
        det["_month"] = pd.NA
    if "crd" in date_col_map:
        crd["_month"] = _coerce_month(crd[date_col_map["crd"]])
    else:
        crd["_month"] = pd.NA
    if "inv" in date_col_map:
        inv["_month"] = _coerce_month(inv[date_col_map["inv"]])
    else:
        inv["_month"] = pd.NA

    def _sum_month(df_src, smn_key, col):
        if df_src is None or df_src.empty or col not in df_src.columns:
            return [0.0] * num_months
        smn_c = "SalesmanNumber" if "SalesmanNumber" in df_src.columns else None
        if not smn_c:
            return [0.0] * num_months
        mask = df_src[smn_c].map(_norm_smn) == _norm_smn(smn_key)
        filtered = df_src.loc[mask]
        if filtered.empty:
            return [0.0] * num_months
        g = filtered.groupby("_month")[col].sum()
        return [float(g.get(m, 0.0) or 0.0) for m in range(1, num_months + 1)]

    def _find_col(df_src, candidates):
        if df_src is None or df_src.empty:
            return None
        norm = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in df_src.columns}
        for cand in candidates:
            key = cand.strip().lower().replace(" ", "").replace("_", "")
            if key in norm:
                return norm[key]
        return None

    def _data_vals(label, vals):
        v: dict[int, object] = {2: label}
        for mi in range(num_months):
            v[4 + mi] = float(vals[mi] or 0.0)
        v[ytd_col] = sum(float(x or 0.0) for x in vals)
        return v

    # --- Write salesman blocks ---
    for idx, (smn_key, pct) in enumerate(commissioned):
        smn = _norm_smn(smn_key)
        name = smn_to_name.get(smn, "")

        # Row 0: header
        header_vals: dict[int, object] = {
            2: f"{smn} - {name}".strip(" -"),
            ytd_col: "YTD Total",
        }
        for mi in range(1, num_months + 1):
            header_vals[3 + mi] = dt(year, mi, 1)
        ws.append(_block_row(0, header_vals))

        # Compute monthly data for this salesman
        subtotal_col = _find_col(inv, ["SubTotal Invoices", "SubTotal", "Subtotal", "Sub Total", "SubTotalAmount"])
        tariff_col = _find_col(det, ["Tariff Charges", "Tariff", "TariffCharge"])
        freight_col = _find_col(det, ["Freight Charges", "Freight", "FreightCharge"])
        cc_col = _find_col(det, ["CC Charges", "CC", "Credit Card Charges", "CCCharge"])
        credits_col = _find_col(crd, ["Total Invoice", "Total Invoices", "Total", "Amount"])

        subtotal = _sum_month(inv, smn, subtotal_col) if subtotal_col else [0.0] * num_months
        tariff = _sum_month(det, smn, tariff_col) if tariff_col else [0.0] * num_months
        freight = _sum_month(det, smn, freight_col) if freight_col else [0.0] * num_months
        cc = _sum_month(det, smn, cc_col) if cc_col else [0.0] * num_months
        credits_vals = _sum_month(crd, smn, credits_col) if credits_col else [0.0] * num_months

        totinv_vals = [
            float(subtotal[mi] or 0) + float(tariff[mi] or 0)
            + float(freight[mi] or 0) + float(cc[mi] or 0)
            for mi in range(num_months)
        ]
        net_vals = [
            totinv_vals[mi] + float(credits_vals[mi] or 0)
            - float(freight[mi] or 0) - float(cc[mi] or 0)
            for mi in range(num_months)
        ]
        pct_float = float(pct or 0)
        comm_vals = [net_vals[mi] * pct_float for mi in range(num_months)]

        # Rows 1-6: data rows (all use template 1)
        ws.append(_block_row(1, _data_vals("SubTotal Invoices:", subtotal)))
        ws.append(_block_row(1, _data_vals("Total Tariff Charges:", tariff)))
        ws.append(_block_row(1, _data_vals("Total Freight Charges:", freight)))
        ws.append(_block_row(1, _data_vals("Total CC Charges:", cc)))
        ws.append(_block_row(1, _data_vals("Total Invoices: (SubTotal+Tariff+Freight+CC)", totinv_vals)))
        ws.append(_block_row(1, _data_vals("Total Credits:", credits_vals)))

        # Row 7: Net Commission Amount (template 6)
        ws.append(_block_row(6, _data_vals("Net Commission Amount (Less Freight and CC)", net_vals)))

        # Row 8: Commission (template 7) -- override col C to percentage format
        comm_row_vals: dict[int, object] = {2: "Commission:", 3: pct_float}
        for mi in range(num_months):
            comm_row_vals[4 + mi] = comm_vals[mi]
        comm_row_vals[ytd_col] = sum(comm_vals)
        comm_row_cells = _block_row(7, comm_row_vals)
        comm_row_cells[2].number_format = '0.00%'
        ws.append(comm_row_cells)

        # Row 9: Total Payable (template 8)
        ws.append(_block_row(8, {
            2: f"Total Payable: {smn} - {name}".strip(" -"),
            ytd_col: sum(comm_vals),
        }))

        # Row 10: separator (template 9)
        ws.append(_block_row(9))
