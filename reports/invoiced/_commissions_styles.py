"""
Commissions sheet styling infrastructure.

Ported from the legacy invoiced_master_pipeline.py to match the exact
Commissions Format.xlsx layout: per-salesman blocks with Jan-Dec + YTD columns.
"""

from datetime import datetime

from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side

# fmt: off
COMMISSIONS_STYLE_DEFS = [
    # 0: default (Arial Narrow 10, no border, General)
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 1: Aptos Narrow 11, no border, General
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Aptos Narrow", "sz": 11.0, "bold": False, "italic": False, "underline": None, "color": "00000000", "family": 2.0, "scheme": "minor"},
     "locked": True, "number_format": "General"},
    # 2: Arial Narrow 10, accounting currency
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'},
    # 3: Arial Narrow 14 Bold (title)
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Arial Narrow", "sz": 14.0, "bold": True, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 4: bottom thin border, General
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": "thin", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 5: bottom thin border, Aptos Narrow 11
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": "thin", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Aptos Narrow", "sz": 11.0, "bold": False, "italic": False, "underline": None, "color": "00000000", "family": 2.0, "scheme": "minor"},
     "locked": True, "number_format": "General"},
    # 6: blue header, bold white, left+top thin, General
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": "thin", "left_color": "00000000", "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FF5B9BD5", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": True, "italic": False, "underline": None, "color": "FFFFFFFF", "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 7: blue header, bold white, top thin, General
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FF5B9BD5", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": True, "italic": False, "underline": None, "color": "FFFFFFFF", "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 8: blue header, bold white, top thin, mmm-yy (month date format)
    {"alignment": {"horizontal": "center", "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FF5B9BD5", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": True, "italic": False, "underline": None, "color": "FFFFFFFF", "family": 2.0, "scheme": None},
     "locked": True, "number_format": "mmm-yy"},
    # 9: blue header, bold white, center, wrapText, right+top thin, General (YTD Total col)
    {"alignment": {"horizontal": "center", "vertical": "center", "wrapText": True, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": "thin", "right_color": "00000000", "top_style": "thin", "top_color": "00000000", "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FF5B9BD5", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": True, "italic": False, "underline": None, "color": "FFFFFFFF", "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 10: left aligned, left thin border, data row label
    {"alignment": {"horizontal": "left", "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": "thin", "left_color": "00000000", "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 11: left aligned, no border
    {"alignment": {"horizontal": "left", "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 12: right thin border, accounting currency
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": "thin", "right_color": "00000000", "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'},
    # 13: Aptos Narrow 11, accounting currency
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": None, "top_color": None, "bottom_style": None, "bottom_color": None, "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": None, "fgColor": "00000000", "bgColor": "00000000"},
     "font": {"name": "Aptos Narrow", "sz": 11.0, "bold": False, "italic": False, "underline": None, "color": "00000000", "family": 2.0, "scheme": "minor"},
     "locked": True, "number_format": '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'},
    # 14: net row: bold, left+top+bottom-medium thin, grey fill, left aligned
    {"alignment": {"horizontal": "left", "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": "thin", "left_color": "00000000", "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFEBEEF1", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": True, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 15: net row: bold, center, top+bottom-medium thin, grey fill, 0% format
    {"alignment": {"horizontal": "center", "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFEBEEF1", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": True, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "0%"},
    # 16: net row: bold, top+bottom-medium thin, grey fill, accounting currency
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFEBEEF1", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": True, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'},
    # 17: commission row: yellow fill, left+bottom-medium+top thin
    {"alignment": {"horizontal": "left", "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": "thin", "left_color": "00000000", "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 18: commission row: yellow fill, bottom-medium+top thin
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 19: commission row: yellow fill, bottom-medium+top thin, red-parens currency
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'},
    # 20: payable row: yellow fill, left+bottom-medium+top thin, left indent 2
    {"alignment": {"horizontal": "left", "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 2.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 21: payable row: yellow fill, bottom-medium+top thin, General
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 22: payable row first data col: yellow, accounting currency
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'},
    # 23: payable row col E special: yellow
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 24: payable row col O: yellow, right thin border
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": "thin", "right_color": "00000000", "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": "General"},
    # 25: payable row col P: yellow, accounting currency, bottom-medium+top thin
    {"alignment": {"horizontal": None, "vertical": None, "wrapText": None, "shrinkToFit": None, "textRotation": 0, "indent": 0.0, "relativeIndent": 0.0, "justifyLastLine": None, "readingOrder": 0.0},
     "border": {"left_style": None, "left_color": None, "right_style": None, "right_color": None, "top_style": "thin", "top_color": "00000000", "bottom_style": "medium", "bottom_color": "00000000", "diagDown": False, "diagUp": False, "diag_style": None, "diag_color": None},
     "fill": {"patternType": "solid", "fgColor": "FFFFFF00", "bgColor": "FF000000"},
     "font": {"name": "Arial Narrow", "sz": 10.0, "bold": False, "italic": False, "underline": None, "color": None, "family": 2.0, "scheme": None},
     "locked": True, "number_format": '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'},
]
# fmt: on

COMMISSIONS_TOP_STYLE_MAP = [
    (1, 1, 0), (1, 2, 1), (1, 3, 0), (1, 4, 1), (1, 5, 0), (1, 6, 0), (1, 7, 0), (1, 8, 0),
    (1, 9, 0), (1, 10, 0), (1, 11, 0), (1, 12, 0), (1, 13, 0), (1, 14, 0), (1, 15, 2), (1, 16, 0),
    (2, 1, 0), (2, 2, 3), (2, 3, 0), (2, 4, 1), (2, 5, 0), (2, 6, 0), (2, 7, 0), (2, 8, 0),
    (2, 9, 0), (2, 10, 0), (2, 11, 0), (2, 12, 0), (2, 13, 0), (2, 14, 0), (2, 15, 2), (2, 16, 0),
    (3, 1, 0), (3, 2, 1), (3, 3, 4), (3, 4, 5), (3, 5, 0), (3, 6, 0), (3, 7, 0), (3, 8, 0),
    (3, 9, 0), (3, 10, 0), (3, 11, 0), (3, 12, 0), (3, 13, 0), (3, 14, 0), (3, 15, 2), (3, 16, 0),
]

COMMISSIONS_BLOCK_STYLE_MAP = [
    (0, 1, 1), (0, 2, 6), (0, 3, 7), (0, 4, 8), (0, 5, 8), (0, 6, 8), (0, 7, 8), (0, 8, 8),
    (0, 9, 8), (0, 10, 8), (0, 11, 8), (0, 12, 8), (0, 13, 8), (0, 14, 8), (0, 15, 8), (0, 16, 9),
    (1, 1, 1), (1, 2, 10), (1, 3, 11), (1, 4, 2), (1, 5, 2), (1, 6, 2), (1, 7, 2), (1, 8, 2),
    (1, 9, 2), (1, 10, 2), (1, 11, 2), (1, 12, 2), (1, 13, 2), (1, 14, 2), (1, 15, 2), (1, 16, 12),
    (2, 1, 1), (2, 2, 10), (2, 3, 11), (2, 4, 2), (2, 5, 2), (2, 6, 2), (2, 7, 2), (2, 8, 2),
    (2, 9, 2), (2, 10, 2), (2, 11, 2), (2, 12, 2), (2, 13, 2), (2, 14, 2), (2, 15, 13), (2, 16, 12),
    (3, 1, 1), (3, 2, 10), (3, 3, 11), (3, 4, 2), (3, 5, 2), (3, 6, 2), (3, 7, 2), (3, 8, 2),
    (3, 9, 2), (3, 10, 2), (3, 11, 2), (3, 12, 2), (3, 13, 2), (3, 14, 2), (3, 15, 2), (3, 16, 12),
    (4, 1, 1), (4, 2, 10), (4, 3, 11), (4, 4, 2), (4, 5, 2), (4, 6, 2), (4, 7, 2), (4, 8, 2),
    (4, 9, 2), (4, 10, 2), (4, 11, 2), (4, 12, 2), (4, 13, 2), (4, 14, 2), (4, 15, 2), (4, 16, 12),
    (5, 1, 1), (5, 2, 10), (5, 3, 11), (5, 4, 2), (5, 5, 2), (5, 6, 2), (5, 7, 2), (5, 8, 2),
    (5, 9, 2), (5, 10, 2), (5, 11, 2), (5, 12, 2), (5, 13, 2), (5, 14, 2), (5, 15, 2), (5, 16, 12),
    (6, 1, 1), (6, 2, 14), (6, 3, 15), (6, 4, 16), (6, 5, 16), (6, 6, 16), (6, 7, 16), (6, 8, 16),
    (6, 9, 16), (6, 10, 16), (6, 11, 16), (6, 12, 16), (6, 13, 16), (6, 14, 16), (6, 15, 16), (6, 16, 12),
    (7, 1, 1), (7, 2, 17), (7, 3, 18), (7, 4, 19), (7, 5, 19), (7, 6, 19), (7, 7, 19), (7, 8, 19),
    (7, 9, 19), (7, 10, 19), (7, 11, 19), (7, 12, 19), (7, 13, 19), (7, 14, 19), (7, 15, 19), (7, 16, 19),
    (8, 1, 1), (8, 2, 20), (8, 3, 21), (8, 4, 22), (8, 5, 23), (8, 6, 21), (8, 7, 21), (8, 8, 21),
    (8, 9, 21), (8, 10, 21), (8, 11, 21), (8, 12, 21), (8, 13, 21), (8, 14, 21), (8, 15, 24), (8, 16, 25),
    (9, 1, 1), (9, 2, 1), (9, 3, 1), (9, 4, 1), (9, 5, 1), (9, 6, 1), (9, 7, 1), (9, 8, 1),
    (9, 9, 1), (9, 10, 1), (9, 11, 1), (9, 12, 1), (9, 13, 1), (9, 14, 1), (9, 15, 1), (9, 16, 1),
]

COMMISSIONS_COL_WIDTHS = {
    "A": 13.0, "B": 48.0, "C": 3.0, "D": 15.0, "E": 13.0, "F": 13.0,
    "G": 13.0, "H": 13.0, "I": 13.0, "J": 13.0, "K": 13.0, "L": 13.0,
    "M": 13.0, "N": 13.0, "O": 15.0, "P": 15.0,
}

COMMISSIONS_TOP_MERGES = ["C2:D2", "C1:D1", "A1:B1", "C3:D3", "A3:B3"]


def build_commissions_styles():
    """Rebuild openpyxl style objects from COMMISSIONS_STYLE_DEFS."""
    built = []
    for d in COMMISSIONS_STYLE_DEFS:
        f = d["font"]
        f_color = Color(rgb=f["color"]) if f.get("color") else None
        font = Font(
            name=f.get("name"), sz=f.get("sz"), bold=f.get("bold"),
            italic=f.get("italic"), underline=f.get("underline"),
            color=f_color, family=f.get("family"), scheme=f.get("scheme"),
        )

        fl = d["fill"]
        if fl.get("patternType") is None:
            fill = PatternFill()
        else:
            fill = PatternFill(
                patternType=fl["patternType"],
                fgColor=fl.get("fgColor"), bgColor=fl.get("bgColor"),
            )

        bd = d["border"]

        def _side(style, rgb):
            if style is None and rgb is None:
                return Side()
            return Side(style=style, color=Color(rgb=rgb) if rgb else None)

        border = Border(
            left=_side(bd.get("left_style"), bd.get("left_color")),
            right=_side(bd.get("right_style"), bd.get("right_color")),
            top=_side(bd.get("top_style"), bd.get("top_color")),
            bottom=_side(bd.get("bottom_style"), bd.get("bottom_color")),
            diagonal=_side(bd.get("diag_style"), bd.get("diag_color")),
            diagonalUp=bd.get("diagUp", False),
            diagonalDown=bd.get("diagDown", False),
        )

        al = d["alignment"]
        alignment = Alignment(
            horizontal=al.get("horizontal"), vertical=al.get("vertical"),
            textRotation=al.get("textRotation"), wrapText=al.get("wrapText"),
            shrinkToFit=al.get("shrinkToFit"), indent=al.get("indent"),
            relativeIndent=al.get("relativeIndent"),
            justifyLastLine=al.get("justifyLastLine"),
            readingOrder=al.get("readingOrder"),
        )

        protection = Protection(locked=d.get("locked", True))
        built.append({
            "font": font, "fill": fill, "border": border,
            "alignment": alignment,
            "number_format": d.get("number_format", "General"),
            "protection": protection,
        })
    return built


def apply_commissions_style(ws, r, c, style_id, built_styles):
    """Apply a pre-built style to a cell."""
    st = built_styles[style_id]
    cell = ws.cell(row=r, column=c)
    cell.font = st["font"]
    cell.fill = st["fill"]
    cell.border = st["border"]
    cell.alignment = st["alignment"]
    cell.number_format = st["number_format"]
    cell.protection = st["protection"]
