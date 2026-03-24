"""
Customer Aging Report Excel writer.

Produces nicely formatted workbooks with:
  - Color-coded aging buckets (green -> yellow -> orange -> red)
  - Currency formatting on money columns
  - Date formatting on payment date
  - Totals row per sheet
  - Auto-sized columns with filters
"""

import logging
import math
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.excel_styles import (
    BORDER_THIN,
    FILL_HEADER_BLUE,
    FILL_TOTALS,
    FMT_CURRENCY,
    FMT_DATE,
    FONT_HEADER,
)
from core.excel_writer import autosize_columns, strip_datetime_tz
from reports.customer_aging.builder import DISPLAY_HEADERS, REPORT_COLUMNS

log = logging.getLogger(__name__)

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

FONT_RED = Font(color="FF0000")
FONT_DARK_RED = Font(bold=True, color="9C0006")

MONEY_COLS = {"AmountDue", "LastPaymentAmount", "Current", "30", "60", "90", "91+"}
DATE_COLS = {"LastPaymentDate"}
AGING_BUCKET_COLS = {"Current": FILL_GREEN, "30": FILL_YELLOW, "60": FILL_ORANGE, "90": FILL_RED, "91+": FILL_RED}

COL_WIDTHS = {
    "CustomerAccount": 12,
    "CustomerName": 35,
    "Salesman": 18,
    "AmountDue": 14,
    "LastPaymentDate": 16,
    "LastPaymentAmount": 18,
    "NumOpenInvoices": 16,
    "Current": 14,
    "30": 12,
    "60": 12,
    "90": 12,
    "91+": 12,
}


def _safe_val(v):
    """Return Excel-safe value."""
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, str) and v.strip().lower() in ("nan", "nat", ""):
        return None
    return v


def _write_sheet(ws, df: pd.DataFrame, sheet_title: str) -> None:
    """Write a single aging sheet with headers, data, totals, and formatting."""
    df = strip_datetime_tz(df)

    # Header row
    for c_idx, header in enumerate(DISPLAY_HEADERS, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_BLUE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    ws.freeze_panes = "A2"

    # Column widths
    for c_idx, col_name in enumerate(REPORT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(col_name, 12)

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, col_name in enumerate(REPORT_COLUMNS, 1):
            val = _safe_val(row.get(col_name))
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.border = BORDER_THIN

            if col_name in MONEY_COLS:
                cell.number_format = FMT_CURRENCY
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = FONT_RED

            if col_name in DATE_COLS and val is not None:
                cell.number_format = FMT_DATE

            # Color-code aging buckets when they have a non-zero value
            if col_name in AGING_BUCKET_COLS:
                if isinstance(val, (int, float)) and abs(val) > 0.005:
                    cell.fill = AGING_BUCKET_COLS[col_name]

    # Totals row
    if not df.empty:
        total_row = len(df) + 2
        ws.cell(row=total_row, column=1, value="TOTALS")
        ws.cell(row=total_row, column=1).font = Font(bold=True)

        for c_idx, col_name in enumerate(REPORT_COLUMNS, 1):
            cell = ws.cell(row=total_row, column=c_idx)
            cell.fill = FILL_TOTALS
            cell.border = BORDER_THIN
            cell.font = Font(bold=True)

            if col_name in MONEY_COLS:
                total = df[col_name].sum() if col_name in df.columns else 0
                cell.value = round(total, 2)
                cell.number_format = FMT_CURRENCY
                if total < 0:
                    cell.font = FONT_DARK_RED
            elif col_name == "NumOpenInvoices":
                cell.value = int(df[col_name].sum()) if col_name in df.columns else 0
            elif col_name == "CustomerAccount":
                cell.value = f"TOTALS ({len(df)} customers)"
                cell.font = Font(bold=True)

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(REPORT_COLUMNS))}{total_row}"


def write_aging_master_workbook(
    sheets: dict[str, pd.DataFrame],
    out_path: str,
) -> None:
    """Write the master Customer Aging workbook with multiple sheets."""
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in sheets.items():
        safe_name = sheet_name[:31]
        ws = wb.create_sheet(safe_name)
        _write_sheet(ws, df, safe_name)

    wb.save(out_path)
    log.info("Wrote master aging workbook: %s (%d sheets)", out_path, len(sheets))


def write_aging_salesman_workbook(
    df: pd.DataFrame,
    salesman_name: str,
    out_path: str,
) -> None:
    """Write a single-sheet workbook for one salesman."""
    if df.empty:
        log.info("No aging data for salesman %s -- skipping", salesman_name)
        return

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = salesman_name[:31]
    _write_sheet(ws, df, salesman_name)
    wb.save(out_path)
    log.info("Wrote salesman aging workbook: %s", out_path)
