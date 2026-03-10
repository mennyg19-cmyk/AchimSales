"""
Shared styles and formatting for Number 4 Report writers (By Item and By Customer).

Uses WriteOnlyCell for streaming write_only mode.
"""

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_TOTALS = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_GRAND = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
FONT_BOLD = Font(bold=True)
FMT_QTY = "#,##0"
FMT_CURRENCY = "$#,##0.00"


def make_cell(ws, value, *, fill=None, bold=False, fmt=None):
    """Create a styled WriteOnlyCell for streaming writes."""
    cell = WriteOnlyCell(ws, value=value)
    cell.border = BORDER
    if fill:
        cell.fill = fill
    if bold:
        cell.font = FONT_BOLD
    if fmt:
        cell.number_format = fmt
    return cell


def styled_row(ws, values, qty_set, currency_set, *, fill=None, bold=False):
    """Build a list of styled WriteOnlyCells from a list of values.

    qty_set / currency_set are sets of 0-based column indices.
    """
    cells = []
    for i, v in enumerate(values):
        fmt = None
        if i in qty_set:
            fmt = FMT_QTY
        elif i in currency_set:
            fmt = FMT_CURRENCY
        cells.append(make_cell(ws, v, fill=fill, bold=bold, fmt=fmt))
    return cells
