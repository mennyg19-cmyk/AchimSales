"""
Number 4 Report - By Item Excel writer (streaming / write_only mode).

Writes sections per item with customer rows, TOTALS per item, and GRAND TOTALS.
Two sheets: "12 Months" and "Year to Date". Same trailing columns as By Customer:
month qty and $, then Total Qty, Total $, Avg Price, Book Price, Salesman.
"""

import logging
import os

import pandas as pd
from openpyxl import Workbook

from reports.number_4._styles import (
    FILL_GRAND,
    FILL_HEADER,
    FILL_TOTALS,
    styled_row,
)

log = logging.getLogger(__name__)


def write_by_item(
    agg_12: pd.DataFrame,
    labels_12: list[tuple[str, str]],
    qty_cols_12: list[str],
    dol_cols_12: list[str],
    agg_ytd: pd.DataFrame,
    labels_ytd: list[tuple[str, str]],
    qty_cols_ytd: list[str],
    dol_cols_ytd: list[str],
    out_path: str,
) -> None:
    """Write By Item workbook with 12 Months and Year to Date sheets."""
    log.info("Writing Number 4 By Item: 12mo=%d rows, YTD=%d rows", len(agg_12), len(agg_ytd))
    wb = Workbook(write_only=True)

    ws_12 = wb.create_sheet("12 Months")
    log.info("  Writing sheet: 12 Months (%d rows)", len(agg_12))
    _write_sheet(ws_12, agg_12, labels_12, qty_cols_12, dol_cols_12)

    ws_ytd = wb.create_sheet("Year to Date")
    log.info("  Writing sheet: Year to Date (%d rows)", len(agg_ytd))
    _write_sheet(ws_ytd, agg_ytd, labels_ytd, qty_cols_ytd, dol_cols_ytd)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    try:
        size_kb = os.path.getsize(out_path) / 1024
        log.info("Excel saved: %.0f KB -- %s", size_kb, out_path)
    except OSError:
        log.info("Wrote: %s", out_path)


def _write_sheet(ws, agg_df, month_labels, qty_cols, dol_cols):
    if agg_df.empty:
        ws.append(["No data"])
        return

    headers = ["Item #", "Item Name", "Customer #", "Customer Name"]
    for lbl_q, lbl_d in month_labels:
        headers.append(lbl_q)
        headers.append(lbl_d)
    headers.extend(["Total Qty", "Total $", "Avg Price", "Book Price", "Salesman"])
    n_month = len(qty_cols)

    qty_set = set()
    currency_set = set()
    for i in range(n_month):
        qty_set.add(4 + i * 2)
        currency_set.add(4 + i * 2 + 1)
    summary_start = 4 + n_month * 2
    qty_set.add(summary_start)
    currency_set.add(summary_start + 1)
    currency_set.add(summary_start + 2)
    currency_set.add(summary_start + 3)

    ws.append(styled_row(ws, headers, qty_set, currency_set, fill=FILL_HEADER, bold=True))

    grand_qty = {qc: 0.0 for qc in qty_cols}
    grand_dol = {dc: 0.0 for dc in dol_cols}
    grand_total_qty = 0.0
    grand_total_dol = 0.0

    for item_key, grp in agg_df.groupby(["Item_#", "Item_Name"], sort=True):
        _item_id, item_name = item_key

        item_qty = {qc: 0.0 for qc in qty_cols}
        item_dol = {dc: 0.0 for dc in dol_cols}
        item_total_qty = 0.0
        item_total_dol = 0.0

        for _, r in grp.iterrows():
            vals = [r["Item_#"], r["Item_Name"], r.get("CustomerAccount", ""), r.get("CustomerName", "")]
            for qc, dc in zip(qty_cols, dol_cols):
                vq = float(r.get(qc, 0))
                vd = float(r.get(dc, 0))
                vals.append(vq)
                vals.append(vd)
                item_qty[qc] += vq
                item_dol[dc] += vd
            tq = float(r["Total_Qty"])
            td = float(r["Total_$"])
            item_total_qty += tq
            item_total_dol += td
            vals.append(tq)
            vals.append(td)
            vals.append(float(r["Avg_Price"]) if pd.notna(r["Avg_Price"]) else None)
            vals.append(float(r["BookPrice"]) if pd.notna(r.get("BookPrice")) else None)
            vals.append(r.get("Salesman", ""))
            ws.append(styled_row(ws, vals, qty_set, currency_set))

        tot_vals = ["TOTALS:", item_name, "", ""]
        for qc, dc in zip(qty_cols, dol_cols):
            tot_vals.append(item_qty[qc])
            tot_vals.append(item_dol[dc])
            grand_qty[qc] += item_qty[qc]
            grand_dol[dc] += item_dol[dc]
        tot_vals.append(item_total_qty)
        tot_vals.append(item_total_dol)
        tot_vals.append(item_total_dol / item_total_qty if item_total_qty else None)
        tot_vals.append(None)
        tot_vals.append("")
        grand_total_qty += item_total_qty
        grand_total_dol += item_total_dol
        ws.append(styled_row(ws, tot_vals, qty_set, currency_set, fill=FILL_TOTALS, bold=True))
        ws.append([])

    grand_vals = ["GRAND TOTALS:", "", "", ""]
    for qc, dc in zip(qty_cols, dol_cols):
        grand_vals.append(grand_qty[qc])
        grand_vals.append(grand_dol[dc])
    grand_vals.append(grand_total_qty)
    grand_vals.append(grand_total_dol)
    grand_vals.append(grand_total_dol / grand_total_qty if grand_total_qty else None)
    grand_vals.append(None)
    grand_vals.append("")
    ws.append(styled_row(ws, grand_vals, qty_set, currency_set, fill=FILL_GRAND, bold=True))
