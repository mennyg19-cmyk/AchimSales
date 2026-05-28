"""Multi-sheet .xlsx export for a report viewer.

Takes the payload produced by ``report_runner.run_report`` and a
per-tab layout override (column order + hidden list) supplied by the
client, and writes one sheet per tab with formatted values.
"""

from __future__ import annotations

import math
import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Excel max chars per cell; anything longer will raise IllegalCharacterError.
_MAX_CELL_CHARS = 32767
# Excel chokes on certain ASCII control characters in cell values.
_BAD_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


_NUMBER_FORMATS = {
    "money":   '"$"#,##0.00;[Red]-"$"#,##0.00',
    "int":     "#,##0",
    "percent": "0.00%",
    "date":    "mm/dd/yyyy",
}


def _apply_layout(tab: dict, layout: dict | None) -> tuple[list[dict], list[dict]]:
    """Return (visible_columns_in_order, rows) after applying the layout."""
    all_cols = list(tab.get("columns") or [])
    by_field = {c["field"]: c for c in all_cols}

    if not layout:
        return all_cols, list(tab.get("rows") or [])

    order = layout.get("order") or [c["field"] for c in all_cols]
    hidden = set(layout.get("hidden") or [])

    visible = [by_field[f] for f in order if f in by_field and f not in hidden]
    # Anything new the client didn't know about keeps its original order at the end.
    seen = {c["field"] for c in visible}
    for c in all_cols:
        if c["field"] not in seen and c["field"] not in hidden:
            visible.append(c)

    return visible, list(tab.get("rows") or [])


def _coerce_for_cell(value, col_type: str):
    """Best-effort coercion + sanitisation for an Excel cell.

    openpyxl raises on:
      - NaN / +inf / -inf floats
      - strings with ASCII control chars (\\x00, etc.)
      - strings longer than 32_767 chars
      - non-jsonable types like Decimal in some versions
    Whatever this function returns must be safe for openpyxl.
    """
    if value is None or value == "":
        return None

    if col_type == "date":
        if isinstance(value, (date, datetime)):
            return value
        try:
            return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        except Exception:
            return _sanitise_text(value)

    if col_type in ("money", "int", "percent"):
        try:
            f = float(value)
        except (TypeError, ValueError):
            return _sanitise_text(value)
        if math.isnan(f) or math.isinf(f):
            return None
        return f

    return _sanitise_text(value)


def _sanitise_text(value):
    """Strip control chars and clamp length so openpyxl is happy."""
    s = str(value)
    if _BAD_CTRL.search(s):
        s = _BAD_CTRL.sub("", s)
    if len(s) > _MAX_CELL_CHARS:
        s = s[: _MAX_CELL_CHARS - 1] + "\u2026"
    return s


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]"""
    bad = ':\\/?*[]'
    clean = "".join("_" if ch in bad else ch for ch in name).strip() or "Sheet"
    clean = clean[:31]
    if clean not in used:
        used.add(clean)
        return clean
    i = 2
    while True:
        candidate = f"{clean[:28]} ({i})"
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


def _autosize(ws, columns: list[dict], row_count: int) -> None:
    for idx, col in enumerate(columns, start=1):
        header_len = len(str(col.get("header") or col["field"]))
        # Cheap guess based on type, good enough for a first pass.
        type_guess = {
            "money":   14,
            "int":     10,
            "percent": 10,
            "date":    12,
            "text":    22,
        }.get(col.get("type", "text"), 16)
        width = min(42, max(header_len + 3, type_guess))
        ws.column_dimensions[get_column_letter(idx)].width = width


_BLOCK_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BLOCK_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_BLOCK_LABEL_FONT = Font(bold=True, size=10)
_BLOCK_NET_FILL = PatternFill("solid", fgColor="DCE6F1")
_BLOCK_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
_BLOCK_TOTAL_FILL = PatternFill("solid", fgColor="2E75B6")
_THIN = Side(style="thin", color="B4C7E7")
_BLOCK_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_commission_cards(ws, tab: dict) -> None:
    """Render the live-style per-salesman YTD commissions pivot.

    Layout matches reports/invoiced/writer._write_commissions_sheet:
    one block per salesman with months across columns and the seven
    standard rows (SubTotal, Tariff, Freight, CC, Total Invoices,
    Credits, Net Commission, Commission rate, Total Payable). A
    blank row separates blocks. Used when ``tab["layout"]`` is
    ``"commission_cards"``.
    """
    year = tab.get("year") or datetime.now().year
    end_month = tab.get("end_month") or len(tab.get("month_labels") or []) or 12
    month_labels: list[str] = list(tab.get("month_labels") or [])
    if len(month_labels) < end_month:
        # Defensive: fill in any missing labels from the canonical list.
        from test.webapp.services.reports.invoiced import _MONTH_LABELS as _ML
        month_labels = list(_ML[:end_month])

    salesmen = tab.get("salesmen") or []

    # Column layout: A=label, B=rate cell, C..(2+end_month)=months, last=YTD Total
    rate_col = 2
    first_month_col = 3
    ytd_col = first_month_col + end_month
    total_cols = ytd_col

    # Title row
    ws.cell(row=1, column=1, value=f"Commissions Summary ({year})").font = Font(bold=True, size=14)

    if not salesmen:
        ws.cell(row=3, column=1, value="No commissioned salesmen for this period.")
        ws.column_dimensions["A"].width = 40
        return

    row = 3
    for s in salesmen:
        sm_num = s.get("salesman_number") or ""
        sm_name = s.get("salesman_name") or ""
        title = f"{sm_num} - {sm_name}".strip(" -")
        pct = float(s.get("commission_pct") or 0.0)

        # Header row: <Salesman ID-Name>  [blank]  Jan Feb Mar Apr ...  YTD Total
        hdr = ws.cell(row=row, column=1, value=title)
        hdr.font = _BLOCK_HEADER_FONT
        hdr.fill = _BLOCK_HEADER_FILL
        for mi in range(end_month):
            c = ws.cell(row=row, column=first_month_col + mi, value=month_labels[mi])
            c.font = _BLOCK_HEADER_FONT
            c.fill = _BLOCK_HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=row, column=ytd_col, value="YTD Total")
        c.font = _BLOCK_HEADER_FONT
        c.fill = _BLOCK_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        row += 1

        line_defs = [
            ("SubTotal Invoices:",                       "subtotal_invoices",  None),
            ("Total Tariff Charges:",                    "tariff_charges",     None),
            ("Total Freight Charges:",                   "freight_charges",    None),
            ("Total CC Charges:",                        "cc_charges",         None),
            ("Total Invoices: (SubTotal+Tariff+Freight+CC)", "total_invoices", None),
            ("Total Credits:",                           "credits",            None),
            ("Net Commission Amount (Less Freight and CC)", "net_commission",  _BLOCK_NET_FILL),
        ]
        for label, field, fill in line_defs:
            ws.cell(row=row, column=1, value=label).font = _BLOCK_LABEL_FONT
            for mi in range(end_month):
                monthly = (s.get("monthly") or [{}])[mi] if mi < len(s.get("monthly") or []) else {}
                val = float(monthly.get(field) or 0.0)
                cell = ws.cell(row=row, column=first_month_col + mi, value=val)
                cell.number_format = _NUMBER_FORMATS["money"]
                if fill is not None:
                    cell.fill = fill
            ytd_val = float((s.get("ytd") or {}).get(field) or 0.0)
            cell = ws.cell(row=row, column=ytd_col, value=ytd_val)
            cell.number_format = _NUMBER_FORMATS["money"]
            cell.font = Font(bold=True)
            if fill is not None:
                cell.fill = fill
            row += 1

        # Commission row: rate in col B, then per-month commission, then YTD
        ws.cell(row=row, column=1, value="Commission:").font = _BLOCK_LABEL_FONT
        rate_cell = ws.cell(row=row, column=rate_col, value=pct)
        rate_cell.number_format = "0.00%"
        rate_cell.alignment = Alignment(horizontal="right")
        rate_cell.font = Font(bold=True)
        for mi in range(end_month):
            monthly = (s.get("monthly") or [{}])[mi] if mi < len(s.get("monthly") or []) else {}
            val = float(monthly.get("commission") or 0.0)
            cell = ws.cell(row=row, column=first_month_col + mi, value=val)
            cell.number_format = _NUMBER_FORMATS["money"]
        ytd_comm = float((s.get("ytd") or {}).get("commission") or 0.0)
        cell = ws.cell(row=row, column=ytd_col, value=ytd_comm)
        cell.number_format = _NUMBER_FORMATS["money"]
        cell.font = Font(bold=True)
        row += 1

        # Total Payable row
        pay_label = f"Total Payable: {title}" if title else "Total Payable:"
        c = ws.cell(row=row, column=1, value=pay_label)
        c.font = _BLOCK_TOTAL_FONT
        c.fill = _BLOCK_TOTAL_FILL
        for mi in range(end_month):
            cell = ws.cell(row=row, column=first_month_col + mi)
            cell.fill = _BLOCK_TOTAL_FILL
        ytd_payable = float((s.get("ytd") or {}).get("total_payable") or 0.0)
        cell = ws.cell(row=row, column=ytd_col, value=ytd_payable)
        cell.number_format = _NUMBER_FORMATS["money"]
        cell.font = _BLOCK_TOTAL_FONT
        cell.fill = _BLOCK_TOTAL_FILL
        row += 2  # blank row between salesmen

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 8
    for col_i in range(first_month_col, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14
    ws.freeze_panes = "A2"


def build_workbook(payload: dict, layouts: dict[str, dict] | None = None) -> bytes:
    """Render the viewer payload into a .xlsx byte string.

    ``layouts`` is ``{tab_key: {"order": [...fields...], "hidden": [...fields...]}}``.
    Missing entries fall back to the tab's declared columns.
    """
    wb = Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()
    tabs = payload.get("tabs") or []
    layouts = layouts or {}

    if not tabs:
        ws = wb.create_sheet(title="Report")
        ws["A1"] = "No data."

    for tab in tabs:
        sheet_name = _safe_sheet_name(tab.get("name") or tab["key"], used_names)
        ws = wb.create_sheet(title=sheet_name)

        # Special-case the commissions tab when the builder gave us
        # the per-salesman monthly structure. Drop into the live-style
        # pivot writer instead of the generic flat-table loop.
        if tab.get("layout") == "commission_cards":
            _write_commission_cards(ws, tab)
            continue

        cols, rows = _apply_layout(tab, layouts.get(tab["key"]))

        for idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=idx, value=col.get("header") or col["field"])
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, col in enumerate(cols, start=1):
                cell = ws.cell(
                    row=r_idx,
                    column=c_idx,
                    value=_coerce_for_cell(row.get(col["field"]), col.get("type", "text")),
                )
                fmt = _NUMBER_FORMATS.get(col.get("type"))
                if fmt:
                    cell.number_format = fmt

        ws.freeze_panes = "A2"
        _autosize(ws, cols, len(rows))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
