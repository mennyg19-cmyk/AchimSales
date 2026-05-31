"""One-time seed of the v3 `salesmen` table from the live config Excel.

The live app keeps salesman master data (number, names, commission rate) in
``config/salesman_map.xlsx``. Per the owner decision (REVIEW-LOG journal #12)
v3 seeds its own editable table from that file once, then owns the data going
forward. We read the .xlsx directly (no import of live code) so v3 stays
decoupled; commission is stored as a fraction (e.g. 0.05), matching how the
live commissions math applies it (net * pct).

Run:  python -m web.data.seed_salesmen [path-to-xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from web.data.connection import Database
from web.data.repositories.salesmen import SalesmanRepository, SalesmanSeed

# repo-root/config/salesman_map.xlsx  (this file is repo-root/v3/web/data/...)
DEFAULT_XLSX = Path(__file__).resolve().parents[3] / "config" / "salesman_map.xlsx"

_COLS = {"Key": "key", "Number": "number", "FullName": "full_name",
         "DisplayName": "display_name", "Email": "email", "Commission %": "commission"}


def _to_float(val) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def read_seeds_from_xlsx(path: Path | str = DEFAULT_XLSX) -> list[SalesmanSeed]:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"salesman map not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

        def cell(vals, name):
            i = idx.get(name)
            return vals[i] if i is not None and i < len(vals) else None

        seeds: list[SalesmanSeed] = []
        for vals in rows:
            if not vals or not vals[0]:
                continue
            seeds.append(SalesmanSeed(
                raw_key=str(cell(vals, "Key") or "").strip(),
                number=str(cell(vals, "Number") or "").strip(),
                full_name=str(cell(vals, "FullName") or "").strip(),
                display_name=str(cell(vals, "DisplayName") or "").strip(),
                email=str(cell(vals, "Email") or "").strip(),
                commission_pct=_to_float(cell(vals, "Commission %")),
            ))
        return seeds
    finally:
        wb.close()


def seed_from_xlsx(db: Database, path: Path | str = DEFAULT_XLSX) -> int:
    return SalesmanRepository(db).upsert_many(read_seeds_from_xlsx(path))


def _main(argv: list[str]) -> int:
    from web.config import load_config
    from web.data.connection import from_config

    path = Path(argv[1]) if len(argv) > 1 else DEFAULT_XLSX
    db = from_config(load_config())
    n = seed_from_xlsx(db, path)
    print(f"Seeded {n} salesmen from {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main(sys.argv))
