"""OData path for Beta: run Live's report runners, shape sheets into v3 tabs.

Used when the per-report source switch is 'odata'. Live runners write Excel;
we read the workbook sheets into the same {key,name,columns,rows} tab shape
the v3 viewer expects.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from report_engine.lib import salesman_key

log = logging.getLogger(__name__)


def _public_run_params(params: dict[str, Any] | None) -> dict[str, Any]:
    """Drop Live-internal underscore keys so a v3 client cannot steer copy paths."""
    return {
        key: value for key, value in (params or {}).items()
        if not str(key).startswith("_")
    }


def build_odata_payload(
    report_key: str,
    params: dict[str, Any],
    visible_keys: set[str] | None,
) -> dict:
    """Run the live OData report and return a v3-compatible payload."""
    from webapp.report_api import run_report

    result = run_report(report_key, _public_run_params(params))
    if not result.get("success"):
        raise RuntimeError(result.get("error") or f"Live OData run failed for {report_key}")

    filepath = result.get("filepath") or ""
    extra_files = result.get("extra_files") or []
    paths = [p for p in [filepath, *_extra_file_paths(extra_files)] if p and Path(p).is_file()]
    if not paths:
        raise RuntimeError(f"Live OData run for {report_key} produced no Excel file")

    tabs = []
    for path in paths:
        tabs.extend(_workbook_to_tabs(path, report_key=report_key, source_path=path))

    if visible_keys is not None:
        tabs = [_scope_tab(tab, visible_keys) for tab in tabs]

    if report_key == "ordered" and not params.get("salesman"):
        tabs = [_attach_ordered_default_group(tab) for tab in tabs]
    elif report_key == "number_4":
        tabs = [_attach_number4_defaults(tab) for tab in tabs]

    row_count = sum(len(t.get("rows") or []) for t in tabs)
    return {
        "report_key": report_key,
        "tabs": tabs,
        "row_count": row_count,
        "data_source": "odata",
    }


# Match SQL Ordered builder: Summary / By Customer / By Order group by Salesman.
_ORDERED_DEFAULT_GROUP_KEYS = {
    "summary", "by_customer", "by_order",
    "summary_by_customer",  # live filtered variant sheet names sometimes differ
}
_ORDERED_DEFAULT_GROUP_NAMES = {
    "summary", "by customer", "by order",
}


def _extra_file_paths(extra_files) -> list[str]:
    """History stores extra_files as {filepath, filename} dicts, not bare paths."""
    paths: list[str] = []
    for item in extra_files or []:
        if isinstance(item, str) and item:
            paths.append(item)
        elif isinstance(item, dict):
            p = item.get("filepath") or item.get("path") or ""
            if p:
                paths.append(p)
    return paths


def _attach_ordered_default_group(tab: dict) -> dict:
    key = str(tab.get("key") or "").strip().lower()
    name = str(tab.get("name") or "").strip().lower()
    if key in _ORDERED_DEFAULT_GROUP_KEYS or name in _ORDERED_DEFAULT_GROUP_NAMES:
        out = dict(tab)
        out["default_group"] = ["Salesman"]
        return out
    return tab


def _number4_version_label(path: str) -> str:
    name = Path(path).name.lower()
    if "item" in name:
        return "By Item"
    if "customer" in name:
        return "By Customer"
    return ""


def _is_summary_row(row: dict) -> bool:
    for value in row.values():
        text = str(value or "").strip()
        if text.startswith("TOTALS") or text.startswith("GRAND TOTALS"):
            return True
    return False


def _attach_number4_defaults(tab: dict) -> dict:
    """Group by item; drop Excel TOTALS rows so the grid grouping is clean."""
    from report_engine.reports.number_4 import canonical_header, order_number4_columns

    out = dict(tab)
    out["default_group"] = ["Item #"]
    rows = [r for r in (tab.get("rows") or []) if not _is_summary_row(r)]
    out["rows"] = rows
    cols = out.get("columns") or []
    if cols and isinstance(cols[0], str):
        out["columns"] = order_number4_columns(
            canonical_header(c) for c in cols
        )
    return out


def _workbook_to_tabs(path: str, *, report_key: str = "", source_path: str = "") -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    tabs: list[dict] = []
    try:
        for sheet in wb.worksheets:
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                continue
            columns = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(header)]
            # Skip fully blank header rows
            if not any(columns):
                continue
            data_rows: list[dict] = []
            for raw in rows_iter:
                if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                    continue
                row = {}
                for i, col in enumerate(columns):
                    if not col:
                        continue
                    val = raw[i] if i < len(raw) else None
                    if hasattr(val, "isoformat"):
                        val = val.isoformat()
                    row[col] = val
                data_rows.append(row)
            title = sheet.title
            prefix = _number4_version_label(source_path or path) if report_key == "number_4" else ""
            if prefix:
                title = f"{prefix} ({sheet.title})"
            key = _slug(title)
            tabs.append({
                "key": key,
                "name": title,
                "columns": [c for c in columns if c],
                "rows": data_rows,
            })
    finally:
        wb.close()
    return tabs


def _slug(title: str) -> str:
    out = "".join(ch.lower() if ch.isalnum() else "_" for ch in (title or "sheet"))
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_") or "sheet"


def _scope_tab(tab: dict, visible_keys: set[str]) -> dict:
    """Best-effort salesman scope on OData sheets (column name varies)."""
    # None is handled by the caller (unrestricted). An empty set is a real
    # scope: this user may see no salesmen, so do not return the full tab.
    if not visible_keys:
        out = dict(tab)
        out["rows"] = []
        return out
    rows = tab.get("rows") or []
    if not rows:
        return tab
    scope_cols = ("Salesman", "SalesGroup", "SalesmanNumber", "Sales Group", "sales_group")
    col = next((c for c in scope_cols if c in (tab.get("columns") or [])), None)
    if col is None:
        # Try first row keys
        sample = rows[0]
        col = next((c for c in scope_cols if c in sample), None)
    if col is None:
        out = dict(tab)
        out["rows"] = []
        return out
    allowed = {salesman_key(str(k)) for k in visible_keys}
    allowed.discard("")
    filtered = [
        r for r in rows if salesman_key(str(r.get(col, ""))) in allowed
    ]
    out = dict(tab)
    out["rows"] = filtered
    return out
