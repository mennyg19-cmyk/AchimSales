"""Styled Excel export from a report payload (openpyxl).

One worksheet per tab, formatted to match the look of the live app's exports:
a bold grey header row, thin cell borders, subtle zebra striping, and real
Excel number formats per column type (currency, integer, percent, date). When
a tab is grouped (via the saved/clicked view), each group gets a subtotal line
and the sheet ends with a grand total - mirroring the legacy test app.

Payload shape: {"tabs": [{"key", "name", "columns": [...], "rows": [{col: val}],
"layout"?}]}. ``columns`` is either a list of header strings or a list of
{"field", "header", "type"} dicts (the viewer shape); both are supported.

``build_workbook(payload, layout)`` accepts the viewer's serialized layout
(``{"views": {<tabKey>: {"group": [...], ...}}}``) to drive per-tab grouping.
``apply_layout`` (web.delivery.layout) should be run on the payload first to
replay column order / hide / sort / filter; this module only adds grouping +
styling on top. ``payload_to_xlsx`` is the no-layout shortcut kept for the
delivery + test callers. Pure transform -> bytes; no Flask, no DB.

Built in openpyxl **write-only (streaming) mode**: rows are appended top-to-bottom
and flushed to a temp file as they go, so a six-tab report with hundreds of
thousands of rows builds in seconds with flat memory instead of materialising
every styled cell in RAM (which made big exports time out).
"""

from __future__ import annotations

import math
import re
from datetime import date, datetime
from io import BytesIO
from itertools import groupby
from typing import Any

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_engine.lib import iso_date

# Excel sheet-title constraints: <=31 chars, none of : \ / ? * [ ]
_INVALID_SHEET = re.compile(r"[:\\/?*\[\]]")
# CSV/Excel formula-injection: a cell whose text starts with one of these can be
# executed as a formula. Prefix with an apostrophe to force it to literal text.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n")
# openpyxl rejects these ASCII control chars in cell text; strip them.
_BAD_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_MAX_CELL_CHARS = 32767

_NUMERIC_TYPES = {"money", "int", "percent"}
# Columns that can be meaningfully summed for subtotal/grand-total rows. Percent
# is deliberately excluded: summing raw percentages is nonsense (the on-screen
# grid excludes percent bottom-calcs too), so percent total cells stay blank.
_SUMMABLE_TYPES = {"money", "int"}
# Unit prices (Extended / Qty). Summing them on a group footer is wrong.
_NEVER_SUM_FIELDS = frozenset({"Net Price"})

# Nested group colors. Outer = darkest. Keep in sync with report.ts nest* helpers.
# Discrete shades so every step stays readable (lerp hits a dead zone around 4.5:1).
_HEADER_SHADES = (
    (30, 64, 175),     # #1E40AF
    (37, 99, 235),     # #2563EB (app primary)
    (96, 165, 250),    # #60A5FA
    (147, 197, 253),   # #93C5FD
    (191, 219, 254),   # #BFDBFE
)
_FOOTER_SHADES = (
    (107, 114, 128),   # #6B7280
    (156, 163, 175),   # #9CA3AF
    (176, 182, 191),   # #B0B6BF
    (189, 196, 204),   # #BDC4CC
)
_GRAND_GREY = (55, 65, 81)             # #374151
_TEXT_DARK = "1E293B"
_TEXT_LIGHT = "FFFFFF"
_FILL_BY_HEX: dict[str, PatternFill] = {}
_FONT_BOLD_BY_HEX: dict[str, Font] = {}

# --- Number formats (match the on-screen grid + the live exports) ---
_FMT = {
    "money": '"$"#,##0.00',
    "int": "#,##0",
    "percent": "0.0%",
    "date": "YYYY-MM-DD",
}

# --- Styling (live palette: grey header, light zebra, darker totals) ---
_HEADER_FILL = PatternFill("solid", fgColor="E0E0E0")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
_TOTAL_FONT = Font(bold=True)
_GROUP_FILL = PatternFill("solid", fgColor="BDD7EE")
_GROUP_FONT = Font(bold=True)
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
# Live salesman YoY column bands (month / YTD / full-year) + red negatives.
_FONT_BAND_BLUE = Font(color="0000CC")
_FONT_BAND_GREEN = Font(color="008000")
_FONT_BAND_PURPLE = Font(color="800080")
_FONT_BAND_RED = Font(color="FF0000")
_BAND_FONTS = (_FONT_BAND_BLUE, _FONT_BAND_GREEN, _FONT_BAND_PURPLE)
_FILL_LIGHT_GREY = PatternFill("solid", fgColor="E8E8E8")
# Salesman identity columns never get YoY color bands.
_SALESMAN_ID_FIELDS = frozenset({
    "Sort Number", "Salesman", "Cust. #", "Customer Name", "SalesmanNumber",
})


def _hex6(rgb: tuple[int, int, int]) -> str:
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def _rel_lum(rgb: tuple[int, int, int]) -> float:
    def _lin(c: int) -> float:
        x = c / 255.0
        return x / 12.92 if x <= 0.04045 else ((x + 0.055) / 1.055) ** 2.4
    return 0.2126 * _lin(rgb[0]) + 0.7152 * _lin(rgb[1]) + 0.0722 * _lin(rgb[2])


def _contrast(a: tuple[int, int, int], b: tuple[int, int, int]) -> float:
    hi, lo = _rel_lum(a), _rel_lum(b)
    if lo > hi:
        hi, lo = lo, hi
    return (hi + 0.05) / (lo + 0.05)


def _contrast_text_hex(rgb: tuple[int, int, int]) -> str:
    white, dark = (255, 255, 255), (30, 41, 59)
    return _TEXT_LIGHT if _contrast(rgb, white) >= _contrast(rgb, dark) else _TEXT_DARK


def _shade_at(shades: tuple[tuple[int, int, int], ...], level: int, depth: int) -> tuple[int, int, int]:
    if depth <= 1:
        return shades[0]
    last = len(shades) - 1
    idx = int(round(level * last / (depth - 1)))
    return shades[max(0, min(idx, last))]


def nest_header_rgb(level: int, depth: int) -> tuple[int, int, int]:
    return _shade_at(_HEADER_SHADES, level, depth)


def nest_footer_rgb(level: int, depth: int, *, grand: bool) -> tuple[int, int, int]:
    if grand:
        return _GRAND_GREY
    # Consecutive steps from the dark end. Stretching to the palest shade made
    # 2-level (Daily Ordered) customer totals look white.
    if depth <= 1:
        return _FOOTER_SHADES[0]
    return _FOOTER_SHADES[max(0, min(int(level), len(_FOOTER_SHADES) - 1))]


def _fill_hex(hex6: str) -> PatternFill:
    fill = _FILL_BY_HEX.get(hex6)
    if fill is None:
        fill = PatternFill("solid", fgColor=hex6)
        _FILL_BY_HEX[hex6] = fill
    return fill


def _font_bold(hex6: str) -> Font:
    font = _FONT_BOLD_BY_HEX.get(hex6)
    if font is None:
        font = Font(bold=True, color=hex6)
        _FONT_BOLD_BY_HEX[hex6] = font
    return font


def _nest_header_style(level: int, depth: int) -> tuple[PatternFill, Font]:
    rgb = nest_header_rgb(level, depth)
    return _fill_hex(_hex6(rgb)), _font_bold(_contrast_text_hex(rgb))


def _nest_footer_style(level: int, depth: int, *, grand: bool) -> tuple[PatternFill, Font]:
    rgb = nest_footer_rgb(level, depth, grand=grand)
    return _fill_hex(_hex6(rgb)), _font_bold(_contrast_text_hex(rgb))


def _col_summable(field: str, ctype: str | None, col: dict | None = None) -> bool:
    if isinstance(col, dict) and col.get("sum") is False:
        return False
    if field in _NEVER_SUM_FIELDS:
        return False
    return ctype in _SUMMABLE_TYPES


def _fulfillment_fill(score: Any) -> PatternFill | None:
    """Red (0) → yellow (0.5) → green (1). Same RGB as the old Ordered writer."""
    try:
        s = float(score)
        if s < 0 or s != s:
            return _FILL_LIGHT_GREY
    except (TypeError, ValueError):
        return None
    s = max(0.0, min(1.0, s))
    red, yellow, green = (255, 199, 206), (255, 235, 156), (198, 239, 206)
    if s <= 0:
        r, g, b = red
    elif s >= 1:
        r, g, b = green
    elif s < 0.5:
        t = s * 2
        r = int(red[0] + (yellow[0] - red[0]) * t)
        g = int(red[1] + (yellow[1] - red[1]) * t)
        b = int(red[2] + (yellow[2] - red[2]) * t)
    else:
        t = (s - 0.5) * 2
        r = int(yellow[0] + (green[0] - yellow[0]) * t)
        g = int(yellow[1] + (green[1] - yellow[1]) * t)
        b = int(yellow[2] + (green[2] - yellow[2]) * t)
    return PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")


def _safe_text(value: Any) -> str:
    """Stringify, neutralise formula injection, strip control chars, clamp length."""
    s = str(value)
    if s[:1] in _FORMULA_TRIGGERS:
        s = "'" + s
    if _BAD_CTRL.search(s):
        s = _BAD_CTRL.sub("", s)
    if len(s) > _MAX_CELL_CHARS:
        s = s[: _MAX_CELL_CHARS - 1] + "\u2026"
    return s


def _coerce(value: Any, col_type: str | None) -> tuple[Any, str | None]:
    """Return (excel_value, number_format) for a typed cell, openpyxl-safe."""
    if value is None or value == "":
        return None, None
    if col_type == "date":
        if isinstance(value, (date, datetime)):
            return value, _FMT["date"]
        s = iso_date(value)
        if s and len(s) == 10 and s[4] == "-" and s[7] == "-":
            try:
                return date.fromisoformat(s), _FMT["date"]
            except ValueError:
                pass
        return _safe_text(value), None
    if col_type in _NUMERIC_TYPES:
        try:
            f = float(value)
        except (TypeError, ValueError):
            return _safe_text(value), None
        if math.isnan(f) or math.isinf(f):
            return None, None
        return f, _FMT[col_type]
    return _safe_text(value), None


def _num(value: Any) -> float | None:
    try:
        f = float(value)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return None


def _infer_salesman_band(field: str) -> int | None:
    """0=blue month YoY, 1=green YTD, 2=purple full year. Follows the field name.

    Used when a column dict has no ``band`` (older cached payloads). Must not
    use Excel column letters — hidden/reordered views move the same field.
    """
    if field in _SALESMAN_ID_FIELDS:
        return None
    if "(YTD Full Year)" in field or field.startswith("Sales Year to Date "):
        return 2
    if "(YTD)" in field or " Jan Thru " in field:
        return 1
    if (
        field.startswith("Sales ")
        or field in ("$ This Year to Last Year", "% This Year to Last Year")
    ):
        return 0
    return None


def _explicit_salesman_bands(columns: list) -> dict[str, int]:
    """``band`` stamped on salesman column dicts (survives hide/reorder)."""
    out: dict[str, int] = {}
    for col in columns or []:
        if not isinstance(col, dict) or col.get("band") is None:
            continue
        field = str(col.get("field") or col.get("header") or "")
        if not field:
            continue
        try:
            out[field] = min(max(int(col["band"]), 0), 2)
        except (TypeError, ValueError):
            continue
    return out


def _salesman_font(field: str, raw: Any, bands: dict[str, int]) -> Font | None:
    idx = bands[field] if field in bands else _infer_salesman_band(field)
    if idx is None:
        return None
    n = _num(raw)
    if n is not None and n < 0:
        return _FONT_BAND_RED
    return _BAND_FONTS[min(idx, 2)]


def _safe_sheet_title(name: str, used: set[str]) -> str:
    title = _INVALID_SHEET.sub(" ", (name or "Sheet").strip())[:31] or "Sheet"
    base, n = title, 2
    while title.lower() in used:
        suffix = f" {n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _columns_meta(columns: list, rows: list) -> list[tuple[str, str, str | None, bool]]:
    """Normalise a tab's columns to (header, field, type, summable) tuples.

    Supports the viewer's {"field","header","type","sum"} dicts, plain header
    strings, or (when no columns are declared) the keys of the first row.
    """
    if columns and isinstance(columns[0], dict):
        out = []
        for c in columns:
            field = str(c.get("field") or c.get("header") or "")
            ctype = c.get("type")
            out.append((
                str(c.get("header") or c.get("field") or ""),
                field,
                ctype,
                _col_summable(field, ctype, c),
            ))
        return out
    if columns:
        return [(str(c), str(c), None, _col_summable(str(c), None)) for c in columns]
    if rows:
        return [(str(k), str(k), None, _col_summable(str(k), None)) for k in rows[0].keys()]
    return []


def _group_sort_key(value: Any) -> tuple:
    if value is None or value == "":
        return (1, "")
    return (0, str(value).lower())


def _autosize(ws, metas: list[tuple[str, str, str | None, bool]]) -> None:
    type_guess = {"money": 14, "int": 10, "percent": 10, "date": 12, "text": 22}
    for idx, (header, _field, ctype, _summable) in enumerate(metas, start=1):
        guess = type_guess.get(ctype or "text", 16)
        width = min(45, max(len(header) + 3, guess))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _cell(ws, value: Any, *, fmt: str | None = None, font: Font | None = None,
          fill: PatternFill | None = None, align: Alignment | None = None,
          border: Border | None = None) -> WriteOnlyCell:
    """A styled write-only cell. Styles (Font/Fill/Border) are shared singletons
    so openpyxl dedups them; only the lightweight cell object is per-value."""
    c = WriteOnlyCell(ws, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if border:
        c.border = border
    if align:
        c.alignment = align
    if fmt:
        c.number_format = fmt
    return c


def _header_cells(ws, metas) -> list[WriteOnlyCell]:
    return [_cell(ws, _safe_text(h), font=_HEADER_FONT, fill=_HEADER_FILL,
                  align=_HEADER_ALIGN, border=_BORDER) for h, _f, _t, _s in metas]


def _data_cells(ws, metas, row: dict, acc: dict[str, float],
                *, salesman_bands: bool = False,
                salesman_band_by_field: dict[str, int] | None = None) -> list[WriteOnlyCell]:
    """Data row cells — styled lightly (number formats only, no borders/zebra).

    The live app applies no per-cell borders or zebra on data rows; only headers,
    totals, and group banners get fills/borders. Skipping these two style objects
    on every cell is the single biggest speedup for large exports (borders alone
    were ~40% of openpyxl's write-time on a 120k-row grid).

    Salesman month tabs get blue/green/purple font by field identity (and red
    for negatives), not by Excel column letter — hidden/reordered views move
    the same numbers to different letters.
    """
    bands = salesman_band_by_field or {}
    cells = []
    for _header, field, ctype, summable in metas:
        value, fmt = _coerce(row.get(field), ctype)
        font = _salesman_font(field, row.get(field), bands) if salesman_bands else None
        fill = _fulfillment_fill(row.get(field)) if field == "Fulfillment %" else None
        cells.append(_cell(ws, value, fmt=fmt, font=font, fill=fill))
        if summable:
            x = _num(row.get(field))
            if x is not None:
                acc[field] = acc.get(field, 0.0) + x
    return cells


def _total_cells(ws, metas, sums: dict[str, float], label: str,
                 *, fill: PatternFill, font: Font) -> list[WriteOnlyCell]:
    # Put the "Total"/"Grand total" label in the first NON-summable column so a
    # numeric first column (e.g. an order number is text, but a money/qty first
    # column isn't) keeps its own subtotal instead of being overwritten by text.
    label_idx = next((i for i, (_h, _f, _t, s) in enumerate(metas) if not s), 0)
    cells = []
    for i, (_header, field, ctype, summable) in enumerate(metas):
        if i == label_idx:
            value, fmt = label, None
        elif summable and field in sums:
            value, fmt = sums[field], _FMT.get(ctype or "", None)
        else:
            value, fmt = None, None
        cells.append(_cell(ws, value, fmt=fmt, font=font, fill=fill, border=_BORDER))
    return cells


def _banner_cells(ws, ncol: int, text: str, *, fill: PatternFill, font: Font) -> list[WriteOnlyCell]:
    cells = [_cell(ws, _safe_text(text), font=font, fill=fill, border=_BORDER)]
    cells += [_cell(ws, None, font=font, fill=fill, border=_BORDER) for _ in range(ncol - 1)]
    return cells


def _sorter_specs(sorters: list | None) -> list[tuple[str, bool]]:
    specs: list[tuple[str, bool]] = []
    seen: set[str] = set()
    for spec in sorters or []:
        if not isinstance(spec, dict):
            continue
        col = spec.get("column") or spec.get("field")
        if not col or col in seen:
            continue
        specs.append((str(col), (spec.get("dir") or "asc").lower() != "desc"))
        seen.add(str(col))
    return specs


def _sort_rows_for_groups(rows: list, group_fields: list[str],
                          sorters: list | None) -> list:
    """Keep group keys consecutive, then apply extra sorts inside each group.

    Excel grouping uses itertools.groupby, which needs adjacent keys. If every
    group field is already in the sorter list (Heshy: customer then order
    number, group by order), honour that sorter order. Otherwise group fields
    are primary so a customer sort cannot split a salesman group.
    """
    sorter_specs = _sorter_specs(sorters)
    sorter_cols = {col for col, _asc in sorter_specs}
    group_covered = bool(group_fields) and all(g in sorter_cols for g in group_fields)
    specs: list[tuple[str, bool]] = []
    seen: set[str] = set()

    def _add(col: str, ascending: bool) -> None:
        if not col or col in seen:
            return
        specs.append((col, ascending))
        seen.add(col)

    if not group_covered:
        for field in group_fields:
            _add(field, True)
    for col, ascending in sorter_specs:
        _add(col, ascending)
    ordered = list(rows)
    for col, ascending in reversed(specs):
        ordered.sort(key=lambda row, c=col: _group_sort_key(row.get(c)),
                     reverse=not ascending)
    return ordered


def _emit_grouped(ws, metas, rows: list, group_fields: list[str],
                  *, salesman_bands: bool,
                  salesman_band_by_field: dict[str, int] | None = None,
                  parent_acc: dict[str, float] | None = None,
                  group_level: int = 0, group_depth: int = 1
                  ) -> dict[str, float]:
    """Nested group banners + per-level totals. Innermost level writes data rows."""
    gf = group_fields[0]
    rest = group_fields[1:]
    glabel = next((h for h, f, _t, _s in metas if f == gf), gf)
    ncol = len(metas)
    banner_fill, banner_font = _nest_header_style(group_level, group_depth)
    total_fill, total_font = _nest_footer_style(group_level, group_depth, grand=False)
    level: dict[str, float] = {}
    for _key, grp_iter in groupby(rows, key=lambda x: _group_sort_key(x.get(gf))):
        grp = list(grp_iter)
        gval = grp[0].get(gf)
        label = gval if gval not in (None, "") else "(blank)"
        ws.append(_banner_cells(ws, ncol, f"{glabel}: {label}",
                                fill=banner_fill, font=banner_font))
        sub: dict[str, float] = {}
        if rest:
            _emit_grouped(ws, metas, grp, rest, salesman_bands=salesman_bands,
                          salesman_band_by_field=salesman_band_by_field,
                          parent_acc=sub, group_level=group_level + 1,
                          group_depth=group_depth)
        else:
            for row in grp:
                ws.append(_data_cells(
                    ws, metas, row, sub, salesman_bands=salesman_bands,
                    salesman_band_by_field=salesman_band_by_field,
                ))
        ws.append(_total_cells(ws, metas, sub, f"Total \u2014 {label}",
                               fill=total_fill, font=total_font))
        for k, val in sub.items():
            level[k] = level.get(k, 0.0) + val
    if parent_acc is not None:
        for k, val in level.items():
            parent_acc[k] = parent_acc.get(k, 0.0) + val
    return level


def _stream_grid(ws, metas, rows: list, group_fields: list[str],
                 *, salesman_bands: bool = False,
                 salesman_band_by_field: dict[str, int] | None = None,
                 sorters: list | None = None) -> None:
    if not metas:
        return
    ncol = len(metas)
    # Worksheet-level properties are written at save time, so they can be set
    # before appending rows (write-only mode only forbids per-cell random access).
    ws.freeze_panes = "A2"
    _autosize(ws, metas)
    if not group_fields and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{len(rows) + 1}"
    ws.append(_header_cells(ws, metas))

    grand_fill, grand_font = _nest_footer_style(0, max(len(group_fields), 1), grand=True)
    if group_fields:
        ordered = _sort_rows_for_groups(rows, group_fields, sorters)
        grand = _emit_grouped(
            ws, metas, ordered, group_fields,
            salesman_bands=salesman_bands,
            salesman_band_by_field=salesman_band_by_field,
            group_level=0, group_depth=len(group_fields),
        )
        if rows:
            ws.append(_total_cells(ws, metas, grand, "Grand total",
                                   fill=grand_fill, font=grand_font))
    else:
        grand: dict[str, float] = {}
        for row in rows:
            ws.append(_data_cells(
                ws, metas, row, grand, salesman_bands=salesman_bands,
                salesman_band_by_field=salesman_band_by_field,
            ))
        if rows:
            ws.append(_total_cells(ws, metas, grand, "Total",
                                   fill=grand_fill, font=grand_font))


def _stream_commission(ws, tab: dict) -> None:
    """Per-salesman monthly + YTD block (live Excel commissions layout)."""
    year = tab.get("year") or ""
    labels = list(tab.get("month_labels") or [])
    salesmen = tab.get("salesmen") or []
    ws.freeze_panes = "A2"
    title = f"Commissions Summary ({year})" if year else "Commissions Summary"
    ws.append([_cell(ws, title, font=Font(bold=True, size=14))])
    if not salesmen:
        ws.column_dimensions["A"].width = 48
        ws.append([])
        ws.append([_cell(ws, "No commissioned salesmen for this period.")])
        return
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 8

    lines = [
        ("SubTotal Invoices:", "subtotal_invoices"),
        ("Total Tariff Charges:", "tariff_charges"),
        ("Total Freight Charges:", "freight_charges"),
        ("Total CC Charges:", "cc_charges"),
        ("Total Invoices: (SubTotal+Tariff+Freight+CC)", "total_invoices"),
        ("Total Credits:", "credits"),
        ("Net Commission Amount (Less Freight and CC)", "net_commission"),
        ("Commission:", "commission"),
    ]
    ws.append([])  # row 2 spacer (block starts on row 3, matching the live layout)
    center = Alignment(horizontal="center")
    yy = str(year)[-2:] if year else ""
    for s in salesmen:
        title_sm = str(s.get("salesman_name") or s.get("salesman") or "").strip()
        banner = [_cell(ws, _safe_text(title_sm), font=_GROUP_FONT, fill=_GROUP_FILL)]
        banner.append(_cell(ws, "", font=_GROUP_FONT, fill=_GROUP_FILL))
        for lab in labels:
            hdr = f"{lab}-{yy}" if yy else lab
            banner.append(_cell(ws, _safe_text(hdr), font=_GROUP_FONT, fill=_GROUP_FILL, align=center))
        banner.append(_cell(ws, "YTD Total", font=_GROUP_FONT, fill=_GROUP_FILL, align=center))
        ws.append(banner)
        monthly = s.get("monthly") or []
        ytd = s.get("ytd") or {}
        pct = float(s.get("commission_pct") or 0.0)
        for label, field in lines:
            line = [_cell(ws, _safe_text(label), font=_TOTAL_FONT)]
            if field == "commission":
                line.append(_cell(ws, pct, fmt=_FMT["percent"]))
            else:
                line.append(_cell(ws, ""))
            for mi in range(len(labels)):
                m = monthly[mi] if mi < len(monthly) else {}
                line.append(_cell(ws, float(m.get(field) or 0.0), fmt=_FMT["money"]))
            line.append(_cell(ws, float(ytd.get(field) or 0.0), fmt=_FMT["money"], font=_TOTAL_FONT))
            ws.append(line)
        pay = [_cell(ws, _safe_text(f"Total Payable: {title_sm}"), font=_TOTAL_FONT)]
        pay.append(_cell(ws, ""))
        for _ in labels:
            pay.append(_cell(ws, ""))
        pay.append(_cell(
            ws,
            float(ytd.get("total_payable") or ytd.get("commission") or 0.0),
            fmt=_FMT["money"],
            font=_TOTAL_FONT,
        ))
        ws.append(pay)
        ws.append([])  # blank row between salesmen

def _tab_groups_and_sorters(
    tab: dict, view: dict, rows: list, known: set[str],
) -> tuple[list[str], list | None]:
    """Resolve Excel groups/sorts for one tab.

    Empty group [] is a saved ungroup (Default view). Only use the builder
    default_group when the view never set group at all.

    A salesman-split Ordered file has empty default_group (the sheet is already
    one rep). Daily Ordered groups By Customer by Salesman only — drop that
    redundant Salesman level on a one-rep file.
    Summary's builder default_layout (Customer Name then Item) fills extra
    sorts when the view did not set sorters, including when default_group is
    already Salesman.
    """
    if "group" in view:
        wanted = view["group"] if isinstance(view["group"], list) else []
        group_was_set = True
    else:
        wanted = tab.get("default_group") if isinstance(tab.get("default_group"), list) else []
        group_was_set = False
    if rows:
        salesman_vals = {row.get("Salesman") for row in rows}
        if len(salesman_vals) <= 1:
            wanted = [g for g in wanted if g != "Salesman"]
    sorters = view.get("sorters") if isinstance(view.get("sorters"), list) else []
    dl = tab.get("default_layout") if isinstance(tab.get("default_layout"), dict) else {}
    if not wanted and not group_was_set:
        wanted = [g for g in (dl.get("group_levels") or []) if isinstance(g, str)]
        if not sorters:
            sorters = _sorters_from_default_layout(dl)
    elif wanted and not sorters:
        # Company Daily Ordered Summary groups by salesman; still take the
        # builder's Customer Name / Item sort so rows are A-Z inside the group.
        sorters = _sorters_from_default_layout(dl)
    group_fields = [g for g in wanted if g in known]
    return group_fields, sorters or None


def _sorters_from_default_layout(dl: dict) -> list:
    out: list[dict] = []
    for spec in dl.get("sort_levels") or []:
        if not isinstance(spec, dict):
            continue
        col = spec.get("field") or spec.get("column")
        if not col:
            continue
        out.append({"column": col, "dir": spec.get("dir") or "asc"})
    return out


def build_workbook(payload: dict[str, Any], layout: dict | None = None) -> bytes:
    from openpyxl import Workbook
    from web.jobs.trace import step as job_step

    wb = Workbook(write_only=True)  # streaming: flat memory, fast on huge reports
    used: set[str] = set()
    views = (layout or {}).get("views") if isinstance(layout, dict) else None
    views = views if isinstance(views, dict) else {}
    tabs = payload.get("tabs") or []
    if not tabs:
        wb.create_sheet(_safe_sheet_title("Report", used))
    for tab in tabs:
        job_step("xlsx", f"sheet {tab.get('name') or tab.get('key') or 'Report'}: "
                 f"{len(tab.get('rows') or [])} rows")
        ws = wb.create_sheet(_safe_sheet_title(tab.get("name", "Report"), used))
        if tab.get("layout") == "commission_cards" and tab.get("salesmen") is not None:
            _stream_commission(ws, tab)
            continue
        rows = list(tab.get("rows") or [])
        metas = _columns_meta(list(tab.get("columns") or []), rows)
        v = views.get(tab.get("key"))
        v = v if isinstance(v, dict) else {}
        # A group field may be hidden (so absent from metas) yet still present in
        # the row dicts - honour it from the row data, not just visible columns.
        known = {f for _h, f, _t, _s in metas} | (set(rows[0].keys()) if rows else set())
        group_fields, sorters = _tab_groups_and_sorters(tab, v, rows, known)
        salesman_bands = (payload.get("report_key") == "salesman"
                          and tab.get("layout") != "commission_cards")
        band_by_field = (_explicit_salesman_bands(list(tab.get("columns") or []))
                         if salesman_bands else None)
        _stream_grid(ws, metas, rows, group_fields, salesman_bands=salesman_bands,
                     salesman_band_by_field=band_by_field, sorters=sorters)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def payload_to_xlsx(payload: dict[str, Any]) -> bytes:
    """No-layout styled export (one styled sheet per tab, no grouping)."""
    return build_workbook(payload, None)
