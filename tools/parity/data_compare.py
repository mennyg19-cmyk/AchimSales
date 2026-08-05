"""Match Excel report data by row keys, not cell position.

Ignores formatting, column order, and columns that exist only on /test.
Reports missing rows/columns and numeric gaps with pattern summaries
(e.g. "all missing rows share InvoiceDate=2026-07-01").
"""

from __future__ import annotations

import math
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Normalize header text → canonical field name.
_HEADER_ALIASES: dict[str, str] = {
    "customeraccount": "customer_account",
    "customer account": "customer_account",
    "cust. #": "customer_account",
    "cust #": "customer_account",
    "cust#": "customer_account",
    "cust": "customer_account",  # slug of "Cust. #"
    "customername": "customer_name",
    "customer name": "customer_name",
    "invoicenumber": "invoice_number",
    "invoice number": "invoice_number",
    "salesordernumber": "sales_order_number",
    "sales order number": "sales_order_number",
    "invoicedate": "invoice_date",
    "invoice date": "invoice_date",
    "orderdate": "order_date",
    "order date": "order_date",
    "salesmannumber": "salesman_number",
    "salesman number": "salesman_number",
    "salesman": "salesman_code",
    "salesmanname": "salesman_name",
    "salesman name": "salesman_name",
    "item#": "item_number",
    "itemnumber": "item_number",
    "item number": "item_number",
    "itemname": "item_name",
    "item name": "item_name",
    "line description": "item_name",
    "linenumber": "line_number",
    "line number": "line_number",
    "invoicecount": "invoice_count",
    "subtotal invoices": "subtotal_invoices",
    "total invoices": "total_invoices",
    "total invoice": "total_invoice",
    "total tariff charges": "tariff_charges",
    "tariff charges": "tariff_charges",
    "total freight charges": "freight_charges",
    "freight charges": "freight_charges",
    "total cc charges": "cc_charges",
    "cc charges": "cc_charges",
    "total misc charges": "misc_charges",
    "misc charges": "misc_charges",
    "po #": "po_number",
    "qtyordered": "qty_ordered",
    "qtyshipped": "qty_shipped",
    "qtycancelled": "qty_cancelled",
    "qtyreleased": "qty_released",
    "qty shipping": "qty_released",
    "qtyshipping": "qty_released",
    "qtyopen": "qty_open",
    "ordered $": "ordered_dollars",
    "shipped $": "shipped_dollars",
    "cancelled $": "cancelled_dollars",
    "released $": "released_dollars",
    "shipping $": "released_dollars",
    "open $": "open_dollars",
}

# Soft (display) fields: cosmetic mismatches don't fail the sheet.
_SOFT_FIELDS = frozenset({
    "customer_name",
    "salesman_name",
    "item_name",
    "itemname",
    "line_description",
    "sales_order_name",
    "salesordername",
})

# Prefer these as row identity, in order (first available wins as primary set).
_KEY_CANDIDATES: tuple[tuple[str, ...], ...] = (
    ("invoice_number",),
    ("sales_order_number", "line_number", "item_number"),
    ("sales_order_number", "item_number"),
    ("sales_order_number",),
    ("customer_account", "item_number"),
    ("customer_name", "item_number"),
    ("customer_account", "salesman_code"),
    ("customer_account", "salesman_number"),
    ("customer_account",),
    ("customer_name", "salesman_code"),
    ("salesman_code",),
    ("salesman_number",),
    ("item_number",),
)

# Excel export group/total rows — not real data keys.
_SKIP_KEY_PREFIXES = ("salesman:", "total", "total —", "total -")



@dataclass
class SheetDataResult:
    sheet: str
    status: str  # MATCH | DIFF | SKIP | MISSING_TEST
    key_fields: list[str] = field(default_factory=list)
    compared_columns: list[str] = field(default_factory=list)
    ignored_extra_in_test: list[str] = field(default_factory=list)
    missing_columns_in_test: list[str] = field(default_factory=list)
    live_rows: int = 0
    test_rows: int = 0
    matched_rows: int = 0
    missing_in_test: list[str] = field(default_factory=list)
    missing_in_live: list[str] = field(default_factory=list)
    value_diffs: list[str] = field(default_factory=list)
    soft_diffs: list[str] = field(default_factory=list)
    patterns: list[str] = field(default_factory=list)
    note: str = ""
    missing_in_test_count: int = 0
    missing_in_live_count: int = 0
    value_diff_count: int = 0

    @property
    def hard_diff_count(self) -> int:
        return (
            self.missing_in_test_count
            + self.missing_in_live_count
            + self.value_diff_count
            + len(self.missing_columns_in_test)
        )

    @property
    def is_match(self) -> bool:
        return self.status == "MATCH"


@dataclass
class DataComparisonResult:
    live_path: str
    test_path: str
    sheets: list[SheetDataResult] = field(default_factory=list)
    missing_sheets_in_test: list[str] = field(default_factory=list)
    extra_sheets_in_test: list[str] = field(default_factory=list)

    @property
    def total_diffs(self) -> int:
        return sum(s.hard_diff_count for s in self.sheets) + len(self.missing_sheets_in_test)

    @property
    def is_match(self) -> bool:
        if self.missing_sheets_in_test:
            return False
        return all(s.is_match or s.status == "SKIP" for s in self.sheets)

    def summary(self) -> str:
        lines = [
            f"Data comparison: {self.live_path} vs {self.test_path}",
            f"Hard differences: {self.total_diffs}",
            f"Result: {'MATCH' if self.is_match else 'DIFFERENCES FOUND'}",
        ]
        if self.missing_sheets_in_test:
            lines.append(f"Missing sheets in /test: {', '.join(self.missing_sheets_in_test)}")
        if self.extra_sheets_in_test:
            lines.append(
                f"Extra sheets in /test (ignored): {', '.join(self.extra_sheets_in_test)}"
            )
        for sheet in self.sheets:
            lines.append("")
            lines.append(f"## Sheet: {sheet.sheet} [{sheet.status}]")
            if sheet.note:
                lines.append(f"  {sheet.note}")
            if sheet.key_fields:
                lines.append(f"  Row key: {', '.join(sheet.key_fields)}")
            lines.append(
                f"  Rows live={sheet.live_rows} test={sheet.test_rows} "
                f"matched={sheet.matched_rows}"
            )
            if sheet.ignored_extra_in_test:
                lines.append(
                    f"  Extra columns in /test (ignored): "
                    f"{', '.join(sheet.ignored_extra_in_test)}"
                )
            if sheet.missing_columns_in_test:
                lines.append(
                    f"  Missing columns in /test: {', '.join(sheet.missing_columns_in_test)}"
                )
            if sheet.patterns:
                lines.append("  Patterns:")
                for p in sheet.patterns:
                    lines.append(f"    - {p}")
            if sheet.missing_in_test:
                lines.append(
                    f"  Missing in /test ({sheet.missing_in_test_count}):"
                )
                for row in sheet.missing_in_test:
                    lines.append(f"    {row}")
            if sheet.missing_in_live:
                lines.append(
                    f"  Extra in /test only ({sheet.missing_in_live_count}):"
                )
                for row in sheet.missing_in_live:
                    lines.append(f"    {row}")
            if sheet.value_diffs:
                lines.append(f"  Value diffs ({sheet.value_diff_count}):")
                for d in sheet.value_diffs:
                    lines.append(f"    {d}")
            if sheet.soft_diffs:
                lines.append(
                    f"  Soft/cosmetic text diffs (not failing): {len(sheet.soft_diffs)}"
                )
        return "\n".join(lines)


def compare_workbooks_data(
    live_path: str | Path,
    test_path: str | Path,
    *,
    tolerance: float = 0.01,
    max_examples: int = 50,
) -> DataComparisonResult:
    live_path = str(live_path)
    test_path = str(test_path)
    result = DataComparisonResult(live_path=live_path, test_path=test_path)

    wb_live = load_workbook(live_path, data_only=True, read_only=True)
    wb_test = load_workbook(test_path, data_only=True, read_only=True)
    try:
        live_sheets = set(wb_live.sheetnames)
        test_sheets = set(wb_test.sheetnames)
        result.missing_sheets_in_test = sorted(live_sheets - test_sheets)
        result.extra_sheets_in_test = sorted(test_sheets - live_sheets)

        for name in wb_live.sheetnames:
            if name not in test_sheets:
                result.sheets.append(SheetDataResult(
                    sheet=name,
                    status="MISSING_TEST",
                    note="Sheet present on live, missing on /test.",
                ))
                continue
            result.sheets.append(
                _compare_sheet(
                    name,
                    list(wb_live[name].iter_rows(values_only=True)),
                    list(wb_test[name].iter_rows(values_only=True)),
                    tolerance=tolerance,
                    max_examples=max_examples,
                )
            )
        for name in wb_test.sheetnames:
            if name not in live_sheets:
                result.sheets.append(SheetDataResult(
                    sheet=name,
                    status="SKIP",
                    note="Extra sheet on /test only — ignored per parity rules.",
                ))
    finally:
        wb_live.close()
        wb_test.close()
    return result


def _compare_sheet(
    sheet: str,
    live_rows: list[tuple],
    test_rows: list[tuple],
    *,
    tolerance: float,
    max_examples: int,
) -> SheetDataResult:
    live_table = _extract_table(live_rows)
    test_table = _extract_table(test_rows)
    if live_table is None or test_table is None:
        return SheetDataResult(
            sheet=sheet,
            status="SKIP",
            note="Non-tabular sheet (no clear header row) — skipped for key-based compare.",
            live_rows=_nonempty_row_count(live_rows),
            test_rows=_nonempty_row_count(test_rows),
        )

    live_headers, live_data = live_table
    test_headers, test_data = test_table
    live_canon = [_canon_header(h) for h in live_headers]
    test_canon = [_canon_header(h) for h in test_headers]

    live_map = {c: i for i, c in enumerate(live_canon) if c}
    test_map = {c: i for i, c in enumerate(test_canon) if c}

    shared = sorted(set(live_map) & set(test_map))
    extra_test = sorted(set(test_map) - set(live_map))
    missing_test = sorted(set(live_map) - set(test_map))

    # Customer Activity sheets have last_order_date + account; key by account so a
    # different "last order" is a value diff, not a missing/extra row.
    if "last_order_date" in shared and "customer_account" in shared:
        key_fields = ["customer_account"]
    else:
        key_fields = _pick_key(shared) or _pick_key(sorted(set(live_map) | set(test_map)))
    if not key_fields:
        key_fields = [c for c in shared if c not in _SOFT_FIELDS][:4] or shared[:2]
    if not key_fields:
        return SheetDataResult(
            sheet=sheet,
            status="SKIP",
            note="Could not determine a row key.",
            ignored_extra_in_test=extra_test,
            missing_columns_in_test=missing_test,
        )

    compare_cols = [c for c in shared if c not in key_fields]
    live_by_key = _index_rows(live_data, live_map, key_fields)
    test_by_key = _index_rows(test_data, test_map, key_fields)

    missing_in_test: list[str] = []
    missing_in_live: list[str] = []
    value_diffs: list[str] = []
    soft_diffs: list[str] = []
    missing_test_meta: list[dict[str, Any]] = []
    missing_live_meta: list[dict[str, Any]] = []
    value_diff_meta: list[dict[str, Any]] = []

    for key, live_row in live_by_key.items():
        test_row = test_by_key.get(key)
        if test_row is None:
            missing_in_test.append(_row_label(key, key_fields, live_row, live_map))
            missing_test_meta.append(_row_meta(live_row, live_map))
            continue
        for col in compare_cols:
            lv = live_row[live_map[col]] if col in live_map else None
            tv = test_row[test_map[col]] if col in test_map else None
            if _values_equal(lv, tv, tolerance, soft=(col in _SOFT_FIELDS)):
                continue
            msg = f"{_fmt_key(key, key_fields)} | {col}: live={_repr(lv)} test={_repr(tv)}"
            meta = {"col": col, "live": lv, "test": tv, **_row_meta(live_row, live_map)}
            if col in _SOFT_FIELDS:
                soft_diffs.append(msg)
            else:
                value_diffs.append(msg)
                value_diff_meta.append(meta)

    for key, test_row in test_by_key.items():
        if key not in live_by_key:
            missing_in_live.append(_row_label(key, key_fields, test_row, test_map))
            missing_live_meta.append(_row_meta(test_row, test_map))

    patterns = _build_patterns(
        missing_test_meta=missing_test_meta,
        missing_live_meta=missing_live_meta,
        value_diff_meta=value_diff_meta,
        missing_in_test_n=len(missing_in_test),
        missing_in_live_n=len(missing_in_live),
        live_n=len(live_by_key),
        test_n=len(test_by_key),
    )

    out = SheetDataResult(
        sheet=sheet,
        status="MATCH",
        key_fields=list(key_fields),
        compared_columns=compare_cols,
        ignored_extra_in_test=extra_test,
        missing_columns_in_test=missing_test,
        live_rows=len(live_by_key),
        test_rows=len(test_by_key),
        matched_rows=len(set(live_by_key) & set(test_by_key)),
        missing_in_test=_cap(missing_in_test, max_examples),
        missing_in_live=_cap(missing_in_live, max_examples),
        value_diffs=_cap(value_diffs, max_examples),
        soft_diffs=_cap(soft_diffs, max_examples),
        patterns=patterns,
        missing_in_test_count=len(missing_in_test),
        missing_in_live_count=len(missing_in_live),
        value_diff_count=len(value_diffs),
    )
    hard = out.hard_diff_count
    out.status = "MATCH" if hard == 0 else "DIFF"
    return out


def _cap(items: list[str], max_examples: int) -> list[str]:
    if len(items) <= max_examples:
        return items
    return items[:max_examples] + [f"... +{len(items) - max_examples} more"]


def _nonempty_row_count(rows: list[tuple]) -> int:
    return sum(
        1 for r in rows
        if any(c is not None and str(c).strip() for c in (r or ()))
    )


def _canon_header(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    key = re.sub(r"\s+", " ", text).strip().lower()
    if key in _HEADER_ALIASES:
        return _HEADER_ALIASES[key]
    slug = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    return _HEADER_ALIASES.get(slug, slug)


def _extract_table(rows: list[tuple]) -> tuple[list[Any], list[tuple]] | None:
    """Find the first plausible header row and return (headers, data_rows)."""
    for idx, row in enumerate(rows):
        cells = list(row or ())
        non_null = [c for c in cells if c is not None and str(c).strip() != ""]
        if len(non_null) < 2:
            continue
        strish = sum(1 for c in non_null if isinstance(c, str) and not _looks_number(c))
        if strish < max(2, len(non_null) // 2):
            continue
        last = 0
        for i, c in enumerate(cells):
            if c is not None and str(c).strip() != "":
                last = i
        headers = list(cells[: last + 1])
        data = [tuple(r[: last + 1]) if r else tuple() for r in rows[idx + 1:]]
        data = [r for r in data if any(c is not None and str(c).strip() != "" for c in r)]
        return headers, data
    return None


def _looks_number(v: Any) -> bool:
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def _pick_key(available: list[str] | set[str]) -> list[str] | None:
    avail = set(available)
    for cand in _KEY_CANDIDATES:
        if all(c in avail for c in cand):
            return list(cand)
    return None


def _index_rows(
    rows: list[tuple],
    col_map: dict[str, int],
    key_fields: list[str],
) -> dict[tuple, tuple]:
    out: dict[tuple, tuple] = {}
    for row in rows:
        key = tuple(
            _norm_key_part(row[col_map[f]]) if f in col_map and col_map[f] < len(row) else ""
            for f in key_fields
        )
        if all(k == "" for k in key):
            continue
        if _is_skip_row(key):
            continue
        if key not in out:
            out[key] = row
    return out


def _is_skip_row(key: tuple) -> bool:
    for part in key:
        low = str(part).strip().lower()
        if not low:
            continue
        if low == "total" or low.startswith(_SKIP_KEY_PREFIXES):
            return True
    return False


def _norm_key_part(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    # Customer accounts often differ only by leading zeros (00011184 vs 11184).
    if s.isdigit():
        return str(int(s))
    return s


def _values_equal(a: Any, b: Any, tolerance: float, *, soft: bool = False) -> bool:
    if _is_empty(a) and _is_empty(b):
        return True
    if _is_empty(a) or _is_empty(b):
        return False

    if isinstance(a, datetime):
        a = a.date()
    if isinstance(b, datetime):
        b = b.date()
    if isinstance(a, date) and isinstance(b, date):
        return a == b
    # Same calendar day across datetime objects vs Excel text (ISO / MM/DD/YYYY).
    da, db = _date_bucket(a), _date_bucket(b)
    if da is not None and db is not None:
        return da == db
    if isinstance(a, date) or isinstance(b, date):
        return _norm_key_part(a) == _norm_key_part(b)

    na = _as_number(a)
    nb = _as_number(b)
    if na is not None and nb is not None:
        if math.isnan(na) and math.isnan(nb):
            return True
        return abs(na - nb) <= tolerance

    sa = str(a).strip()
    sb = str(b).strip()
    if sa == sb:
        return True
    if soft and _names_equivalent(sa, sb):
        return True
    return False


def _names_equivalent(a: str, b: str) -> bool:
    """'Meir Grego' ~= 'Grego, Meir'."""
    def parts(s: str) -> set[str]:
        s = s.replace(",", " ")
        return {p.lower() for p in s.split() if p}
    pa, pb = parts(a), parts(b)
    return bool(pa) and pa == pb


def _is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _as_number(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("$", "")
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _repr(v: Any) -> str:
    if v is None:
        return "None"
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float):
        return f"{v:.4f}"
    s = str(v)
    return s[:60] + "..." if len(s) > 60 else s


def _fmt_key(key: tuple, fields: list[str]) -> str:
    return ", ".join(f"{f}={k}" for f, k in zip(fields, key))


def _row_label(key: tuple, fields: list[str], row: tuple, col_map: dict[str, int]) -> str:
    base = _fmt_key(key, fields)
    extras = []
    for f in ("invoice_date", "order_date", "customer_account", "customer_name", "salesman_code"):
        if f in col_map and f not in fields:
            extras.append(f"{f}={_repr(row[col_map[f]] if col_map[f] < len(row) else None)}")
    return base if not extras else f"{base} ({', '.join(extras[:3])})"


def _row_meta(row: tuple, col_map: dict[str, int]) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    for f in ("invoice_date", "order_date", "customer_account", "salesman_code", "salesman_number"):
        if f in col_map and col_map[f] < len(row):
            meta[f] = row[col_map[f]]
    return meta


def _date_bucket(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # Excel serial dates sometimes arrive as floats (openpyxl data_only quirks).
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        serial = float(v)
        if 20000 <= serial <= 80000:  # ~1954–2119
            from datetime import timedelta
            # Excel's day 0 is 1899-12-30 for openpyxl/Windows.
            return (date(1899, 12, 30) + timedelta(days=int(serial))).isoformat()
        return None
    s = str(v).strip()
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def _build_patterns(
    *,
    missing_test_meta: list[dict[str, Any]],
    missing_live_meta: list[dict[str, Any]],
    value_diff_meta: list[dict[str, Any]],
    missing_in_test_n: int,
    missing_in_live_n: int,
    live_n: int,
    test_n: int,
) -> list[str]:
    patterns: list[str] = []

    if missing_in_test_n:
        pct = 100.0 * missing_in_test_n / live_n if live_n else 0
        patterns.append(
            f"{missing_in_test_n} live row(s) missing on /test "
            f"({pct:.0f}% of live rows)."
        )
        patterns.extend(_date_concentration(missing_test_meta, "missing on /test"))

    if missing_in_live_n:
        pct = 100.0 * missing_in_live_n / test_n if test_n else 0
        patterns.append(
            f"{missing_in_live_n} /test-only row(s) not on live "
            f"({pct:.0f}% of /test rows)."
        )
        patterns.extend(_date_concentration(missing_live_meta, "extra on /test"))

    if value_diff_meta:
        by_col: Counter[str] = Counter(m["col"] for m in value_diff_meta)
        top = by_col.most_common(5)
        patterns.append(
            "Value diffs by column: " + ", ".join(f"{c}={n}" for c, n in top)
        )
        patterns.extend(_date_concentration(value_diff_meta, "value-diff rows"))

        for col, _n in top[:3]:
            subset = [m for m in value_diff_meta if m["col"] == col]
            live_empty = sum(
                1 for m in subset if _is_empty(m["live"]) or _as_number(m["live"]) == 0
            )
            test_empty = sum(
                1 for m in subset if _is_empty(m["test"]) or _as_number(m["test"]) == 0
            )
            if live_empty == len(subset) and len(subset) >= 3:
                patterns.append(
                    f"All {col} diffs: live is empty/zero, /test has a value."
                )
            elif test_empty == len(subset) and len(subset) >= 3:
                patterns.append(
                    f"All {col} diffs: /test is empty/zero, live has a value."
                )

    return patterns


def _date_concentration(metas: list[dict[str, Any]], label: str) -> list[str]:
    if not metas:
        return []
    dates: list[str] = []
    for m in metas:
        for f in ("invoice_date", "order_date"):
            if f in m:
                b = _date_bucket(m[f])
                if b:
                    dates.append(b)
                    break
    if not dates:
        return []
    counts = Counter(dates)
    top_date, top_n = counts.most_common(1)[0]
    share = top_n / len(dates)
    out = []
    if share >= 0.4 and top_n >= 3:
        out.append(
            f"Common denominator ({label}): {top_n}/{len(dates)} rows share date {top_date}."
        )
    if len(counts) <= 8:
        detail = ", ".join(f"{d}={n}" for d, n in counts.most_common(8))
        out.append(f"Date breakdown ({label}): {detail}")
    else:
        detail = ", ".join(f"{d}={n}" for d, n in counts.most_common(5))
        out.append(f"Top dates ({label}): {detail}")
    return out
