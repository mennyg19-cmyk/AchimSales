"""Dump customer_activity All-sheet hard diffs from postfix parity xlsx."""
from openpyxl import load_workbook
from pathlib import Path

p = Path(".scratch/parity/20260804-193031-postfix")
out = p / "customer_activity_diffs.md"


def find_header_row(ws, must_have=("Customer Account",)):
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20), start=1):
        vals = [str(c or "").strip() for c in row]
        if all(m in vals for m in must_have):
            return i, vals
    return None, None


def sheet_map(path, sheet="All"):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr_row, hdr = find_header_row(ws)
    wb.close()
    if hdr is None:
        raise RuntimeError(f"no header in {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    ia = hdr.index("Customer Account")
    idate = hdr.index("Last Order Date")
    ipo = hdr.index("PO #")
    iso = hdr.index("Sales Order Number")
    iname = hdr.index("Customer Name") if "Customer Name" in hdr else None
    ism = hdr.index("Salesman") if "Salesman" in hdr else None
    outm = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= hdr_row:
            continue
        acct = str(r[ia] or "").strip()
        if not acct:
            continue

        def cell(idx):
            if idx is None:
                return ""
            v = r[idx]
            if v is None:
                return ""
            return str(v).strip()

        def dnorm(v):
            if v is None:
                return ""
            s = str(v).strip()
            if s.upper() in ("N/A", "NONE", ""):
                return ""
            if hasattr(v, "date"):
                return v.date().isoformat()
            return s[:10]

        po = cell(ipo)
        if po.upper() in ("N/A", "NONE", "NULL"):
            po = ""
        so = cell(iso)
        if so.upper() in ("N/A", "NONE", "NULL"):
            so = ""
        outm[acct] = {
            "name": cell(iname),
            "salesman": cell(ism),
            "date": dnorm(r[idate]),
            "po": po,
            "so": so,
        }
    wb.close()
    return outm


lm = sheet_map(p / "customer_activity__live.xlsx")
tm = sheet_map(p / "customer_activity__test.xlsx")
common = sorted(set(lm) & set(tm), key=lambda a: (lm[a]["name"].lower(), a))

rows = []
for a in common:
    L, T = lm[a], tm[a]
    diffs = []
    if L["date"] != T["date"]:
        diffs.append(("last_order_date", L["date"] or "(blank)", T["date"] or "(blank)"))
    if L["po"] != T["po"]:
        diffs.append(("po_number", L["po"] or "(blank)", T["po"] or "(blank)"))
    if L["so"] != T["so"]:
        diffs.append(("sales_order_number", L["so"] or "(blank)", T["so"] or "(blank)"))
    if diffs:
        rows.append((a, L, T, diffs))

lines = [
    "# Customer Activity diffs — All sheet",
    "",
    "Source: `.scratch/parity/20260804-193031-postfix/`",
    f"Accounts with any hard field diff: **{len(rows)}** (of {len(common)} matched)",
    "",
    "| Account | Customer | Salesman | Field | Live | Test |",
    "|---------|----------|----------|-------|------|------|",
]
for a, L, T, diffs in rows:
    for field, lv, tv in diffs:
        name = (L["name"] or T["name"]).replace("|", "/")
        sm = (L["salesman"] or T["salesman"]).replace("|", "/")
        lines.append(f"| {a} | {name} | {sm} | {field} | {lv} | {tv} |")

date_only = [r for r in rows if any(f == "last_order_date" for f, _, _ in r[3])]
lines += [
    "",
    f"## Accounts with last_order_date mismatch: {len(date_only)}",
    "",
    "| Account | Live date | Test date | Live SO | Test SO | Live PO | Test PO |",
    "|---------|-----------|-----------|---------|---------|---------|---------|",
]
for a, L, T, diffs in date_only:
    ld = L["date"] or "(blank)"
    td = T["date"] or "(blank)"
    lso = L["so"] or "(blank)"
    tso = T["so"] or "(blank)"
    lpo = L["po"] or "(blank)"
    tpo = T["po"] or "(blank)"
    lines.append(f"| {a} | {ld} | {td} | {lso} | {tso} | {lpo} | {tpo} |")

text = "\n".join(lines) + "\n"
out.write_text(text, encoding="utf-8")
print(f"wrote {out} accounts={len(rows)} date_mismatches={len(date_only)}")
