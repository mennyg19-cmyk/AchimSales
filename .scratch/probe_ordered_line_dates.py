"""Export ordered Full Data coverage diffs + OData header vs line created dates.

LIVE: headers by OrderCreationDateTime, then all lines.
TEST: lines by CreatedDateTime (SP). Line created date from OData =
SalesLineCDREntities.SysCreatedDateTime; header =
SalesOrderHeadersV3.OrderCreationDateTime.
"""
from __future__ import annotations

import csv
from collections import Counter
from datetime import date
from pathlib import Path

from urllib.parse import urlencode

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

from config.settings import (  # noqa: E402
    get_client_id,
    get_client_secret,
    get_d365_env_url,
    get_tenant_id,
)
from core.auth import get_d365_token  # noqa: E402

BATCH_SIZE = 25


def fetch_cross_company_batched(
    data_base: str,
    entity: str,
    token: str,
    filter_field: str,
    values: list[str],
    select: list[str],
) -> pd.DataFrame:
    """Batched OData with cross-company=true (core.odata does not send this)."""
    session = requests.Session()
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "OData-MaxVersion": "4.0",
        "OData-Version": "4.0",
    }
    chunks: list[pd.DataFrame] = []
    n_batches = max(1, (len(values) + BATCH_SIZE - 1) // BATCH_SIZE)

    def pull(filter_expr: str) -> None:
        query = {
            "$filter": filter_expr,
            "$select": ",".join(select),
            "cross-company": "true",
        }
        url = f"{data_base.rstrip('/')}/{entity}?{urlencode(query)}"
        while url:
            resp = session.get(url, headers=headers, timeout=120)
            if resp.status_code >= 400:
                raise requests.HTTPError(
                    f"{resp.status_code}: {resp.text[:400]}", response=resp)
            payload = resp.json()
            rows = payload.get("value") or []
            if rows:
                chunks.append(pd.DataFrame(rows))
            url = payload.get("@odata.nextLink")

    for i in range(0, len(values), BATCH_SIZE):
        batch = values[i : i + BATCH_SIZE]
        escaped = [str(v).replace("'", "''") for v in batch]
        filt = " or ".join(f"{filter_field} eq '{v}'" for v in escaped)
        try:
            pull(filt)
        except requests.HTTPError as exc:
            print(f"  {entity}: batch failed ({exc}); retrying singles…")
            for one in batch:
                try:
                    pull(f"{filter_field} eq '{str(one).replace(chr(39), chr(39)+chr(39))}'")
                except requests.HTTPError as one_exc:
                    print(f"    skip {one}: {one_exc}")
        bi = i // BATCH_SIZE + 1
        if bi == 1 or bi == n_batches or bi % 5 == 0:
            print(f"  {entity}: batch {bi}/{n_batches}")
    if not chunks:
        return pd.DataFrame()
    return pd.concat(chunks, ignore_index=True)

PARITY = Path(".scratch/parity/20260804-193031-postfix")
OUT_DIR = PARITY / "ordered_line_date_probe"
PERIOD_START = date(2026, 7, 1)
PERIOD_END = date(2026, 7, 31)
TZ_LIVE_EDGE = "2026-07-01"
TZ_TEST_EDGE = "2026-07-31"


def dnorm(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v).strip()
    if not s or s.upper() == "N/A":
        return ""
    # handle Z timestamps
    if "T" in s:
        return s[:10]
    return s[:10]


def parse_day(s: str) -> date | None:
    s = dnorm(s)
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def in_period(d: date | None) -> bool:
    return d is not None and PERIOD_START <= d <= PERIOD_END


def norm_line(v) -> str:
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s


def load_full(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = wb["Full Data"].iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(rows)]
    data = list(rows)
    wb.close()
    return hdr, data


def idx(hdr, *names):
    for n in names:
        if n in hdr:
            return hdr.index(n)
    raise KeyError(names)


def row_map(hdr, data):
    iso = idx(hdr, "SalesOrderNumber")
    iln = idx(hdr, "LineNumber")
    iitem = idx(hdr, "Item#", "Item Number")
    idate = idx(hdr, "OrderDate")
    iacct = idx(hdr, "CustomerAccount")
    iname = idx(hdr, "CustomerName", "SalesOrderName")
    istat = idx(hdr, "Status")
    out = {}
    for r in data:
        so = str(r[iso] or "").strip()
        ln = norm_line(r[iln])
        item = str(r[iitem] or "").strip()
        if not so:
            continue
        key = (so, ln, item)
        out[key] = {
            "so": so,
            "line": ln,
            "item": item,
            "order_date_report": dnorm(r[idate]),
            "account": str(r[iacct] or "").strip(),
            "name": str(r[iname] or "").strip() if iname is not None else "",
            "status": str(r[istat] or "").strip() if istat is not None else "",
        }
    return out


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    data_base = get_d365_env_url().rstrip("/")
    if not data_base.lower().endswith("/data"):
        data_base = data_base + "/data"

    print("Loading parity Full Data…")
    L = row_map(*load_full(PARITY / "ordered__live.xlsx"))
    T = row_map(*load_full(PARITY / "ordered__test.xlsx"))
    only_live = sorted(set(L) - set(T))
    only_test = sorted(set(T) - set(L))
    print(f"live_only={len(only_live)} test_only={len(only_test)}")

    export_path = OUT_DIR / "full_data_coverage_diffs.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "coverage_diffs"
    ws.append([
        "side", "sales_order", "line_number", "item", "report_order_date",
        "customer_account", "customer_name", "status", "tz_edge",
    ])
    for k in only_live:
        r = L[k]
        ws.append([
            "live_only", r["so"], r["line"], r["item"], r["order_date_report"],
            r["account"], r["name"], r["status"],
            "yes" if r["order_date_report"] == TZ_LIVE_EDGE else "no",
        ])
    for k in only_test:
        r = T[k]
        ws.append([
            "test_only", r["so"], r["line"], r["item"], r["order_date_report"],
            r["account"], r["name"], r["status"],
            "yes" if r["order_date_report"] == TZ_TEST_EDGE else "no",
        ])
    wb.save(export_path)
    print(f"Wrote {export_path}")

    sos = sorted({L[k]["so"] for k in only_live} | {T[k]["so"] for k in only_test})
    print(f"Unique SOs: {len(sos)}  OData base: {data_base}")

    print("Acquiring token…")
    token = get_d365_token(get_tenant_id(), get_client_id(), get_client_secret(), get_d365_env_url())

    print("Fetching headers (OrderCreationDateTime, cross-company)…")
    hdr_df = fetch_cross_company_batched(
        data_base, "SalesOrderHeadersV3", token,
        filter_field="SalesOrderNumber",
        values=sos,
        select=["SalesOrderNumber", "OrderCreationDateTime", "InvoiceCustomerAccountNumber", "dataAreaId"],
    )
    header_by_so: dict[str, dict] = {}
    for _, row in hdr_df.iterrows() if hdr_df is not None and not hdr_df.empty else []:
        so = str(row.get("SalesOrderNumber") or "").strip()
        if so and so not in header_by_so:
            header_by_so[so] = {
                "header_order_creation": dnorm(row.get("OrderCreationDateTime")),
                "account": str(row.get("InvoiceCustomerAccountNumber") or "").strip(),
                "data_area": str(row.get("dataAreaId") or "").strip(),
            }
    print(f"Headers: {len(header_by_so)} / {len(sos)} SOs")
    missing_hdr = [s for s in sos if s not in header_by_so]
    if missing_hdr:
        print(f"  missing headers ({len(missing_hdr)}): {missing_hdr[:12]}")

    print("Fetching lines (SalesLineCDREntities.SysCreatedDateTime, cross-company)…")
    line_df = fetch_cross_company_batched(
        data_base, "SalesLineCDREntities", token,
        filter_field="SalesId",
        values=sos,
        select=["SalesId", "LineNum", "ItemId", "SysCreatedDateTime", "CreatedOn"],
    )
    line_meta: dict[tuple[str, str, str], dict] = {}
    if line_df is not None and not line_df.empty:
        for _, row in line_df.iterrows():
            so = str(row.get("SalesId") or "").strip()
            ln = norm_line(row.get("LineNum"))
            item = str(row.get("ItemId") or "").strip()
            created = dnorm(row.get("SysCreatedDateTime") or row.get("CreatedOn"))
            line_meta[(so, ln, item)] = {"line_created": created}
            line_meta.setdefault((so, ln, ""), {"line_created": created})
    print(f"Line keys: {len(line_meta)}")

    buckets = Counter()
    probe_rows = []

    def classify(side: str, rec: dict, key: tuple) -> str:
        so, ln, item = key
        hdr = header_by_so.get(so, {})
        header_day = parse_day(hdr.get("header_order_creation") or "")
        meta = line_meta.get(key) or line_meta.get((so, ln, "")) or {}
        line_day = parse_day(meta.get("line_created") or "")
        header_in = in_period(header_day)
        line_in = in_period(line_day)

        if not hdr:
            return "no_header_odata"
        if not meta.get("line_created"):
            return "no_line_created_odata"

        if side == "live_only":
            # header in July, line created outside July → LIVE keeps, TEST drops
            if header_in and not line_in:
                return "explained_header_in_line_out"
            if header_in and line_in:
                return "both_in_period_other_cause"
            return "unexplained_live_only"
        # test_only: line in July, header outside → TEST keeps, LIVE drops
        if line_in and not header_in:
            return "explained_line_in_header_out"
        if line_in and header_in:
            return "both_in_period_other_cause"
        return "unexplained_test_only"

    for side, keys, src in (("live_only", only_live, L), ("test_only", only_test, T)):
        for key in keys:
            rec = src[key]
            so, ln, item = key
            hdr = header_by_so.get(so, {})
            meta = line_meta.get(key) or line_meta.get((so, ln, "")) or {}
            bucket = classify(side, rec, key)
            buckets[f"{side}:{bucket}"] += 1
            hday = hdr.get("header_order_creation", "")
            lday = meta.get("line_created", "")
            probe_rows.append({
                "side": side,
                "sales_order": so,
                "line_number": ln,
                "item": item,
                "report_order_date": rec["order_date_report"],
                "header_order_creation": hday,
                "line_sys_created": lday,
                "header_in_july": "yes" if in_period(parse_day(hday)) else "no",
                "line_in_july": "yes" if in_period(parse_day(lday)) else "no",
                "dates_differ": "yes" if hday and lday and hday != lday else ("no" if hday and lday else ""),
                "bucket": bucket,
                "customer_account": rec["account"],
                "customer_name": rec["name"],
                "status": rec["status"],
                "tz_edge": (
                    "yes" if (
                        (side == "live_only" and rec["order_date_report"] == TZ_LIVE_EDGE)
                        or (side == "test_only" and rec["order_date_report"] == TZ_TEST_EDGE)
                    ) else "no"
                ),
            })

    fields = list(probe_rows[0].keys()) if probe_rows else []
    probe_csv = OUT_DIR / "odata_header_vs_line_dates.csv"
    with probe_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(probe_rows)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "odata_probe"
    if fields:
        ws2.append(fields)
        for row in probe_rows:
            ws2.append([row.get(c, "") for c in fields])
    ws3 = wb2.create_sheet("summary")
    ws3.append(["bucket", "count"])
    for k, n in sorted(buckets.items(), key=lambda x: (-x[1], x[0])):
        ws3.append([k, n])

    nontz = [r for r in probe_rows if r["tz_edge"] == "no"]
    explained = [r for r in nontz if r["bucket"] in (
        "explained_header_in_line_out", "explained_line_in_header_out")]
    dates_differ = [r for r in nontz if r["dates_differ"] == "yes"]
    ws3.append([])
    ws3.append(["non_tz_rows", len(nontz)])
    ws3.append(["non_tz_explained_by_date_gate", len(explained)])
    ws3.append(["non_tz_header_ne_line_created", len(dates_differ)])
    out_xlsx = OUT_DIR / "odata_header_vs_line_dates.xlsx"
    wb2.save(out_xlsx)

    print(f"\nWrote {probe_csv}")
    print(f"Wrote {out_xlsx}")
    print("\n=== BUCKET COUNTS ===")
    for k, n in sorted(buckets.items(), key=lambda x: (-x[1], x[0])):
        print(f"  {k}: {n}")
    print(f"\nnon-TZ rows: {len(nontz)}")
    print(f"non-TZ explained by header/line date gate: {len(explained)} "
          f"({100 * len(explained) / max(1, len(nontz)):.1f}%)")
    print(f"non-TZ where header date != line SysCreatedDateTime: {len(dates_differ)}")

    unex = [r for r in nontz if r["bucket"] not in (
        "explained_header_in_line_out", "explained_line_in_header_out")]
    print(f"\nnon-TZ not explained by date gate: {len(unex)}")
    for r in unex[:20]:
        print(
            f"  {r['side']} {r['sales_order']} L{r['line_number']} {r['item']} "
            f"report={r['report_order_date']} header={r['header_order_creation']} "
            f"line={r['line_sys_created']} bucket={r['bucket']}"
        )


if __name__ == "__main__":
    main()
