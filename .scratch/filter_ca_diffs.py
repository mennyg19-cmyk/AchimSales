"""CA diffs excluding accounts where SO and PO both match."""
from openpyxl import load_workbook
from pathlib import Path

p = Path(".scratch/parity/20260804-193031-postfix")


def find_header_row(ws, must_have=("Customer Account",)):
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20), start=1):
        vals = [str(c or "").strip() for c in row]
        if all(m in vals for m in must_have):
            return i, vals
    return None, None


def sheet_map(path, sheet="All"):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr_row, hdr = find_header_row(ws)
    wb.close()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    ia = hdr.index("Customer Account")
    idate = hdr.index("Last Order Date")
    ipo = hdr.index("PO #")
    iso = hdr.index("Sales Order Number")
    iname = hdr.index("Customer Name") if "Customer Name" in hdr else None
    ism = hdr.index("Salesman") if "Salesman" in hdr else None
    outm = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= hdr_row:
            continue
        acct = str(r[ia] or "").strip()
        if not acct:
            continue

        def cell(idx):
            if idx is None:
                return ""
            v = r[idx]
            return "" if v is None else str(v).strip()

        def dnorm(v):
            if v is None:
                return ""
            s = str(v).strip()
            if s.upper() in ("N/A", "NONE", ""):
                return ""
            if hasattr(v, "date"):
                return v.date().isoformat()
            return s[:10]

        po = cell(ipo)
        if po.upper() in ("N/A", "NONE", "NULL"):
            po = ""
        so = cell(iso)
        if so.upper() in ("N/A", "NONE", "NULL"):
            so = ""
        outm[acct] = {
            "name": cell(iname),
            "salesman": cell(ism),
            "date": dnorm(r[idate]),
            "po": po,
            "so": so,
        }
    wb.close()
    return outm


lm = sheet_map(p / "customer_activity__live.xlsx")
tm = sheet_map(p / "customer_activity__test.xlsx")
common = sorted(set(lm) & set(tm), key=lambda a: (lm[a]["name"].lower(), a))

left = []
dropped = 0
for a in common:
    L, T = lm[a], tm[a]
    any_diff = L["date"] != T["date"] or L["so"] != T["so"] or L["po"] != T["po"]
    if not any_diff:
        continue
    if L["so"] == T["so"] and L["po"] == T["po"]:
        dropped += 1
        continue
    left.append((a, L, T))

print(f"left={len(left)} dropped_same_so_and_po={dropped}")
print()
print("| Account | Customer | Salesman | Live date | Test date | Live SO | Test SO | Live PO | Test PO |")
print("|---------|----------|----------|-----------|-----------|---------|---------|---------|---------|")
for a, L, T in left:
    name = (L["name"] or "").replace("|", "/")
    sm = (L["salesman"] or "").replace("|", "/")
    ld = L["date"] or "(blank)"
    td = T["date"] or "(blank)"
    lso = L["so"] or "(blank)"
    tso = T["so"] or "(blank)"
    lpo = L["po"] or "(blank)"
    tpo = T["po"] or "(blank)"
    print(f"| {a} | {name} | {sm} | {ld} | {td} | {lso} | {tso} | {lpo} | {tpo} |")
