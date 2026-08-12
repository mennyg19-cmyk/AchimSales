"""Try SalesLineCDREntities entity set for SysCreatedDateTime."""
from dotenv import load_dotenv
load_dotenv()
import requests
from config.settings import get_client_id, get_client_secret, get_d365_env_url, get_tenant_id
from core.auth import get_d365_token

env = get_d365_env_url().rstrip("/")
token = get_d365_token(get_tenant_id(), get_client_id(), get_client_secret(), env)
headers = {
    "Authorization": f"Bearer {token}",
    "Accept": "application/json",
    "OData-MaxVersion": "4.0",
    "OData-Version": "4.0",
}

# discover field names
r = requests.get(
    f"{env}/data/SalesLineCDREntities",
    params={"$top": "1", "cross-company": "true"},
    headers=headers,
    timeout=60,
)
print("top1", r.status_code)
if r.status_code != 200:
    print(r.text[:500])
else:
    row = r.json()["value"][0]
    keys = sorted(row.keys())
    print("keys sample", [k for k in keys if any(x in k.lower() for x in ("sales", "item", "line", "creat", "date", "qty", "invent"))])

# try common sales id field names against a known SO from coverage export
import csv
from pathlib import Path
sos = []
with Path(".scratch/parity/20260804-193031-postfix/ordered_line_date_probe/full_data_coverage_diffs.xlsx").open("rb") as f:
    pass
from openpyxl import load_workbook
wb = load_workbook(".scratch/parity/20260804-193031-postfix/ordered_line_date_probe/full_data_coverage_diffs.xlsx", read_only=True)
rows = wb.active.iter_rows(values_only=True)
hdr = next(rows)
iso = list(hdr).index("sales_order")
for i, r in enumerate(rows):
    sos.append(str(r[iso]))
    if i > 5:
        break
wb.close()
so = sos[0]
print("probing SO", so)

for field in ("SalesId", "SalesOrderNumber", "salesId", "OrderNumber"):
    r = requests.get(
        f"{env}/data/SalesLineCDREntities",
        params={
            "$filter": f"{field} eq '{so}'",
            "$top": "3",
            "cross-company": "true",
        },
        headers=headers,
        timeout=60,
    )
    print(field, r.status_code, r.text[:350].replace("\n", " "))
