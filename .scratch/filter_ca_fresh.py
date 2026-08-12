"""CA live↔test: apply Aug-4 sign-off noise cuts, list leftovers."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

p = Path(r"D:/Projects/Achim/AchimSales/.scratch/parity/20260806-195804-customer_activity")
TODAY = date(2026, 8, 6).isoformat()  # run day Eastern


def find_header_row(ws, must_have=("Customer Account",)):
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20), start=1):
        vals = [str(c or "").strip() for c in row]
        if all(m in vals for m in must_have):
            return i, vals
    return None, None


def sheet_map(path, sheet="All"):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr_row, hdr = find_header_row(ws)
    wb.close()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    ia = hdr.index("Customer Account")
    idate = hdr.index("Last Order Date")
    ipo = hdr.index("PO #")
    iso = hdr.index("Sales Order Number")
    iname = hdr.index("Customer Name") if "Customer Name" in hdr else None
    ism = hdr.index("Salesman") if "Salesman" in hdr else None
    outm = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= hdr_row:
            continue
        acct = str(r[ia] or "").strip()
        if not acct:
            continue

        def cell(idx):
            if idx is None:
                return ""
            v = r[idx]
            return "" if v is None else str(v).strip()

        def dnorm(v):
            if v is None:
                return ""
            s = str(v).strip()
            if s.upper() in ("N/A", "NONE", ""):
                return ""
            if hasattr(v, "date"):
                return v.date().isoformat()
            return s[:10]

        po = cell(ipo)
        if po.upper() in ("N/A", "NONE", "NULL"):
            po = ""
        so = cell(iso)
        if so.upper() in ("N/A", "NONE", "NULL"):
            so = ""
        outm[acct] = {
            "name": cell(iname),
            "salesman": cell(ism),
            "date": dnorm(r[idate]),
            "po": po,
            "so": so,
        }
    wb.close()
    return outm


lm = sheet_map(p / "customer_activity__live.xlsx")
tm = sheet_map(p / "customer_activity__test.xlsx")
print(f"rows live={len(lm)} test={len(tm)} common={len(set(lm)&set(tm))}")
print(f"live_only={sorted(set(lm)-set(tm))[:20]} count={len(set(lm)-set(tm))}")
print(f"test_only={sorted(set(tm)-set(lm))[:20]} count={len(set(tm)-set(lm))}")

raw = []
for a in sorted(set(lm) & set(tm), key=lambda x: (lm[x]["name"].lower(), x)):
    L, T = lm[a], tm[a]
    if L["date"] != T["date"] or L["so"] != T["so"] or L["po"] != T["po"]:
        raw.append((a, L, T))

same_so_po = []
blank_po_test = []
today_dated = []
left = []
for a, L, T in raw:
    if L["so"] == T["so"] and L["po"] == T["po"]:
        same_so_po.append(a)
        continue
    if T["po"] == "" and L["po"] != "":
        blank_po_test.append(a)
        continue
    if L["date"] == TODAY or T["date"] == TODAY:
        today_dated.append(a)
        continue
    left.append((a, L, T))

print()
print(f"raw_diff_accounts={len(raw)}")
print(f"dropped_same_so_and_po={len(same_so_po)}")
print(f"dropped_blank_po_on_test={len(blank_po_test)}")
print(f"dropped_today_dated={len(today_dated)} today={TODAY}")
print(f"LEFT={len(left)}")
print()
print("| Account | Customer | Salesman | Live date | Test date | Live SO | Test SO | Live PO | Test PO |")
print("|---------|----------|----------|-----------|-----------|---------|---------|---------|---------|")
for a, L, T in left:
    name = (L["name"] or "").replace("|", "/")
    sm = (L["salesman"] or "").replace("|", "/")
    print(
        f"| {a} | {name} | {sm} | {L['date'] or '(blank)'} | {T['date'] or '(blank)'} | "
        f"{L['so'] or '(blank)'} | {T['so'] or '(blank)'} | "
        f"{L['po'] or '(blank)'} | {T['po'] or '(blank)'} |"
    )
