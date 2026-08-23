from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

folder = Path(".scratch/parity/20260805-111000-po-audit-retest")

def sheet(path, name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(rows)]
    data = list(rows)
    wb.close()
    return hdr, data

def idx(hdr, *names):
    for n in names:
        if n in hdr:
            return hdr.index(n)
    return None

def dnorm(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.date().isoformat()
    return str(v).strip()[:10]

def norm_line(v):
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        return s
    return s

lh, lr = sheet(folder / "ordered__live.xlsx", "Full Data")
th, tr = sheet(folder / "ordered__test.xlsx", "Full Data")
lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")
lln, tln = idx(lh, "LineNumber"), idx(th, "LineNumber")
li, ti = idx(lh, "Item#", "Item Number"), idx(th, "Item#", "Item Number")
ld, td = idx(lh, "OrderDate"), idx(th, "OrderDate")
lac, tac = idx(lh, "CustomerAccount"), idx(th, "CustomerAccount")
lnm, tnm = idx(lh, "CustomerName", "SalesOrderName"), idx(th, "CustomerName", "SalesOrderName")
lst, tst = idx(lh, "Status"), idx(th, "Status")
lpo, tpo = idx(lh, "PO #", "PO#"), idx(th, "PO #", "PO#")

def keymap(data, iso, iln, iitem):
    m = {}
    for r in data:
        so = str(r[iso] or "").strip()
        if not so or so.upper() == "TOTAL":
            continue
        m[(so, norm_line(r[iln]), str(r[iitem] or "").strip())] = r
    return m

L = keymap(lr, lso, lln, li)
T = keymap(tr, tso, tln, ti)
only_t = set(T) - set(L)
remain = [k for k in only_t if dnorm(T[k][td]) != "2026-07-31" and k[1] not in ("0", "0.0")]

by_so = defaultdict(list)
for k in remain:
    by_so[k[0]].append(k)

print(f"SOs: {len(by_so)}  lines: {len(remain)}\n")
print(f"{'SalesOrder':<14} {'Acct':<12} {'Lines':>5} {'OrderDate':<12} {'Status':<14} {'PO':<20} Customer")
print("-" * 110)
rows_out = []
for so in sorted(by_so, key=lambda s: (-len(by_so[s]), s)):
    keys = by_so[so]
    r0 = T[keys[0]]
    acct = str(r0[tac] or "").strip()
    name = str(r0[tnm] or "").strip()[:40]
    dates = sorted({dnorm(T[k][td]) for k in keys})
    statuses = sorted({str(T[k][tst] or "").strip() for k in keys})
    pos = sorted({str(T[k][tpo] or "").strip() for k in keys if tpo is not None})
    po = pos[0] if len(pos) == 1 else ",".join(pos[:3])
    print(f"{so:<14} {acct:<12} {len(keys):>5} {dates[0]:<12} {statuses[0]:<14} {po:<20} {name}")
    rows_out.append({
        "sales_order": so,
        "customer_account": acct,
        "customer_name": name,
        "line_count": len(keys),
        "order_dates": ",".join(dates),
        "statuses": ",".join(statuses),
        "po": po,
    })

out = folder / "ordered_test_only_21_sos.csv"
import csv
with out.open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
    w.writeheader()
    w.writerows(rows_out)
print(f"\nWrote {out}")
