"""What's left on ordered + invoiced after agreed exclusions."""
from collections import Counter, defaultdict
from openpyxl import load_workbook
from pathlib import Path

p = Path(".scratch/parity/20260804-193031-postfix")


def load_sheet(path, sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    hdr = [str(c or "").strip() for c in rows[0]]
    return hdr, rows[1:]


def idx(hdr, *names):
    for n in names:
        if n in hdr:
            return hdr.index(n)
    return None


def num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def dnorm(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v).strip()
    return "" if not s or s.upper() == "N/A" else s[:10]


# ---- ORDERED Full Data ----
print("=== ORDERED Full Data ===")
lh, lr = load_sheet(p / "ordered__live.xlsx", "Full Data")
th, tr = load_sheet(p / "ordered__test.xlsx", "Full Data")

def keymap(hdr, rows, side):
    iso = idx(hdr, "SalesOrderNumber")
    iln = idx(hdr, "LineNumber")
    iitem = idx(hdr, "Item#", "Item Number")
    out = {}
    for r in rows:
        k = (str(r[iso] or ""), str(r[iln] or ""), str(r[iitem] or ""))
        out[k] = r
    return out

L = keymap(lh, lr, "live")
T = keymap(th, tr, "test")
lso_i = idx(lh, "SalesOrderNumber")
ldate = idx(lh, "OrderDate")
tdate = idx(th, "OrderDate")
lstat, tstat = idx(lh, "Status"), idx(th, "Status")
lrel, trel = idx(lh, "QtyReleased"), idx(th, "QtyReleased", "QTY Shipping")
lship = idx(lh, "QtyShipped")
lord, tord = idx(lh, "QtyOrdered"), idx(th, "QtyOrdered")
lcanc, tcanc = idx(lh, "QtyCancelled"), idx(th, "QtyCancelled")
lopen = idx(lh, "QtyOpen")
# money
ldols = {
    "ordered$": (idx(lh, "Ordered $"), idx(th, "Ordered $")),
    "shipped$": (idx(lh, "Shipped $"), idx(th, "Shipped $")),
    "cancelled$": (idx(lh, "Cancelled $"), idx(th, "Cancelled $")),
    "released$": (idx(lh, "Released $"), idx(th, "Released $")),
    "open$": (idx(lh, "Open $"), idx(th, "Open $")),
}

common = set(L) & set(T)
only_live = set(L) - set(T)
only_test = set(T) - set(L)

# TZ edges for ordered last_month: 2026-07-01 live-only, 2026-07-31 test-only (primary)
TZ_LIVE = "2026-07-01"
TZ_TEST = "2026-07-31"

only_live_nontz = [k for k in only_live if dnorm(L[k][ldate]) != TZ_LIVE]
only_test_nontz = [k for k in only_test if dnorm(T[k][tdate]) != TZ_TEST]
print(f"coverage common={len(common)} live_only={len(only_live)} (non-TZ={len(only_live_nontz)}) "
      f"test_only={len(only_test)} (non-TZ={len(only_test_nontz)})")
print("live_only non-TZ dates", Counter(dnorm(L[k][ldate]) for k in only_live_nontz).most_common(8))
print("test_only non-TZ dates", Counter(dnorm(T[k][tdate]) for k in only_test_nontz).most_common(8))

# Status: which side
status_pairs = Counter()
for k in common:
    ls = str(L[k][lstat] or "").strip()
    ts = str(T[k][tstat] or "").strip()
    if ls != ts:
        status_pairs[(ls, ts)] += 1
print("ALL status mismatch pairs (live -> test):", status_pairs.most_common())
print("Total status mismatches:", sum(status_pairs.values()))

# After dropping known status pairs, anything left?
KNOWN_STATUS = {
    ("Cancelled", "Canceled"),
    ("Open", "Open Order"),
    ("In Process", "Open Order"),
    ("InProcess", "Open Order"),
}
other_status = {pair: n for pair, n in status_pairs.items() if pair not in KNOWN_STATUS}
print("status left after known labels:", other_status or "NONE")

# Value diffs on common AFTER ignoring released-mapping and known status
# "released mapping ok" = live_rel+ship == test_rel
left_fields = Counter()
examples = defaultdict(list)
for k in common:
    lr_, tr_ = L[k], T[k]
    ls = str(lr_[lstat] or "").strip()
    ts = str(tr_[tstat] or "").strip()
    # skip if only status would differ (we ignore known)
    money_or_qty = False
    # qty ordered/cancelled (not released — mapping intentional)
    if abs(num(lr_[lord]) - num(tr_[tord])) > 1e-6:
        left_fields["qty_ordered"] += 1
        money_or_qty = True
        if len(examples["qty_ordered"]) < 3:
            examples["qty_ordered"].append((k, num(lr_[lord]), num(tr_[tord]), ls, ts))
    if abs(num(lr_[lcanc]) - num(tr_[tcanc])) > 1e-6:
        left_fields["qty_cancelled"] += 1
        money_or_qty = True
    for name, (li, ti) in ldols.items():
        if li is None or ti is None:
            continue
        if abs(num(lr_[li]) - num(tr_[ti])) > 0.009:
            left_fields[name] += 1
            money_or_qty = True
            if len(examples[name]) < 3:
                examples[name].append((k, num(lr_[li]), num(tr_[ti]), ls, ts, dnorm(lr_[ldate])))
    # released alone mismatch that is NOT explained by +shipped?
    l_sum = num(lr_[lrel]) + (num(lr_[lship]) if lship is not None else 0)
    t_r = num(tr_[trel])
    if abs(l_sum - t_r) > 1e-6:
        left_fields["released_NOT_explained_by_shipped"] += 1

print("common-line value diffs (ignoring status labels; released checked vs +shipped):", dict(left_fields))
for k, xs in examples.items():
    print(" ex", k, xs)

# By Order leftover after TZ + PO ignore
print("\n=== ORDERED By Order leftover ===")
lh, lr = load_sheet(p / "ordered__live.xlsx", "By Order")
th, tr = load_sheet(p / "ordered__test.xlsx", "By Order")
lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")
ldate, tdate = idx(lh, "OrderDate"), idx(th, "OrderDate")
lstat, tstat = idx(lh, "Status"), idx(th, "Status")
# qty cols
qty_names = [h for h in lh if h.startswith("Qty") or h.endswith(" $")]
print("by order live qty/money cols", qty_names)
L = {str(r[lso] or ""): r for r in lr if r[lso]}
T = {str(r[tso] or ""): r for r in tr if r[tso]}
common = set(L) & set(T)
only_live = [k for k in set(L)-set(T) if dnorm(L[k][ldate]) != TZ_LIVE]
only_test = [k for k in set(T)-set(L) if dnorm(T[k][tdate]) != TZ_TEST]
print(f"by order non-TZ live_only={len(only_live)} test_only={len(only_test)}")
print("live_only dates", Counter(dnorm(L[k][ldate]) for k in only_live).most_common(5))
print("test_only dates", Counter(dnorm(T[k][tdate]) for k in only_test).most_common(5))

# status on by order
sp = Counter()
for k in common:
    ls = str(L[k][lstat] or "").strip() if lstat is not None else ""
    ts = str(T[k][tstat] or "").strip() if tstat is not None else ""
    if ls != ts:
        sp[(ls, ts)] += 1
print("By Order status pairs live->test:", sp.most_common(10))

# ---- INVOICED Full Details leftover ----
print("\n=== INVOICED Full Details leftover ===")
lh, lr = load_sheet(p / "invoiced__live.xlsx", "Full Details")
th, tr = load_sheet(p / "invoiced__test.xlsx", "Full Details")
li, ti = idx(lh, "InvoiceNumber"), idx(th, "InvoiceNumber")
ld, td = idx(lh, "InvoiceDate"), idx(th, "InvoiceDate")
lacc = idx(lh, "CustomerAccount")
fields = {
    "total": (idx(lh, "Total Invoice"), idx(th, "Total Invoice")),
    "subtotal": (idx(lh, "SubTotal Invoices"), idx(th, "SubTotal Invoices")),
    "tariff": (idx(lh, "Tariff Charges"), idx(th, "Tariff Charges")),
    "cc": (idx(lh, "CC Charges"), idx(th, "CC Charges")),
    "freight": (idx(lh, "Freight Charges"), idx(th, "Freight Charges")),
}
L = {str(r[li] or ""): r for r in lr if r[li]}
T = {str(r[ti] or ""): r for r in tr if r[ti]}
TODAY = "2026-08-04"
only_live = [i for i in set(L)-set(T) if dnorm(L[i][ld]) != TODAY]
only_test = list(set(T)-set(L))
print(f"live_only non-today={len(only_live)} test_only={len(only_test)}")
if only_live:
    print(" live_only samples", [(i, dnorm(L[i][ld]), L[i][lacc]) for i in only_live[:10]])
if only_test:
    print(" test_only samples", [(i, dnorm(T[i][td]), T[i][idx(th,'CustomerAccount')]) for i in only_test[:10]])

money_invs = []
for inv in set(L) & set(T):
    diffs = []
    for name, (a, b) in fields.items():
        lv, tv = num(L[inv][a]), num(T[inv][b])
        if abs(lv - tv) > 0.009:
            diffs.append((name, lv, tv))
    if diffs:
        money_invs.append((inv, dnorm(L[inv][ld]), str(L[inv][lacc]), diffs))

print(f"shared invoices with money diffs: {len(money_invs)}")
# group by which fields
by_pat = Counter()
for inv, dt, acc, diffs in money_invs:
    names = tuple(sorted(n for n,_,_ in diffs))
    by_pat[names] += 1
print("money patterns:", dict(by_pat))
print("ALL money invoices:")
for inv, dt, acc, diffs in sorted(money_invs, key=lambda x: (x[2], x[0])):
    dstr = ", ".join(f"{n}: L={lv} T={tv}" for n,lv,tv in diffs)
    print(f"  {acc} {inv} {dt} | {dstr}")

# Credits leftover similarly
print("\n=== INVOICED Credits leftover ===")
lh, lr = load_sheet(p / "invoiced__live.xlsx", "Credits")
th, tr = load_sheet(p / "invoiced__test.xlsx", "Credits")
li, ti = idx(lh, "InvoiceNumber"), idx(th, "InvoiceNumber")
ld, td = idx(lh, "InvoiceDate"), idx(th, "InvoiceDate")
ltot, ttot = idx(lh, "Total Invoice"), idx(th, "Total Invoice")
lacc = idx(lh, "CustomerAccount")
L = {str(r[li] or ""): r for r in lr if r[li]}
T = {str(r[ti] or ""): r for r in tr if r[ti]}
# credits can have duplicate invoice numbers (reversal pairs) - count by (inv, total) or just list
print(f"credit rows live={len(lr)} test={len(tr)} unique inv live={len(L)} test={len(T)}")
only_live = [i for i in set(L)-set(T) if dnorm(L[i][ld]) != TODAY]
only_test = list(set(T)-set(L))
print(f"credits unique live_only non-today={len(only_live)} test_only={len(only_test)}")
money = []
for inv in set(L)&set(T):
    if abs(num(L[inv][ltot]) - num(T[inv][ttot])) > 0.009:
        money.append((str(L[inv][lacc]), inv, dnorm(L[inv][ld]), num(L[inv][ltot]), num(T[inv][ttot])))
print(f"credits shared money diffs (by unique inv key - may undercount pairs): {len(money)}")
for row in money[:20]:
    print(" ", row)
