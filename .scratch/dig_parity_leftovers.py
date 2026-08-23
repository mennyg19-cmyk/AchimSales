"""Extra dig on remaining invoiced money + ordered coverage after noise cuts."""
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

folder = Path(".scratch/parity/20260805-111000-po-audit-retest")


def sheet(path, name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(rows)]
    data = list(rows)
    wb.close()
    return hdr, data


def idx(hdr, *names):
    for n in names:
        if n in hdr:
            return hdr.index(n)
    lower = {h.lower(): i for i, h in enumerate(hdr)}
    for n in names:
        if n.lower() in lower:
            return lower[n.lower()]
    return None


def num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def dnorm(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v).strip()
    return s[:10] if s else ""


# --- invoiced money detail ---
lh, lr = sheet(folder / "invoiced__live.xlsx", "Full Details")
th, tr = sheet(folder / "invoiced__test.xlsx", "Full Details")
li, ti = idx(lh, "InvoiceNumber"), idx(th, "InvoiceNumber")
cols = ["InvoiceDate", "CustomerAccount", "Total Invoice", "SubTotal Invoices",
        "Tariff Charges", "CC Charges", "Freight Charges", "SalesOrderNumber"]
L = {str(r[li] or "").strip(): r for r in lr if r[li]}
T = {str(r[ti] or "").strip(): r for r in tr if r[ti]}
print("=== money leftovers ===")
for inv in ("IN00963267", "IN00828240"):
    a, b = L.get(inv), T.get(inv)
    if not a or not b:
        print(inv, "missing", a is None, b is None)
        continue
    print(inv)
    for c in cols:
        ia, ib = idx(lh, c), idx(th, c)
        print(f"  {c}: live={a[ia]!r} test={b[ib]!r}")

# Audit sheet row counts
for side, path in (("live", "invoiced__live.xlsx"), ("test", "invoiced__test.xlsx")):
    wb = load_workbook(folder / path, read_only=True, data_only=True)
    name = "Audit - Reversals"
    n = wb[name].max_row - 1 if name in wb.sheetnames else None
    print(f"Audit - Reversals {side}: rows~{n} sheets_ok={name in wb.sheetnames}")
    wb.close()

# Ordered: remaining test_only after noise (no double-count)
lh, lr = sheet(folder / "ordered__live.xlsx", "Full Data")
th, tr = sheet(folder / "ordered__test.xlsx", "Full Data")
lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")
lln, tln = idx(lh, "LineNumber"), idx(th, "LineNumber")
li, ti = idx(lh, "Item#", "Item Number"), idx(th, "Item#", "Item Number")
ld, td = idx(lh, "OrderDate"), idx(th, "OrderDate")


def norm_line(v):
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s


def keymap(data, iso, iln, iitem):
    m = {}
    for r in data:
        so = str(r[iso] or "").strip()
        if not so or so.upper() == "TOTAL":
            continue
        m[(so, norm_line(r[iln]), str(r[iitem] or "").strip())] = r
    return m

L = keymap(lr, lso, lln, li)
T = keymap(tr, tso, tln, ti)
only_l, only_t = set(L) - set(T), set(T) - set(L)

def is_tz_live(k):
    return dnorm(L[k][ld]) == "2026-07-01"

def is_tz_test(k):
    return dnorm(T[k][td]) == "2026-07-31"

def is_frac(k):
    return "." in str(k[1])

def is_line0(k):
    return k[1] in ("0", "0.0")

remain_l = [k for k in only_l if not is_tz_live(k) and not is_frac(k)]
remain_t = [k for k in only_t if not is_tz_test(k) and not is_line0(k)]
print(f"ordered remain live_only={len(remain_l)} test_only={len(remain_t)}")
print("remain live dates", Counter(dnorm(L[k][ld]) for k in remain_l))
print("remain test dates sample", Counter(dnorm(T[k][td]) for k in remain_t).most_common(8))
print("remain test SO count", len({k[0] for k in remain_t}))

# InProcess status — treat as Open Order?
print("=== status InProcess ===")
ls, ts = idx(lh, "Status"), idx(th, "Status")
n = 0
for k in set(L) & set(T):
    a, b = str(L[k][ls] or ""), str(T[k][ts] or "")
    if a != b and {a.lower(), b.lower()} == {"inprocess", "open order"}:
        n += 1
        if n <= 3:
            print(" ", k, a, "->", b)
print("InProcess↔Open Order count", n)
