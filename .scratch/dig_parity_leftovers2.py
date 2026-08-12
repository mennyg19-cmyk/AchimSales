from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

folder = Path(".scratch/parity/20260805-111000-po-audit-retest")

def sheet(path, name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(rows)]
    data = list(rows)
    wb.close()
    return hdr, data

def idx(hdr, *names):
    for n in names:
        if n in hdr: return hdr.index(n)
    return None

def dnorm(v):
    if v is None: return ""
    if hasattr(v, "date"): return v.date().isoformat()
    return str(v).strip()[:10]

def norm_line(v):
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"): s = s[:-2]
    try:
        f = float(s)
        if f == int(f): return str(int(f))
    except ValueError:
        pass
    return s

lh, lr = sheet(folder / "ordered__live.xlsx", "Full Data")
th, tr = sheet(folder / "ordered__test.xlsx", "Full Data")
lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")
lln, tln = idx(lh, "LineNumber"), idx(th, "LineNumber")
li, ti = idx(lh, "Item#", "Item Number"), idx(th, "Item#", "Item Number")
ld, td = idx(lh, "OrderDate"), idx(th, "OrderDate")
ls, ts = idx(lh, "Status"), idx(th, "Status")

def keymap(data, iso, iln, iitem):
    m = {}
    for r in data:
        so = str(r[iso] or "").strip()
        if not so or so.upper() == "TOTAL": continue
        m[(so, norm_line(r[iln]), str(r[iitem] or "").strip())] = r
    return m

L, T = keymap(lr, lso, lln, li), keymap(tr, tso, tln, ti)
only_l, only_t = set(L) - set(T), set(T) - set(L)
remain_l = [k for k in only_l if dnorm(L[k][ld]) != "2026-07-01" and "." not in str(k[1])]
remain_t = [k for k in only_t if dnorm(T[k][td]) != "2026-07-31" and k[1] not in ("0", "0.0")]
print("remain live_only", len(remain_l), dict(Counter(dnorm(L[k][ld]) for k in remain_l)))
print("remain test_only", len(remain_t), "SOs", len({k[0] for k in remain_t}))
print("remain test dates", Counter(dnorm(T[k][td]) for k in remain_t).most_common(6))
n = 0
for k in set(L) & set(T):
    a, b = str(L[k][ls] or ""), str(T[k][ts] or "")
    if a.replace(" ", "").lower() == "inprocess" and b.lower() == "open order":
        n += 1
print("InProcess -> Open Order", n)
