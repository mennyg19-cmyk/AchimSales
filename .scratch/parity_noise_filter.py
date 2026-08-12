"""Apply locked noise filters to a parity run folder (ordered + invoiced).

Noise / intentional (do not treat as fail):
Ordered:
  - TZ edges: OrderDate on period first day (live_only) or last day (test_only)
  - Status spelling: Cancelled↔Canceled; Open/In Process ↔ Open Order
  - Qty: LIVE QtyReleased+QtyShipped == TEST QtyReleased (QTY Shipping)
  - Line identity: fractional LineNum live_only / LineNum 0 test_only (same SO)
  - Known coverage: late lines / HeadersV3 gaps (reported separately, not hard fail)

Invoiced:
  - TZ: invoice date == today (or period edge days) live_only
  - SO# one side blank (cosmetic)
  - Audit - Reversals: expect present after SP fix; missing is a real gap

PO: now expected to match (no longer blank-on-test noise).
"""
from __future__ import annotations

import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

TODAY = date.today().isoformat()


def headers_rows(path: Path, sheet: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None, None
    rows = wb[sheet].iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(rows)]
    data = list(rows)
    wb.close()
    return hdr, data


def idx(hdr, *names):
    lower = {h.lower(): i for i, h in enumerate(hdr)}
    for n in names:
        if n in hdr:
            return hdr.index(n)
        if n.lower() in lower:
            return lower[n.lower()]
    return None


def num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def dnorm(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v).strip()
    return "" if not s or s.upper() == "N/A" else s[:10]


def norm_line(v):
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s


def status_equiv(a: str, b: str) -> bool:
    a, b = a.strip(), b.strip()
    if a == b:
        return True
    pair = {a.lower(), b.lower()}
    if pair <= {"cancelled", "canceled"}:
        return True
    opens = {"open", "in process", "open order"}
    if pair <= opens:
        return True
    return False


def analyze_ordered(folder: Path) -> list[str]:
    out = ["## Ordered (last_month)"]
    live = folder / "ordered__live.xlsx"
    test = folder / "ordered__test.xlsx"
    if not live.exists() or not test.exists():
        out.append("missing xlsx")
        return out

    # Full Data
    lh, lr = headers_rows(live, "Full Data")
    th, tr = headers_rows(test, "Full Data")
    lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")
    lln, tln = idx(lh, "LineNumber"), idx(th, "LineNumber")
    li, ti = idx(lh, "Item#", "Item Number"), idx(th, "Item#", "Item Number")
    ld, td = idx(lh, "OrderDate"), idx(th, "OrderDate")
    lrel, trel = idx(lh, "QtyReleased"), idx(th, "QtyReleased", "QTY Shipping")
    lship = idx(lh, "QtyShipped")
    ls, ts = idx(lh, "Status"), idx(th, "Status")
    lpo = idx(lh, "PO #", "PO#")
    tpo = idx(th, "PO #", "PO#")

    def keymap(hdr, data, iso, iln, iitem):
        m = {}
        for r in data:
            so = str(r[iso] or "").strip()
            if not so or so.upper() == "TOTAL":
                continue
            key = (so, norm_line(r[iln]), str(r[iitem] or "").strip())
            m[key] = r
        return m

    L = keymap(lh, lr, lso, lln, li)
    T = keymap(th, tr, tso, tln, ti)
    common = set(L) & set(T)
    only_l, only_t = set(L) - set(T), set(T) - set(L)

    # Coverage with TZ filter (July last_month: first=07-01 last=07-31 — detect from data)
    dates_l = [dnorm(L[k][ld]) for k in only_l]
    dates_t = [dnorm(T[k][td]) for k in only_t]
    # Infer period edges from majority
    all_dates = [dnorm(L[k][ld]) for k in L] + [dnorm(T[k][td]) for k in T]
    months = [d[:7] for d in all_dates if len(d) >= 7]
    month = Counter(months).most_common(1)[0][0] if months else TODAY[:7]
    edge_start, edge_end = f"{month}-01", f"{month}-31"
    # fix end day for month length
    y, m = int(month[:4]), int(month[5:7])
    if m == 12:
        edge_end = f"{y}-12-31"
    else:
        from calendar import monthrange
        edge_end = f"{y}-{m:02d}-{monthrange(y, m)[1]:02d}"

    live_tz = sum(1 for d in dates_l if d == edge_start)
    test_tz = sum(1 for d in dates_t if d == edge_end)
    # Fractional live_only
    frac = sum(1 for k in only_l if "." in str(k[1]) and str(k[1]) not in ("",))
    line0 = sum(1 for k in only_t if k[1] in ("0", "0.0"))

    qty_ok = qty_bad = 0
    status_noise = status_real = 0
    status_pairs = Counter()
    for k in common:
        lr_, tr_ = L[k], T[k]
        if abs((num(lr_[lrel]) + (num(lr_[lship]) if lship is not None else 0)) - num(tr_[trel])) < 1e-6:
            qty_ok += 1
        else:
            qty_bad += 1
        a, b = str(lr_[ls] or ""), str(tr_[ts] or "")
        if a != b:
            if status_equiv(a, b):
                status_noise += 1
            else:
                status_real += 1
                status_pairs[(a, b)] += 1

    # By Order PO
    lh2, lr2 = headers_rows(live, "By Order")
    th2, tr2 = headers_rows(test, "By Order")
    lso2, tso2 = idx(lh2, "SalesOrderNumber"), idx(th2, "SalesOrderNumber")
    lpo2, tpo2 = idx(lh2, "PO #", "PO#"), idx(th2, "PO #", "PO#")
    Lo = {str(r[lso2] or "").strip(): r for r in lr2 if r[lso2]}
    To = {str(r[tso2] or "").strip(): r for r in tr2 if r[tso2]}
    both = set(Lo) & set(To)
    po_match = po_live_only = po_test_only = po_both_blank = po_diff = 0
    for so in both:
        pv = str(Lo[so][lpo2] or "").strip() if lpo2 is not None else ""
        tw = str(To[so][tpo2] or "").strip() if tpo2 is not None else ""
        if not pv and not tw:
            po_both_blank += 1
        elif pv and not tw:
            po_live_only += 1
        elif tw and not pv:
            po_test_only += 1
        elif pv == tw:
            po_match += 1
        else:
            po_diff += 1

    out += [
        f"Full Data: live={len(L)} test={len(T)} common={len(common)} "
        f"live_only={len(only_l)} test_only={len(only_t)}",
        f"  live_only on {edge_start} (TZ noise): {live_tz}",
        f"  test_only on {edge_end} (TZ noise): {test_tz}",
        f"  live_only fractional LineNum: {frac}",
        f"  test_only LineNum 0: {line0}",
        f"  remaining coverage after those cuts: "
        f"live_only={len(only_l)-live_tz-frac} test_only={len(only_t)-test_tz-line0}",
        f"Qty LIVE(rel+ship)==TEST(rel): {qty_ok}/{len(common)} (bad {qty_bad})",
        f"Status: noise(spelling/label)={status_noise} real={status_real}",
        f"  top real status pairs: {status_pairs.most_common(8)}",
        f"By Order PO on {len(both)} shared SOs: match={po_match} "
        f"live_has_test_blank={po_live_only} test_has_live_blank={po_test_only} "
        f"both_blank={po_both_blank} both_filled_diff={po_diff}",
        f"  PO fill rate test (shared): "
        f"{100*(po_match+po_test_only+po_diff)/max(1,len(both)):.1f}% "
        f"(was ~0% when stubbed)",
    ]
    return out


def analyze_invoiced(folder: Path) -> list[str]:
    out = ["## Invoiced (YTD)"]
    live = folder / "invoiced__live.xlsx"
    test = folder / "invoiced__test.xlsx"
    if not live.exists() or not test.exists():
        out.append("missing xlsx")
        return out

    wb_l = load_workbook(live, read_only=True, data_only=True)
    wb_t = load_workbook(test, read_only=True, data_only=True)
    sheets_l, sheets_t = set(wb_l.sheetnames), set(wb_t.sheetnames)
    wb_l.close()
    wb_t.close()
    out.append(f"Sheets live-only: {sorted(sheets_l - sheets_t)}")
    out.append(f"Sheets test-only: {sorted(sheets_t - sheets_l)}")
    audit_ok = "Audit - Reversals" in sheets_t
    out.append(f"Audit - Reversals on /test: {'YES' if audit_ok else 'MISSING'}")

    lh, lr = headers_rows(live, "Full Details")
    th, tr = headers_rows(test, "Full Details")
    if not lh:
        out.append("No Full Details sheet")
        return out

    li, ti = idx(lh, "InvoiceNumber"), idx(th, "InvoiceNumber")
    ld, td = idx(lh, "InvoiceDate"), idx(th, "InvoiceDate")
    ltot, ttot = idx(lh, "Total Invoice"), idx(th, "Total Invoice")
    ltar, ttar = idx(lh, "Tariff Charges"), idx(th, "Tariff Charges")
    lcc, tcc = idx(lh, "CC Charges"), idx(th, "CC Charges")
    lfr, tfr = idx(lh, "Freight Charges"), idx(th, "Freight Charges")
    lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")

    L = {str(r[li] or "").strip(): r for r in lr if r[li]}
    T = {str(r[ti] or "").strip(): r for r in tr if r[ti]}
    common = set(L) & set(T)
    only_l, only_t = set(L) - set(T), set(T) - set(L)
    live_today = sum(1 for i in only_l if dnorm(L[i][ld]) == TODAY)
    live_other = len(only_l) - live_today

    money = Counter()
    examples = defaultdict(list)
    so_blank = 0
    for inv in common:
        a, b = L[inv], T[inv]
        for name, ia, ib in (
            ("total", ltot, ttot), ("tariff", ltar, ttar),
            ("cc", lcc, tcc), ("freight", lfr, tfr),
        ):
            if ia is None or ib is None:
                continue
            if abs(num(a[ia]) - num(b[ib])) > 0.009:
                money[name] += 1
                if len(examples[name]) < 5:
                    examples[name].append((inv, dnorm(a[ld]), num(a[ia]), num(b[ib])))
        sa = str(a[lso] or "").strip() if lso is not None else ""
        sb = str(b[tso] or "").strip() if tso is not None else ""
        if (not sa and sb) or (sa and not sb):
            so_blank += 1

    out += [
        f"Full Details: live={len(L)} test={len(T)} common={len(common)} "
        f"live_only={len(only_l)} test_only={len(only_t)}",
        f"  live_only dated today ({TODAY}) TZ noise: {live_today}",
        f"  live_only other (real?): {live_other}",
        f"  SO one-side blank (cosmetic): {so_blank}",
        f"Money diffs on common: {dict(money)}",
    ]
    for k, xs in examples.items():
        out.append(f"  sample {k}: {xs}")
    return out


def main():
    folder = Path(sys.argv[1] if len(sys.argv) > 1 else ".")
    lines = [f"# Noise-filtered parity — {folder}", f"Today={TODAY}", ""]
    lines += analyze_invoiced(folder)
    lines.append("")
    lines += analyze_ordered(folder)
    text = "\n".join(lines) + "\n"
    out = folder / "NOISE_FILTERED_VERDICT.md"
    out.write_text(text, encoding="utf-8")
    print(text)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
