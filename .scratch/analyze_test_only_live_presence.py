"""For test_only coverage (ignore live_only fractional), are they on LIVE outside July?"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PARITY = Path(".scratch/parity/20260804-193031-postfix")
PROBE = PARITY / "ordered_line_date_probe"


def load_full(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = wb["Full Data"].iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(rows)]
    data = list(rows)
    wb.close()
    return hdr, data


def idx(hdr, *names):
    for n in names:
        if n in hdr:
            return hdr.index(n)
    raise KeyError(names)


def main():
    enriched = pd.read_csv(PROBE / "odata_header_vs_line_dates_enriched.csv")
    test = enriched[enriched["side"] == "test_only"].copy()
    test = test[test["sales_order"].astype(str).str.upper() != "TOTAL"]
    print(f"test_only rows (excl Total): {len(test)}")
    print(test.groupby("bucket").size().sort_values(ascending=False).to_string())
    print()

    # LIVE Full Data: which SOs appear at all in the July LIVE export?
    lhdr, ldata = load_full(PARITY / "ordered__live.xlsx")
    iso = idx(lhdr, "SalesOrderNumber")
    iln = idx(lhdr, "LineNumber")
    iitem = idx(lhdr, "Item#", "Item Number")
    idate = idx(lhdr, "OrderDate")
    live_sos = set()
    live_keys = set()
    live_so_dates: dict[str, set] = {}
    for r in ldata:
        so = str(r[iso] or "").strip()
        if not so or so.upper() == "TOTAL":
            continue
        live_sos.add(so)
        ln = str(r[iln] or "").strip()
        item = str(r[iitem] or "").strip()
        live_keys.add((so, ln, item))
        d = str(r[idate] or "")[:10]
        live_so_dates.setdefault(so, set()).add(d)

    # Per test_only SO: is that SO on LIVE July at all (other lines)?
    by_bucket = Counter()
    so_summary = []
    for so, g in test.groupby("sales_order"):
        buckets = sorted(g["bucket"].unique())
        primary = buckets[0] if len(buckets) == 1 else "+".join(buckets)
        n = len(g)
        header = str(g["header_created"].iloc[0] or "")
        header_in = str(g["header_in_july"].iloc[0])
        line_days = sorted({str(x)[:10] for x in g["line_sys_created"].dropna().astype(str) if str(x) not in ("", "nan")})
        on_live_july = so in live_sos
        live_dates = sorted(live_so_dates.get(so, []))
        # Answer classes
        if primary == "explained_line_in_header_out" or (
            header and header_in == "no" and any(
                (str(x)[:10] >= "2026-07-01" and str(x)[:10] <= "2026-07-31")
                for x in line_days
            )
        ):
            cls = "header_outside_july_line_in_july__LIVE_would_have_under_header_month"
        elif primary == "test_only_missing_from_live_odata_headers":
            cls = "not_in_LIVE_HeadersV3_at_all"
        elif primary == "tz_edge":
            cls = "tz_edge_Aug1_UTC__likely_LIVE_Aug_or_Jul_depending_TZ"
        elif primary == "no_line_created_odata":
            cls = "no_CDR_line_match_LineNum0"
        else:
            cls = f"other:{primary}"

        if on_live_july:
            cls = cls + "__BUT_SO_HAS_OTHER_LINES_ON_LIVE_JULY"

        by_bucket[cls] += n
        so_summary.append({
            "sales_order": so,
            "test_only_lines": n,
            "bucket": primary,
            "header_created": header,
            "header_in_july": header_in,
            "line_created_days": ",".join(line_days[:5]),
            "so_on_live_july_export": "yes" if on_live_july else "no",
            "live_july_order_dates": ",".join(live_dates[:5]),
            "answer_class": cls,
        })

    print("=== test_only lines by answer class ===")
    for k, n in sorted(by_bucket.items(), key=lambda x: -x[1]):
        print(f"  {n:4d}  {k}")

    print("\n=== SO-level counts ===")
    sdf = pd.DataFrame(so_summary)
    print(sdf.groupby("answer_class").agg(
        sos=("sales_order", "count"),
        lines=("test_only_lines", "sum"),
    ).sort_values("lines", ascending=False).to_string())

    # How many test_only SOs appear on LIVE July with *other* lines?
    overlap = sdf[sdf.so_on_live_july_export == "yes"]
    print(f"\nSOs with test_only lines that ALSO appear on LIVE July (other lines): "
          f"{len(overlap)} SOs / {overlap.test_only_lines.sum()} test_only lines")
    if len(overlap):
        print(overlap[["sales_order", "test_only_lines", "bucket", "header_created", "line_created_days"]].head(20).to_string())

    out = PROBE / "test_only_live_presence.csv"
    sdf.sort_values(["answer_class", "test_only_lines"], ascending=[True, False]).to_csv(out, index=False)
    print(f"\nWrote {out}")

    # Direct answer rollup
    in_live_outside = by_bucket.get(
        "header_outside_july_line_in_july__LIVE_would_have_under_header_month", 0)
    not_in_live = by_bucket.get("not_in_LIVE_HeadersV3_at_all", 0)
    tz = by_bucket.get("tz_edge_Aug1_UTC__likely_LIVE_Aug_or_Jul_depending_TZ", 0)
    other = len(test) - in_live_outside - not_in_live - tz
    print("\n=== DIRECT ANSWER ===")
    print(f"After dropping live_only LineNum diffs, remaining = {len(test)} test_only lines.")
    print(f"1) On LIVE under a different month (header outside July, line in July): {in_live_outside}")
    print(f"2) Not visible to LIVE OData HeadersV3 at all: {not_in_live}")
    print(f"3) TZ edge (OData Aug 1 / report Jul 31): {tz}")
    print(f"4) Other (no CDR line 0, etc.): {other}")


if __name__ == "__main__":
    main()
