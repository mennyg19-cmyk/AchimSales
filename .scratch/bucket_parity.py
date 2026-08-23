"""Bucket invoiced + ordered postfix diffs for breakdown."""
from collections import Counter, defaultdict
from openpyxl import load_workbook
from pathlib import Path
from datetime import date

p = Path(".scratch/parity/20260804-193031-postfix")
TODAY = date.today().isoformat()  # 2026-08-04


def headers_and_rows(path, sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    # skip title rows for live CA-style; for invoiced/ordered first row is usually header
    first = next(rows)
    hdr = [str(c or "").strip() for c in first]
    # if not enough header-like cells, scan
    if sum(1 for h in hdr if h) < 3:
        for r in rows:
            cand = [str(c or "").strip() for c in r]
            if sum(1 for h in cand if h) >= 3 and any("Invoice" in h or "Qty" in h or "Customer" in h for h in cand):
                hdr = cand
                break
        data = list(rows)
    else:
        data = list(rows)
    wb.close()
    return hdr, data


def idx(hdr, *names):
    for n in names:
        if n in hdr:
            return hdr.index(n)
    return None


def num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def dnorm(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v).strip()
    if not s or s.upper() == "N/A":
        return ""
    return s[:10]


print("=== INVOICED Full Details ===")
lh, lr = headers_and_rows(p / "invoiced__live.xlsx", "Full Details")
th, tr = headers_and_rows(p / "invoiced__test.xlsx", "Full Details")
li, ti = idx(lh, "InvoiceNumber"), idx(th, "InvoiceNumber")
ld, td = idx(lh, "InvoiceDate"), idx(th, "InvoiceDate")
ltot, ttot = idx(lh, "Total Invoice"), idx(th, "Total Invoice")
ltar, ttar = idx(lh, "Tariff Charges"), idx(th, "Tariff Charges")
lcc, tcc = idx(lh, "CC Charges"), idx(th, "CC Charges")
lfr, tfr = idx(lh, "Freight Charges"), idx(th, "Freight Charges")
lsub, tsub = idx(lh, "SubTotal Invoices"), idx(th, "SubTotal Invoices")
lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")
lacc, tacc = idx(lh, "CustomerAccount"), idx(th, "CustomerAccount")

L = {}
for r in lr:
    inv = str(r[li] or "").strip()
    if inv:
        L[inv] = r
T = {}
for r in tr:
    inv = str(r[ti] or "").strip()
    if inv:
        T[inv] = r

only_live = set(L) - set(T)
only_test = set(T) - set(L)
common = set(L) & set(T)
print(f"rows live={len(L)} test={len(T)} common={len(common)} live_only={len(only_live)} test_only={len(only_test)}")

live_only_dates = Counter(dnorm(L[i][ld]) for i in only_live)
print("live_only date breakdown:", dict(live_only_dates.most_common(5)))

# value diffs among common
buckets = Counter()
examples = defaultdict(list)
money_fields = [
    ("total", ltot, ttot),
    ("subtotal", lsub, tsub),
    ("tariff", ltar, ttar),
    ("cc", lcc, tcc),
    ("freight", lfr, tfr),
]
for inv in common:
    lr_, tr_ = L[inv], T[inv]
    so_l = str(lr_[lso] or "").strip() if lso is not None else ""
    so_t = str(tr_[tso] or "").strip() if tso is not None else ""
    money_diff = False
    for name, li_, ti_ in money_fields:
        if abs(num(lr_[li_]) - num(tr_[ti_])) > 0.009:
            buckets[f"money_{name}"] += 1
            money_diff = True
            if len(examples[f"money_{name}"]) < 5:
                examples[f"money_{name}"].append(
                    (inv, dnorm(lr_[ld]), num(lr_[li_]), num(tr_[ti_]), str(lr_[lacc]))
                )
    if so_l != so_t:
        # cosmetic if one blank
        if (not so_l and so_t) or (so_l and not so_t):
            buckets["so_one_blank"] += 1
        else:
            buckets["so_both_filled_diff"] += 1
    if money_diff:
        buckets["any_money"] += 1

print("common value buckets:", dict(buckets))
for k, xs in examples.items():
    print(k, xs)

# Credits sheet reversal-ish: negative totals count
print("\n=== INVOICED Credits sample extents ===")
# already know audit missing because no mixed +/-

print("\n=== ORDERED Full Data ===")
lh, lr = headers_and_rows(p / "ordered__live.xlsx", "Full Data")
th, tr = headers_and_rows(p / "ordered__test.xlsx", "Full Data")
print("live qty cols", [h for h in lh if "Qty" in h or "Status" in h or "PO" in h])
print("test qty cols", [h for h in th if "Qty" in h or "Status" in h or "PO" in h])

lso = idx(lh, "SalesOrderNumber")
lln = idx(lh, "LineNumber")
litem = idx(lh, "Item#", "Item Number")
tso = idx(th, "SalesOrderNumber")
tln = idx(th, "LineNumber")
titem = idx(th, "Item#", "Item Number")
lrel, trel = idx(lh, "QtyReleased"), idx(th, "QtyReleased", "QTY Shipping")
lship = idx(lh, "QtyShipped")
lstat, tstat = idx(lh, "Status"), idx(th, "Status")
lpo = idx(lh, "PO #", "PO#", "CustomerRequisition")
# PO may be missing from Full Data on live - check
print("live has PO?", lpo is not None, "test has PO?", idx(th, "PO #", "PO#") is not None)
ldate, tdate = idx(lh, "OrderDate"), idx(th, "OrderDate")

L = {}
for r in lr:
    key = (str(r[lso] or ""), str(r[lln] or ""), str(r[litem] or ""))
    L[key] = r
T = {}
for r in tr:
    key = (str(r[tso] or ""), str(r[tln] or ""), str(r[titem] or ""))
    T[key] = r

only_live = set(L) - set(T)
only_test = set(T) - set(L)
common = set(L) & set(T)
print(f"lines live={len(L)} test={len(T)} common={len(common)} live_only={len(only_live)} test_only={len(only_test)}")

live_only_dates = Counter(dnorm(L[k][ldate]) for k in only_live)
test_only_dates = Counter(dnorm(T[k][tdate]) for k in only_test)
print("live_only dates", dict(live_only_dates.most_common(5)))
print("test_only dates", dict(test_only_dates.most_common(5)))

# among common: released sum match, status, etc
rel_sum_match = rel_only_match = status_diff = 0
status_pairs = Counter()
for k in common:
    lr_, tr_ = L[k], T[k]
    l_rel = num(lr_[lrel]) if lrel is not None else 0
    l_ship = num(lr_[lship]) if lship is not None else 0
    t_rel = num(tr_[trel]) if trel is not None else 0
    if abs((l_rel + l_ship) - t_rel) < 1e-6:
        rel_sum_match += 1
    if abs(l_rel - t_rel) < 1e-6:
        rel_only_match += 1
    ls = str(lr_[lstat] or "").strip() if lstat is not None else ""
    ts = str(tr_[tstat] or "").strip() if tstat is not None else ""
    if ls != ts:
        status_diff += 1
        status_pairs[(ls, ts)] += 1

print(f"common released+shipped==test released: {rel_sum_match}/{len(common)}")
print(f"common live released==test released: {rel_only_match}/{len(common)}")
print(f"status diffs: {status_diff}")
print("top status pairs:", status_pairs.most_common(15))

# By Order coverage / PO
print("\n=== ORDERED By Order ===")
lh, lr = headers_and_rows(p / "ordered__live.xlsx", "By Order")
th, tr = headers_and_rows(p / "ordered__test.xlsx", "By Order")
lso, tso = idx(lh, "SalesOrderNumber"), idx(th, "SalesOrderNumber")
lpo, tpo = idx(lh, "PO #", "PO#"), idx(th, "PO #", "PO#")
ldate, tdate = idx(lh, "OrderDate"), idx(th, "OrderDate")
L = {str(r[lso] or ""): r for r in lr if r[lso]}
T = {str(r[tso] or ""): r for r in tr if r[tso]}
only_live = set(L) - set(T)
only_test = set(T) - set(L)
common = set(L) & set(T)
print(f"orders live={len(L)} test={len(T)} common={len(common)} live_only={len(only_live)} test_only={len(only_test)}")
print("live_only dates", dict(Counter(dnorm(L[k][ldate]) for k in only_live).most_common(5)))
print("test_only dates", dict(Counter(dnorm(T[k][tdate]) for k in only_test).most_common(5)))
po_blank_test = sum(1 for k in common if not str(T[k][tpo] or "").strip()) if tpo is not None else None
po_blank_live = sum(1 for k in common if not str(L[k][lpo] or "").strip()) if lpo is not None else None
po_live_has_test_blank = sum(
    1 for k in common
    if lpo is not None and tpo is not None
    and str(L[k][lpo] or "").strip() and not str(T[k][tpo] or "").strip()
)
print(f"common PO blank live={po_blank_live} test={po_blank_test} live_has_test_blank={po_live_has_test_blank}")
