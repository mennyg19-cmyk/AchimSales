"""
Read salesman mappings from the editable Excel file (salesman_map.xlsx).

The Excel file has columns:
  Key, Number, FullName, DisplayName, Email,
  Recv_Ordered, Recv_Invoiced, Recv_Salesman, Recv_Number4, Recv_CustomerActivity,
  Recv_MasterSalesman, Recv_MasterCustomerActivity,
  CC, BCC, Commission %

This module is the primary data source for salesman lookups.
salesman_map.py delegates here and falls back to its hardcoded dict
if the Excel file is missing.
"""

import logging
import os
import re
from functools import lru_cache
from typing import NamedTuple

from openpyxl import load_workbook

log = logging.getLogger(__name__)

_XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "salesman_map.xlsx")

REPORT_KEYS = ("ordered", "invoiced", "salesman", "number_4", "customer_activity",
                "master_salesman", "master_customer_activity",
                "customer_aging_report")

_SUBSCRIPTION_COLUMNS = {
    "Recv_Ordered": "ordered",
    "Recv_Invoiced": "invoiced",
    "Recv_Salesman": "salesman",
    "Recv_Number4": "number_4",
    "Recv_CustomerActivity": "customer_activity",
    "Recv_MasterSalesman": "master_salesman",
    "Recv_MasterCustomerActivity": "master_customer_activity",
    "Recv_CustomerAging": "customer_aging_report",
}


class SalesmanRecord(NamedTuple):
    number: str
    full_name: str
    display_name: str
    email: str
    subscriptions: dict[str, bool]
    cc: list[str] = []
    bcc: list[str] = []
    commission_pct: float = 0.0


DEFAULT_SALESMAN = SalesmanRecord(
    number="?unassigned",
    full_name="Unassigned",
    display_name="Unassigned",
    email="",
    subscriptions={k: False for k in REPORT_KEYS},
    cc=[],
    bcc=[],
    commission_pct=0.0,
)


def _norm_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower()) if s else ""


def _to_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "1", "yes", "y")


def _parse_email_list(val) -> list[str]:
    """Split a semicolon-separated email string into a clean list."""
    if not val:
        return []
    raw = str(val).strip()
    if not raw:
        return []
    return [addr.strip() for addr in raw.split(";") if addr.strip()]


@lru_cache(maxsize=1)
def load_salesman_map(path: str | None = None) -> dict[str, SalesmanRecord]:
    """Load the salesman map from the Excel file.

    Returns {normalized_key: SalesmanRecord}.
    """
    xlsx = path or _XLSX_PATH
    if not os.path.isfile(xlsx):
        log.warning("Salesman map Excel not found: %s", xlsx)
        return {}

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active

    header_map: dict[str, int] = {}
    result: dict[str, SalesmanRecord] = {}
    first_row = True

    for row in ws.iter_rows(values_only=False):
        if first_row:
            for cell in row:
                if cell.value:
                    header_map[str(cell.value).strip()] = cell.column - 1
            first_row = False
            continue

        vals = [cell.value for cell in row]
        if not vals or not vals[0]:
            continue

        key = _norm_key(str(vals[0]))
        number = str(vals[1] or "").strip() if len(vals) > 1 else ""
        full_name = str(vals[2] or "").strip() if len(vals) > 2 else ""
        display_name = str(vals[3] or "").strip() if len(vals) > 3 else ""
        email = str(vals[4] or "").strip() if len(vals) > 4 else ""

        subs: dict[str, bool] = {}
        for col_name, report_key in _SUBSCRIPTION_COLUMNS.items():
            col_idx = header_map.get(col_name)
            if col_idx is not None and col_idx < len(vals):
                subs[report_key] = _to_bool(vals[col_idx])
            else:
                subs[report_key] = bool(email)

        cc_idx = header_map.get("CC")
        cc = _parse_email_list(vals[cc_idx] if cc_idx is not None and cc_idx < len(vals) else None)
        bcc_idx = header_map.get("BCC")
        bcc = _parse_email_list(vals[bcc_idx] if bcc_idx is not None and bcc_idx < len(vals) else None)

        comm_idx = header_map.get("Commission %")
        comm_pct = 0.0
        if comm_idx is not None and comm_idx < len(vals) and vals[comm_idx] is not None:
            try:
                comm_pct = float(vals[comm_idx])
            except (TypeError, ValueError):
                pass

        result[key] = SalesmanRecord(
            number=number,
            full_name=full_name,
            display_name=display_name,
            email=email,
            subscriptions=subs,
            cc=cc,
            bcc=bcc,
            commission_pct=comm_pct,
        )

    wb.close()
    log.info("Loaded %d salesmen from %s", len(result), xlsx)
    return result


def lookup_salesman_xl(sales_group: str) -> SalesmanRecord:
    """Look up salesman by raw sales group string."""
    key = _norm_key(sales_group)
    return load_salesman_map().get(key, DEFAULT_SALESMAN)


def get_salesman_email(sales_group: str) -> str:
    return lookup_salesman_xl(sales_group).email


def get_salesman_cc_bcc(sales_group: str) -> tuple[list[str], list[str]]:
    """Return (cc_list, bcc_list) for a salesman."""
    rec = lookup_salesman_xl(sales_group)
    return list(rec.cc), list(rec.bcc)


def get_salesman_display_name_xl(sales_group: str) -> str:
    return lookup_salesman_xl(sales_group).display_name


def get_salesman_full_name_xl(sales_group: str) -> str:
    return lookup_salesman_xl(sales_group).full_name


def get_salesman_number_xl(sales_group: str) -> str:
    return lookup_salesman_xl(sales_group).number


def wants_report(sales_group: str, report_key: str) -> bool:
    """Check if a salesman wants to receive a specific report.

    report_key must be one of: ordered, invoiced, salesman, number_4, customer_activity,
    master_salesman, master_customer_activity
    """
    rec = lookup_salesman_xl(sales_group)
    if not rec.email:
        return False
    return rec.subscriptions.get(report_key, False)


def pad_salesman_number(num: str) -> str:
    """Zero-pad numeric salesman numbers to 3 digits."""
    s = str(num).strip()
    if re.fullmatch(r"\d+", s):
        return s.zfill(3)
    return s


def has_report_subscription_column(report_key: str, path: str | None = None) -> bool:
    """Return whether the salesman spreadsheet explicitly defines this report."""
    column_name = next(
        (name for name, key in _SUBSCRIPTION_COLUMNS.items() if key == report_key),
        None,
    )
    if column_name is None:
        return False

    xlsx = path or _XLSX_PATH
    if not os.path.isfile(xlsx):
        return False

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        header_row = next(
            wb.active.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = {str(value).strip() for value in header_row if value}
        return column_name in headers
    finally:
        wb.close()


def get_report_subscribers(
    report_key: str, path: str | None = None,
) -> list[tuple[str, str, list[str], list[str]]]:
    """Return [(display_name, email, cc, bcc)] for everyone subscribed to a report.

    Useful for master report distribution where the recipient is not
    necessarily a salesman in the data -- just someone who wants the file.
    """
    result = []
    for key, rec in load_salesman_map(path).items():
        if rec.email and rec.subscriptions.get(report_key, False):
            result.append((rec.display_name or key, rec.email, list(rec.cc), list(rec.bcc)))
    return result


def get_commission_pct_map() -> dict[str, float]:
    """Return {normalized_salesman_number: commission_rate} from the Excel file.

    Only includes salesmen with a non-zero commission percentage.
    Keys are salesman numbers with leading zeros stripped (e.g. "12", "90").
    """
    result: dict[str, float] = {}
    for _key, rec in load_salesman_map().items():
        if rec.commission_pct and rec.commission_pct > 0:
            normalized = str(rec.number).strip().lstrip("0") or "0"
            result[normalized] = rec.commission_pct
    return result
